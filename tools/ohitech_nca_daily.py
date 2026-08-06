#!/usr/bin/env python3
"""NCA 보고서들을 **일별 × 판매방식**으로 접어 Retail분만 뽑는다.

★비례배분 없음 — 「판매방식」 컬럼의 실측 그대로다. 달마다 비중이 완전히 달라서
  (2025-07 0% · 2025-08 94.5% · 2025-11 53.7% · 2026-04~06 100%) 배분은 쓸 수 없다.

출력: /tmp/claude-501/nca_retail_by_day.json  = {"YYYY-MM-DD": <Retail NCA 광고비>}
"""
import collections
import glob
import json
import os
import warnings
import zipfile

import openpyxl

warnings.filterwarnings("ignore")

OUT = "/tmp/claude-501/nca_retail_by_day.json"


def main():
    by_day = collections.defaultdict(lambda: collections.Counter())
    seen_files = []
    bad = []
    # 같은 날이 두 파일에 들어 있으면(월 단위 + 주 단위 재수집) 중복 합산이 된다.
    # 파일별로 날짜 집합을 보고, 이미 채운 날은 건너뛴다 — 먼저 읽은 파일이 이긴다.
    claimed: dict[str, str] = {}
    for p in sorted(glob.glob(os.path.expanduser("~/.ohisell_nca/*.xlsx"))):
        if not zipfile.is_zipfile(p):
            bad.append(os.path.basename(p))
            continue
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        hdr = list(next(it))
        i = {h: n for n, h in enumerate(hdr)}
        C, S, D = i["집행 광고비"], i["판매방식"], i["날짜"]
        rows = 0
        for r in it:
            raw = str(r[D]).split(".")[0]
            day = f"{raw[:4]}-{raw[4:6]}-{raw[6:8]}"
            owner = claimed.get(day)
            if owner is None:
                claimed[day] = p
            elif owner != p:
                continue        # 다른 파일이 이미 그 날을 채웠다 — 중복 합산 금지
            by_day[day][r[S]] += float(r[C] or 0)
            rows += 1
        seen_files.append((os.path.basename(p), rows))
    retail = {d: round(v.get("Retail", 0.0)) for d, v in sorted(by_day.items())}
    json.dump(retail, open(OUT, "w"))
    print("깨진 파일:", bad or "없음")
    for f, n in seen_files:
        print(f"  {f:24} 행 {n:,}")
    tot = sum(retail.values())
    allsum = sum(sum(v.values()) for v in by_day.values())
    print(f"\n일수 {len(retail)}  NCA 전체 {allsum:,.0f}  그중 Retail {tot:,}  → {OUT}")
    ms = collections.Counter()
    for d, v in retail.items():
        ms[d[:7]] += v
    for m in sorted(ms):
        print(f"  {m}  Retail NCA {ms[m]:>12,}")


if __name__ == "__main__":
    main()
