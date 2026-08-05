#!/usr/bin/env python3
"""월별 Billboard XLSX → 판매방식=Retail 일별 광고비 집계 + 완결성 검산 + prod 소급 적재 (D-21).

왜 있나: 1P 광고비가 2025-07-24~2026-03-16 8개월 통째로 0이었다(D-20, ref 46). `report/SALES`는
  **계정 총액**이라 3P가 섞인다 — 2025-07은 광고비의 96%가 3P였다. 그래서 Billboard 옵션
  보고서의 「판매방식」으로 Retail만 골라 넣는다(B안).

★검산은 "Retail 비중"이 아니라 **완결성**으로 한다(교훈 #144):
    Billboard 전체(Retail+3P+RG) ≈ report/SALES PA   ← 둘 다 PA 전용이라 정의가 같다
  ref 46 §5의 "Retail이 계정 총액의 95~98%"를 그대로 쓰면 3P→1P 전환월인 2025-07(3.71%)을
  **정상인데 실패로 판정**한다. 비중은 알아내려는 값이지 합격 조건이 아니다.

준비: 같은 디렉터리에 (1) 월별 Billboard XLSX `YYYY-MM.xlsx`
        ← `ohitech_ad_option_report.py YYYYMMDD YYYYMMDD YYYY-MM.xlsx`
      (2) `sales_totals.json` ← `ohitech_ad_sales_totals.py <시작> <끝> sales_totals.json`

사용: ohitech_ad_retail_backfill.py            # 집계·검산만 (출력 days.json)
      ohitech_ad_retail_backfill.py --push     # 검산 통과 시 prod ingest 까지(snapshot upsert=멱등)

⚠️비-PA(ALL−PA)는 Billboard에 없어 이 경로로는 안 들어간다(8개월 4,527,079원=1.17%).
  D-10 머니룰은 ALL 차감이므로 그만큼 덜 차감한다 — ref 46 §5-1④(Jino 결정 대기).
"""
import collections
import glob
import json
import os
import sys
import urllib.request
import warnings

import openpyxl

warnings.filterwarnings("ignore")

SP = os.path.dirname(os.path.abspath(__file__))
GAP_END = "2026-03-16"   # 이후는 수기 XLSX가 이미 채움
GAP_START = "2025-07-01"


def load_month(path):
    """(date -> [retail_cost, retail_conv1d, retail_conv14d]), (판매방식 -> 총액)"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = next(it)
    I = {h: i for i, h in enumerate(hdr)}
    days = collections.defaultdict(lambda: [0.0, 0.0, 0.0])
    by_sm = collections.Counter()
    rows = 0
    for r in it:
        rows += 1
        sm = str(r[I["판매방식"]]).strip()
        cost = float(r[I["광고비"]] or 0)
        by_sm[sm] += cost
        if sm != "Retail":
            continue
        raw = r[I["날짜"]]
        d = str(int(float(raw)))            # 20251001.0 → '20251001'
        day = f"{d[0:4]}-{d[4:6]}-{d[6:8]}"
        days[day][0] += cost
        days[day][1] += float(r[I["총 전환매출액(1일)"]] or 0)
        days[day][2] += float(r[I["총 전환매출액(14일)"]] or 0)
    wb.close()
    return days, by_sm, rows


def main():
    push = "--push" in sys.argv
    sales = json.load(open(f"{SP}/sales_totals.json"))
    all_days = {}
    print(f"{'월':8s} {'행수':>7s} {'Retail':>13s} {'3P':>11s} {'RG':>7s} "
          f"{'SALES(PA)':>13s} {'커버':>7s} {'R비중':>6s} {'SALES(ALL)':>13s}")
    ok = True
    for path in sorted(glob.glob(f"{SP}/2*.xlsx")):
        label = os.path.basename(path)[:-5]
        days, by_sm, rows = load_month(path)
        for day, v in days.items():
            if not (GAP_START <= day <= GAP_END):
                continue
            all_days[day] = v
        retail = by_sm.get("Retail", 0.0)
        p3 = by_sm.get("3P", 0.0)
        rg = sum(v for k, v in by_sm.items() if k not in ("Retail", "3P"))
        s_pa = sum(x["pa"] for d, x in sales.items() if d[:7] == label and GAP_START <= d <= GAP_END)
        s_all = sum(x["all"] for d, x in sales.items() if d[:7] == label and GAP_START <= d <= GAP_END)
        # 월 전체를 받았지만 03월은 03-16까지만 유효 → Retail도 창으로 다시 합산해 비교
        retail_win = sum(v[0] for d, v in days.items() if GAP_START <= d <= GAP_END)
        # ★검산은 "Retail 비중"이 아니라 **완결성**으로 한다:
        #   Billboard 전체(Retail+3P+RG) ≈ report/SALES PA 여야 한다(둘 다 PA 전용).
        #   Retail 비중은 달마다 정당하게 다르다 — 2025-07은 3P 96%(1P 전환 직전)라
        #   "95~98%" 규칙을 쓰면 정상 데이터를 실패로 오판한다.
        bb_total = retail + p3 + rg
        cover = 100 * bb_total / s_pa if s_pa else 0
        r_pa = 100 * retail_win / s_pa if s_pa else 0
        flag = "" if 99.0 <= cover <= 101.0 else "  ⚠️"
        if flag:
            ok = False
        print(f"{label:8s} {rows:7,d} {retail_win:13,.0f} {p3:11,.0f} {rg:7,.0f} "
              f"{s_pa:13,.0f} {cover:6.2f}% {r_pa:6.1f}% {s_all:13,.0f}{flag}")

    tot = sum(v[0] for v in all_days.values())
    s_pa = sum(x["pa"] for d, x in sales.items() if GAP_START <= d <= GAP_END)
    s_all = sum(x["all"] for d, x in sales.items() if GAP_START <= d <= GAP_END)
    print(f"\n{'합계':8s} {len(all_days):3d}일  Retail={tot:,.0f}  "
          f"SALES PA={s_pa:,.0f} ({100*tot/s_pa:.2f}%)  ALL={s_all:,.0f} ({100*tot/s_all:.2f}%)")
    print(f"비PA(=ALL−PA) 잔차 {s_all - s_pa:,.0f}원 — Billboard(PA 전용)에는 없음")

    days_payload = [{"date": d, "ad_spend": round(v[0]), "conv_sales": round(v[1])}
                    for d, v in sorted(all_days.items())]
    with open(f"{SP}/days.json", "w") as f:
        json.dump(days_payload, f, ensure_ascii=False)
    print(f"days.json 저장: {len(days_payload)}일")

    if not push:
        return 0 if ok else 1
    if not ok:
        print("검산 실패 — push 중단", file=sys.stderr)
        return 1
    cfg = json.load(open(os.path.expanduser("~/.ohisell_ohitech_ad.json")))
    url = cfg["prod_base_url"].rstrip("/") + "/api/coupang/ops/rocket/ad-cost/ingest"
    body = {"vendor_id": cfg.get("ad_vendor_code", "A01029796"), "days": days_payload}
    req = urllib.request.Request(url, data=json.dumps(body).encode(),
                                 headers={"Content-Type": "application/json",
                                          "X-Ingest-Token": cfg["ingest_token"]})
    with urllib.request.urlopen(req, timeout=120) as r:
        print("ingest:", r.status, r.read().decode()[:400])
    return 0


if __name__ == "__main__":
    sys.exit(main())
