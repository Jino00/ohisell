"""오하이테크 광고센터 CDP 공용 헬퍼 — 세션 판정 · GraphQL · 대용량 다운로드.

왜 모듈로 빼는가: 아래 두 결함이 **같은 코드를 각자 복사해 둔 탓에** 도구마다 따로 났다.
`ohitech_ad_option_report.py`·`ohitech_nca_report.py`가 이 모듈만 쓰게 해 한 번만 고친다.
(데몬 런타임에는 설치되지 않는 repo 전용 도구다 — `install_local_runtime.sh` 목록 밖.)

## 결함 ① 세션이 죽은 것을 "데이터가 없는 것"으로 읽었다 (2026-08-06 실사고)
탭이 몇 분만 유휴여도 `/user/login`으로 밀려나 있고 그 상태의 `fetch`는
`TypeError: Failed to fetch`로 죽는다. 그런데 옛 `gql()`은 예외를 잡지 않아 CDP가 준 빈 객체를
그대로 받았고, 호출부는 그것을 **"캠페인 0개"**로 출력한 뒤 정상 종료했다 — 9개월치 수집이
조용히 건너뛰어졌고 요약표에는 성공처럼 보였다.
→ 이 모듈은 **fetch throw · 비200 · HTML 본문(로그인 페이지) · GraphQL errors를 전부 예외**로
  올린다. "0건"은 그 관문을 통과한 뒤에만 나올 수 있으므로 **진짜 0건**을 뜻한다.

## 결함 ② 큰 보고서를 "깨졌다"고 오진했다 (2026-08-06 — ★진단이 두 번 바뀌었다)
2025-10 NCA를 한 달 통으로 받으면 72,882,972바이트가 오는데 zip이 아니다(매직 `22eb82a0`).
`openpyxl`이 `BadZipFile`로 죽으니 **다운로드가 깨진 줄 알았다.** 틀렸다.

조각 전송으로 다시 받아 보니 **길이가 정확히 같고**(72,882,972) 내용은 멀쩡했다 —
그 바이트를 UTF-8로 읽으면 `"날짜"\t"과금 방식"\t"판매방식"…`이다. **쿠팡은 행이 많으면
xlsx 대신 탭 구분 TSV를 준다**(이 건은 188,094행). 다운로드는 처음부터 정상이었고,
진짜 결함은 **포맷이 바뀌는데 파서가 xlsx만 가정한 것 + 받은 것을 확인하지 않은 것**이다.
검증: TSV 한 달 통짜 합계 3,679,536원이 주 단위 5개 합계와 **차이 0원**(Retail 3,658,680 /
3P 20,856까지 동일).

→ `sniff_format()`으로 무엇이 왔는지 **판정**하고 `read_report()`가 둘 다 읽는다.
→ `download_binary()`의 조각 전송은 남긴다: 원인은 아니었지만 바이트마다 문자열을 잇는 옛
  방식보다 빠르고, 받은 길이를 서버가 말한 값과 **대조**하므로 진짜로 잘린 경우를 잡는다.
"""
from __future__ import annotations

import json
import time
import urllib.error
import urllib.request

import websocket

PORT = 9224
GQL_URL = "https://advertising.coupang.com/marketing-reporting/v2/graphql"
EXCEL_URL = "https://advertising.coupang.com/marketing-reporting/v2/api/excel-report?id="

# 페이지에 잠깐 두는 다운로드 버퍼. 끝나면 지운다.
_BUF = "__ohisellDL"


class SessionDead(RuntimeError):
    """로그인이 끊겼거나 오리진이 죽었다 — 재무장/재로그인이 필요하다는 뜻."""


class GraphQLError(RuntimeError):
    """서버가 GraphQL errors를 돌려줬다(스키마·권한·인자 문제)."""


def connect(port: int | None = None, *, rearm: bool = True, timeout_s: int = 40):
    """광고센터 탭에 붙고, 기본으로 **재무장**(리로드)까지 하고 돌려준다.

    재무장을 기본값으로 둔 이유: 유휴 탭은 거의 항상 `/user/login`에 있고, 그걸 모르고 부르면
    결함 ①이 그대로 재현된다. 리로드 뒤에도 로그인 페이지면 `SessionDead`를 올린다 —
    호출부가 "0건"으로 오해할 여지를 남기지 않는다.
    """
    # 기본 인자로 박지 않고 **호출 시점에** 읽는다 — 그래야 테스트가 PORT를 바꿔 끼울 수 있다.
    port = PORT if port is None else port
    try:
        ts = json.load(urllib.request.urlopen(f"http://127.0.0.1:{port}/json", timeout=5))
    except (urllib.error.URLError, OSError, ValueError) as ex:
        raise SessionDead(f"CDP {port}에 붙지 못했다: {str(ex)[:120]} — 광고 Chrome이 떠 있는지 볼 것") from ex
    pages = [t for t in ts if t.get("type") == "page"
             and "advertising.coupang.com" in t.get("url", "")]
    if not pages:
        raise SessionDead(f"CDP {port}에 advertising.coupang.com 탭이 없다 — 광고 Chrome을 띄울 것")
    ws = websocket.create_connection(pages[0]["webSocketDebuggerUrl"],
                                     suppress_origin=True, max_size=None)
    mid = [0]

    def call(method, params=None):
        mid[0] += 1
        ws.send(json.dumps({"id": mid[0], "method": method, "params": params or {}}))
        while True:
            msg = json.loads(ws.recv())
            if msg.get("id") == mid[0]:
                return msg

    call.close = ws.close  # type: ignore[attr-defined]
    call("Runtime.enable")
    call("Page.enable")
    if rearm:
        call("Page.reload", {"ignoreCache": True})
        deadline = time.monotonic() + timeout_s
        href = ""
        while time.monotonic() < deadline:
            time.sleep(2)
            href = evaluate(call, "location.href") or ""
            if href and "/user/login" not in href:
                return call
        ws.close()
        raise SessionDead(
            f"재무장 후에도 로그인 페이지({href[:80]}) — 페처 갱신으로 자가복구시킬 것: "
            "POST /api/coupang/ops/rocket/ad-cost/request-refresh")
    return call


def evaluate(call, expression: str, *, await_promise: bool = False):
    """Runtime.evaluate 얇은 래퍼. 값이 안 오면 예외 — 조용히 None을 흘리지 않는다."""
    r = call("Runtime.evaluate", {"expression": expression,
                                  "awaitPromise": await_promise, "returnByValue": True})
    res = r.get("result", {}).get("result", {})
    if "value" not in res:
        raise SessionDead("CDP가 값을 주지 않았다(오리진 죽음 의심): " + json.dumps(r)[:300])
    return res["value"]


_FETCH_TMPL = """(async(p)=>{try{
  const r = await fetch(%s, {method:'POST',
    headers:{'content-type':'application/json'},
    body: JSON.stringify(p), credentials:'include'});
  return JSON.stringify({s:r.status, b: await r.text()});
}catch(e){ return JSON.stringify({err:String(e)}); }})(%s)"""


def gql(call, body: dict, *, url: str = GQL_URL) -> dict:
    """GraphQL POST. **실패는 전부 예외로 올린다**(결함 ①의 재발 방지선).

    빈 응답·로그인 HTML·비200·GraphQL errors 중 무엇도 조용히 통과하지 않는다.
    """
    raw = evaluate(call, _FETCH_TMPL % (json.dumps(url), json.dumps(body)), await_promise=True)
    d = raw if isinstance(raw, dict) else json.loads(raw)
    if "err" in d:
        raise SessionDead(f"fetch가 죽었다: {d['err'][:160]} — 오리진 재무장 필요")
    if d.get("s") != 200:
        raise SessionDead(f"HTTP {d.get('s')}: {str(d.get('b'))[:200]}")
    text = str(d.get("b") or "")
    if text.lstrip().startswith("<"):
        raise SessionDead("본문이 HTML이다 — 로그인 페이지로 밀려난 상태(200이라도 데이터가 아니다)")
    j = json.loads(text)
    if j.get("errors"):
        raise GraphQLError(json.dumps(j["errors"], ensure_ascii=False)[:400])
    return j


_DL_START = """(async(u)=>{try{
  const r = await fetch(u, {credentials:'include'});
  if(!r.ok) return JSON.stringify({s:r.status,n:0});
  const b = await r.arrayBuffer();
  window.%s = new Uint8Array(b);
  return JSON.stringify({s:r.status, n: window.%s.length});
}catch(e){ return JSON.stringify({err:String(e)}); }})(%s)"""

_DL_SLICE = """(()=>{const u=window.%s.subarray(%d,%d);let s='';const CH=0x8000;
for(let i=0;i<u.length;i+=CH)s+=String.fromCharCode.apply(null,u.subarray(i,i+CH));
return btoa(s);})()"""


def download_binary(call, url: str, *, chunk: int = 2 * 1024 * 1024,
                    progress=None) -> bytes:
    """바이너리를 **조각으로** 끌어온다 (결함 ②의 재발 방지선).

    한 번에 전부 base64로 만들어 돌려주면 큰 보고서에서 조용히 깨진다(2025-10 = 72MB).
    버퍼를 페이지에 남기고 `chunk`씩 가져와 파이썬에서 잇는다. 길이가 서버가 말한 값과
    다르면 예외 — **짧은 파일을 성공으로 쓰지 않는다.**
    """
    import base64

    raw = evaluate(call, _DL_START % (_BUF, _BUF, json.dumps(url)), await_promise=True)
    d = raw if isinstance(raw, dict) else json.loads(raw)
    if "err" in d:
        raise SessionDead(f"다운로드 fetch가 죽었다: {d['err'][:160]}")
    if d.get("s", 0) >= 400:
        raise RuntimeError(f"다운로드 HTTP {d['s']}")
    total = int(d.get("n") or 0)
    if total == 0:
        raise RuntimeError("다운로드 본문이 비었다(0바이트)")
    out = bytearray()
    try:
        off = 0
        while off < total:
            end = min(off + chunk, total)
            b64 = evaluate(call, _DL_SLICE % (_BUF, off, end))
            out.extend(base64.b64decode(b64))
            off = end
            if progress:
                progress(off, total)
    finally:
        try:
            evaluate(call, f"(()=>{{delete window.{_BUF};return 1;}})()")
        except Exception:  # noqa: BLE001 — 정리 실패가 본 작업을 덮지 않게 한다
            pass
    if len(out) != total:
        raise RuntimeError(f"다운로드가 잘렸다: {len(out):,}/{total:,} 바이트")
    return bytes(out)


def sniff_format(data: bytes) -> str:
    """받은 바이트가 무엇인지 판정한다 — `'xlsx'` | `'tsv'` | `'unknown'`.

    ★쿠팡 보고서는 **행이 많으면 xlsx가 아니라 TSV로 온다**(2025-10 NCA = 188,094행 / 72MB).
      확장자를 믿거나 xlsx라고 가정하면 `BadZipFile`이 나고 "다운로드가 깨졌다"고 오진한다.
    """
    if data[:4] == b"PK\x03\x04":
        return "xlsx"
    head = data[:4096].decode("utf-8", "replace")
    if "\t" in head and ("날짜" in head or "캠페인" in head):
        return "tsv"
    return "unknown"


def suggest_suffix(fmt: str) -> str:
    return ".xlsx" if fmt == "xlsx" else (".tsv" if fmt == "tsv" else ".bin")


def read_report(path: str):
    """광고 보고서를 **포맷 무관**하게 읽어 `(헤더리스트, 행 이터레이터)`로 돌려준다.

    xlsx면 openpyxl, TSV면 csv. 호출부가 포맷을 신경 쓰지 않게 해서 같은 오진이 재발하지
    않게 한다. 행은 헤더와 같은 길이의 리스트다(짧은 행은 버린다).
    """
    import csv
    import io as _io

    with open(path, "rb") as f:
        head = f.read(8)
    if head[:4] == b"PK\x03\x04":
        import warnings

        import openpyxl
        warnings.filterwarnings("ignore")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        hdr = [str(h) if h is not None else "" for h in next(it)]

        def rows():
            for r in it:
                yield ["" if v is None else v for v in r]
        return hdr, rows()

    text = open(path, "rb").read().decode("utf-8", "replace")
    rd = csv.reader(_io.StringIO(text), delimiter="\t", quotechar='"')
    hdr = next(rd)

    def rows_tsv():
        n = len(hdr)
        for r in rd:
            if len(r) >= n:
                yield r
    return hdr, rows_tsv()


def poll_report(call, report_id: str, report_type: str, q_list: str,
                *, tries: int = 120, interval: float = 3.0) -> None:
    """보고서 생성 완료까지 폴링. 타임아웃이면 예외(조용한 실패 금지)."""
    for _ in range(tries):
        time.sleep(interval)
        d = gql(call, {"variables": {"reportType": report_type, "page": 1,
                                     "pageSize": 20, "duration": 30},
                       "query": q_list})
        reports = ((d.get("data") or {}).get("reportList") or {}).get("reports") or []
        me = [r for r in reports if str(r.get("id")) == str(report_id)]
        status = str(me[0].get("status") if me else "?").lower()
        if status in ("completed", "done", "success"):
            return
    raise RuntimeError(f"보고서 {report_id} 생성이 {tries * interval:.0f}초 안에 안 끝났다")
