# ohitech_billboard_recon.py — [임시·S0 정찰] 오하이테크(9224) Billboard 옵션 보고서 라이브 캡처
# 목적(D-13 GATE): 1P 로켓배송 광고가 옵션(keyword) granularity Billboard 보고서를 제공하는지,
#   XLSX가 오픽스 keyword 포맷(파서 호환)인지, 옵션ID가 조인 가능한지 라이브 검증.
# 추정 금지: 오픽스 페처의 GraphQL 쿼리·_gql·_DL_JS + 오하이테크 _cdp_page(9224)를 그대로 재사용.
# S0 종료 후 이 파일은 제거한다(shipped 코드 아님).
import sys
import json
import base64
from datetime import datetime, timedelta, timezone

sys.path.insert(0, "/Users/jino/Library/Mobile Documents/com~apple~CloudDocs/1Personal/AI Program/Ohiselling/tools")

import ad_cost_browser_fetcher as ofix   # 오픽스 Billboard 흐름(쿼리/_gql/_DL_JS)
import ohitech_ad_fetcher as ohi          # 오하이테크 9224 CDP attach

KST = timezone(timedelta(hours=9))
OUT_XLSX = "/tmp/ohitech_billboard_recon.xlsx"


def _window() -> tuple[int, int]:
    """최근 7일(어제까지) YYYYMMDD 정수 (레퍼런스 16 D-2)."""
    end = (datetime.now(KST).date() - timedelta(days=1))
    start = end - timedelta(days=6)
    return int(start.strftime("%Y%m%d")), int(end.strftime("%Y%m%d"))


def main() -> int:
    cfg = ohi.load_config()
    start_i, end_i = _window()
    print(f"[S0] window {start_i}~{end_i} (cdp_port={cfg.get('cdp_port')})")

    # ★Chrome 수명은 페처와 같은 계약으로(2026-07-27 버튼-only 전환): _cdp_page는 이미 떠 있는
    #   Chrome만 attach하므로, 여기서 _owned_chrome으로 기동을 보장한다.
    with ohi._owned_chrome(cfg):
      with ohi._cdp_page(cfg) as (page, _browser):
        # 0) same-origin 확보 — advertising.coupang.com으로 navigate(동작하는 페처와 동일).
        try:
            page.goto(ohi.DASH_URL, wait_until="domcontentloaded", timeout=40000)
        except Exception as e:  # noqa: BLE001
            print(f"[S0] 대시보드 goto 경고(계속): {str(e)[:120]}")
        page.wait_for_timeout(2000)  # Akamai/세션 안정화
        if ohi._is_logged_out(page.url):
            # ★2026-08-03: 세션 만료가 더 이상 정찰의 중단 사유가 아니다(PR #175 병합).
            #   ①SSO → ②Keychain → ③알림 순으로 스스로 복구한다. 이 정찰이 7/11에 멈춰 있던
            #   이유가 바로 이 자리였다(`[S0][세션만료]`).
            print(f"[S0] 세션 만료 감지(url={page.url[:70]}) — 자가 복구 시도")
            if ohi._recover_session(page, cfg) != ohi.coupang_auth.OK:
                print("[S0][세션만료] 자가 복구 실패 — Chrome 창에서 재로그인 필요. (게이트 판정 아님)")
                return 4
            print("[S0] 세션 자가 복구 성공 — 정찰 계속")

        # 1) 캠페인 목록 (세션 스코프 — 9224=오하이테크 계정)
        data = ofix._gql(page, [{
            "operationName": "GetCampaignListInBillboard",
            "variables": {"startDate": start_i, "endDate": end_i, "reportType": "pa"},
            "query": ofix._Q_CAMPAIGNS,
        }])
        campaigns = (data or {}).get("getCampaignList") if data else None
        print(f"[S0] 캠페인 응답: {json.dumps(campaigns, ensure_ascii=False)[:600] if campaigns is not None else 'NONE/실패'}")
        if not campaigns:
            print("[S0][GATE-FAIL] 오하이테크 Billboard 캠페인 0개 — 옵션 granularity 미지원 가능. 중단·Jino 보고.")
            return 3
        campaign_ids = [str(c["id"]) for c in campaigns if c.get("id")]
        print(f"[S0] 캠페인 {len(campaign_ids)}개: {campaign_ids}")

        # 2) 보고서 생성 요청 (daily/keyword)
        data = ofix._gql(page, [{
            "variables": {
                "reportType": "pa", "startDate": start_i, "endDate": end_i,
                "dateGroup": "daily", "granularity": "keyword",
                "excludeIfNoClickCount": True, "campaignIds": campaign_ids,
            },
            "query": ofix._M_REQUEST_REPORT,
        }])
        req = (data or {}).get("requestReport") if data else None
        report_id = str(req["id"]) if req and req.get("id") else None
        print(f"[S0] requestReport: id={report_id} status={req.get('status') if req else None} "
              f"granularity={req.get('granularity') if req else None}")
        if not report_id:
            print("[S0][GATE-FAIL] requestReport id 없음 — 옵션 보고서 생성 불가. 중단·Jino 보고.")
            return 3

        # 3) 완료 폴링
        completed, waited = False, 0
        while waited < 300:
            page.wait_for_timeout(3000)
            waited += 3
            data = ofix._gql(page, [{
                "variables": {"reportType": "pa", "page": 1, "pageSize": 20,
                              "duration": 90, "onlyScheduledReport": False},
                "query": ofix._Q_REPORT_LIST,
            }])
            reports = ((data or {}).get("reportList") or {}).get("reports") if data else None
            mine = next((r for r in (reports or []) if str(r.get("id")) == report_id), None)
            if mine:
                st = str(mine.get("status")).lower()
                print(f"[S0] poll +{waited}s status={st} granularity={mine.get('granularity')}")
                if st == "completed":
                    completed = True
                    break
                if st in ("failed", "error"):
                    print(f"[S0][GATE-FAIL] 보고서 생성 실패 status={st}. 중단·Jino 보고.")
                    return 3
        if not completed:
            print("[S0][GATE-FAIL] 300s 내 보고서 미완료. 중단·Jino 보고.")
            return 3

        # 4) 다운로드
        dl = page.evaluate(ofix._DL_JS, ofix.EXCEL_REPORT_URL + report_id)
        if not dl or not dl.get("b64"):
            print(f"[S0][GATE-FAIL] 다운로드 본문 없음 status={dl and dl.get('status')}. 중단·Jino 보고.")
            return 3
        content = base64.b64decode(dl["b64"])
        if content[:4] != b"PK\x03\x04":
            print(f"[S0][GATE-FAIL] xlsx(zip) 아님(세션 만료 가능) head={content[:16]!r}. 중단·Jino 보고.")
            return 3
        with open(OUT_XLSX, "wb") as f:
            f.write(content)
        print(f"[S0] XLSX 저장: {OUT_XLSX} ({len(content)} bytes)")

    # 5) 컬럼 검증 (파서 호환 — 오픽스 keyword 포맷: [8]광고집행 옵션ID, [10]전환 옵션ID)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(OUT_XLSX, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(max_row=6, values_only=True))
        print(f"[S0] 시트='{ws.title}' 총행≈{ws.max_row} 열수={ws.max_column}")
        for i, r in enumerate(rows):
            print(f"[S0] row{i}: {[str(c)[:22] if c is not None else '' for c in r]}")
    except ImportError:
        print("[S0] openpyxl 없음 — XLSX 수동 검사 필요(/tmp/ohitech_billboard_recon.xlsx)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
