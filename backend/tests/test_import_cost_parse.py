# test_import_cost_parse.py — 서류 업로드 → 폼 초안 (D-CPP-48)
#
# ★픽스처가 prod보다 관대하면 안 된다(교훈 #292). 그래서 통관경비서 픽스처는 **pypdf가 실제
#   서류에서 뽑은 텍스트 그대로**다 — 라벨과 값이 줄바꿈으로 갈리고 순서까지 뒤섞인,
#   내가 «정리해 준» 형태가 아니라 추출기가 실제로 내놓는 형태.
#   초판 파서는 정리된 텍스트로만 검증돼 실파일에서 0줄을 뽑았다.
from __future__ import annotations

import io
from decimal import Decimal as D

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base, get_db
from app.main import app
from app.services.import_cost.allocator import CostLine, InvoiceLine, allocate, costing_pool
from app.services.import_cost.parser import CustomsDocParseError, parse_customs_expense


@pytest.fixture
def client():
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    Base.metadata.create_all(engine)
    TestingSession = sessionmaker(bind=engine, autoflush=False, autocommit=False)

    def _override_get_db():
        db = TestingSession()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = _override_get_db
    tc = TestClient(app)
    yield tc
    app.dependency_overrides.clear()


# ──────────────────────────────────────────────
# 통관경비서 — pypdf 6.16.1이 SETR2608170216에서 실제로 뽑은 텍스트
# (표를 셀 단위로 뱉어 앞 7행은 「금액→이름→세액」, 관세 이후 3행은 「이름→금액→세액」)
# ──────────────────────────────────────────────
REAL_PDF_TEXT = """:
 23,100.00
CNY
INV Value
287.800
Gross W/T
백마보세창고
[02006103]
창고명
19 CARTONS / 1.01
수량/Meaure
HANSUNG INCHEON /
3594E
선명/항차
2026-08-18
ETA
SCREEN PROTECTOR CLEANING KITS
품명
SHENZHEN OTAO
TECHNOLOGY LIMITED
Shipper NM
비       고
예상세액
내       역
예상공급가액
(주)세아트랜스
2026-08-18
오하이테크  505-88-02821

* 통관자금 예상비용 명세표 및 송금요청
SETR2608170216
HBL NO :
날             짜
담    당    자
수           신
수입화물 통관비 예상내역서
:
:
209.88
환율
3,434
신고금액 ($)
4,862,657
과세금액 (₩)

면허일자
2026-08-18
신고일자
15443-26-701565M
신고번호
156,550
OCEAN FREIGHT(해상운임)

0
18,000
OVER WEIGHT CHARGES

0
35,000
C/O(원산지증명서비용)

0
44,000
신고비

0
18,400
PICKUP CHARGE IN CHINA

0
25,000
DOCUMENT FEE

0
90,000
국내운송료 ( 라보 * 1 )

9,000
B/L Sub Total
386,950
9,000


관세
249,670

0
부가세
511,230

0
통관수수료
25,000

2,500
1,172,850

11,500
소       계

1,184,350
합       계
:
송금계좌번호
기업은행 : 476-050803-01-015
-1,184,350
잔       액
1,184,350
합       계
0
가입금액

처리일자
비          고
"""


def test_expense_parse_from_real_extractor_output():
    """실 추출 텍스트에서 비용 10줄 · pool 661,620 · 부가세 1줄 제외."""
    ex = parse_customs_expense(REAL_PDF_TEXT)
    assert len(ex.cost_lines) == 10, [c.item_name for c in ex.cost_lines]
    assert costing_pool(ex.cost_lines) == D("661620")
    non_costing = [c.item_name for c in ex.cost_lines if not c.is_costing]
    assert non_costing == ["부가세"]
    # 소계·합계 행이 데이터로 섞이지 않았다 — 섞이면 통관비가 2배가 된다
    names = " ".join(c.item_name for c in ex.cost_lines)
    assert "Sub Total" not in names and "소" not in names.replace("소계", "")
    # 소계 뒤 송금계좌 구역이 가짜 행으로 딸려오지 않았다
    assert not any("기업은행" in c.item_name for c in ex.cost_lines)


def test_expense_parse_header_fields():
    """라벨과 값이 뒤바뀐 레이아웃에서도 헤더가 **정확히** 나온다.

    초판은 정규식이 「환율\\n3,434」에 걸려 fx_rate=3을 냈다 — 값이 «있는데 틀린 것»이
    «없는 것»보다 나쁘다. 그래서 형태 검증을 통과한 후보만 채택한다.
    """
    ex = parse_customs_expense(REAL_PDF_TEXT)
    assert ex.hbl_no == "SETR2608170216"
    assert ex.declaration_no == "15443-26-701565M"
    assert ex.fx_rate == D("209.88")
    assert ex.customs_value_krw == D("4862657")
    assert ex.currency == "CNY"
    assert ex.declared_inv_value == D("23100.00")
    assert ex.shipper_name == "SHENZHEN OTAO TECHNOLOGY LIMITED"
    assert ex.vessel == "HANSUNG INCHEON / 3594E"
    assert ex.carton_count == 19
    assert ex.gross_weight_kg == D("287.800")
    assert ex.declaration_date is not None and ex.declaration_date.isoformat() == "2026-08-18"


def test_expense_self_check_rejects_partial_parse():
    """행을 하나 지우면 **자기검산이 막는다** — 조용히 적은 금액으로 채워지지 않는다."""
    broken = REAL_PDF_TEXT.replace("249,670\n\n0\n부가세", "부가세")  # 관세 행 훼손
    with pytest.raises(CustomsDocParseError, match="자기검산"):
        parse_customs_expense(broken)


def test_expense_parse_raises_on_garbage():
    with pytest.raises(CustomsDocParseError):
        parse_customs_expense("이건 통관경비서가 아니다")


# ──────────────────────────────────────────────
# CI/PL 엑셀 — 실 서류와 같은 레이아웃을 xlsx로 만든다
# ──────────────────────────────────────────────
CI_ROWS = [
    ("116\n（20260708-1）", "Privacy Glass_iP16 Pro 2ea", 50, 19.2),
    (None, "Privacy Glass_iP15 Pro 2ea", 50, 19.2),
    (None, "Glass_Ip17Pro", 500, 12.2),
    (None, "Glass_Ip16 Pro", 350, 12.2),
    (None, "Glass_Ip16 Plus", 50, 12.2),
    (None, "Glass_iP15", 50, 12.2),
    (None, "Glass_iP15 pro", 200, 12.2),
    (None, "Glass_iP14promax", 50, 12.2),
    (None, "Glass_iP13/13pro", 50, 12.2),
    ("115\n（20260617-1）", "Glass_iP15 promax", 100, 12.2),
    (None, "Glass_iP15 plus", 50, 12.2),
    ("114\n（20260528-1）", "Privacy Glass_iP14pro 2ea", 50, 19.2),
    ("Order 112", "Glass_iP13promax", 50, 12.2),
    ("Order 100", "Glass_iP12promax", 50, 12.2),
    (None, "cleaning kits", 2400, 0.8),
]

# (박스, 품명, 수량, 박스당수량, 박스수, 총중량, 규격, CBM) — 박스 공유분은 첫 행에만 값이 있다
PL_ROWS = [
    ("1", "Privacy Glass_iP16 Pro 2ea", 50, 100, 1, 15.1, "50.5*44*24", 0.053328),
    (None, "Privacy Glass_iP15 Pro 2ea", 50, None, None, None, None, None),
    ("2-6", "Glass_Ip17Pro", 500, 100, 5, 74.0, "50.5*44*24", 0.26664),
    ("7-9", "Glass_Ip16 Pro", 300, 100, 3, 45.3, "50.5*44*24", 0.159984),
    ("10", "Glass_Ip16 Plus", 50, 100, 1, 14.9, "50.5*44*24", 0.053328),
    (None, "Glass_Ip16 Pro", 50, None, None, None, None, None),
    ("11", "Glass_iP15", 50, 50, 1, 14.9, "50.5*44*24", 0.053328),
    ("12-13", "Glass_iP15 pro", 200, 100, 2, 30.0, "50.5*44*24", 0.106656),
    ("14", "Glass_iP14promax", 50, 100, 1, 15.2, "50.5*44*24", 0.053328),
    (None, "Glass_iP13/13pro", 50, None, None, None, None, None),
    ("15", "Glass_iP15 promax", 100, 100, 1, 15.0, "50.5*44*24", 0.053328),
    ("16", "Glass_iP15 plus", 50, 100, 1, 15.1, "50.5*44*24", 0.053328),
    (None, "Privacy Glass_iP14pro 2ea", 50, None, None, None, None, None),
    ("17", "Glass_iP13promax", 50, 100, 1, 15.3, "50.5*44*24", 0.053328),
    (None, "Glass_iP12promax", 50, None, None, None, None, None),
    ("18-19", "cleaning kits", 2400, 1200, 2, 33.0, "50.5*44*24", 0.106656),
]


def _build_workbook() -> bytes:
    wb = Workbook()
    ci = wb.active
    ci.title = "CI"
    ci["B8"] = "COMMERCIAL  INVOICE"
    # ★라벨을 실물과 똑같이 넣는다. 초판 픽스처는 값만 넣어 «실물보다 관대»했고,
    #   파서가 열 위치 대신 라벨로 찾도록 고치자 그 관대함이 드러났다(교훈 #292).
    ci["F9"] = "Invoice Date:"
    ci["G9"] = 46245.0            # Invoice Date (엑셀 시리얼)
    ci["F10"] = "Invoice No.:"
    ci["G10"] = "SO-WSOH-116-115-114"
    ci["B14"], ci["D14"] = "Order No.", "Item "
    ci["E14"], ci["F14"], ci["G14"] = "Quantity   (pcs)", "Unit Price     (CNY)", "Total(CNY)"
    r = 15
    for order, item, qty, price in CI_ROWS:
        if order:
            ci.cell(row=r, column=2, value=order)
        ci.cell(row=r, column=4, value=item)
        ci.cell(row=r, column=5, value=qty)
        ci.cell(row=r, column=6, value=price)
        ci.cell(row=r, column=7, value=round(qty * price, 2))
        r += 1
    ci.cell(row=r, column=2, value="Total QTY:")
    ci.cell(row=r, column=5, value=sum(q for _, _, q, _ in CI_ROWS))
    ci.cell(row=r, column=6, value="Total(RMB):")
    ci.cell(row=r, column=7, value=23100.0)

    pl = wb.create_sheet("PL")
    pl["A13"], pl["C13"], pl["D13"] = "Ctn No.:/箱号", "Description of Goods", " Shipment Qty"
    pl["E13"], pl["F13"], pl["G13"] = "Quantity", "CTN QTY", "G.W"
    pl["H13"], pl["I13"], pl["J13"] = "Total\nG.W", "Measure", "Total volume"
    r = 14
    for ctn, item, qty, per, cnt, gw, meas, cbm in PL_ROWS:
        if ctn:
            pl.cell(row=r, column=1, value=ctn)
        pl.cell(row=r, column=3, value=item)
        pl.cell(row=r, column=4, value=qty)
        if per is not None:
            pl.cell(row=r, column=5, value=per)
            pl.cell(row=r, column=6, value=cnt)
            pl.cell(row=r, column=8, value=gw)
            pl.cell(row=r, column=9, value=meas)
            pl.cell(row=r, column=10, value=cbm)
        r += 1
    pl.cell(row=r, column=3, value="TOTAL:       ")
    pl.cell(row=r, column=4, value=4050)
    pl.cell(row=r, column=6, value=19)
    pl.cell(row=r, column=8, value=287.8)
    pl.cell(row=r, column=10, value=1.013232)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────
# 엔드포인트 — 업로드가 폼을 채운다 (저장하지 않는다)
# ──────────────────────────────────────────────
def _parse(client, *, xlsx: bytes | None = None, text: str | None = None):
    files = {}
    if xlsx is not None:
        files["ci_pl_file"] = ("doc.xlsx", xlsx, "application/vnd.ms-excel")
    data = {"expense_text": text} if text is not None else None
    r = client.post("/api/import-cost/parse", files=files or None, data=data)
    assert r.status_code == 200, r.text
    return r.json()


def test_parse_endpoint_fills_form_from_two_documents(client):
    """★핵심 시나리오: 서류 2개를 올리면 폼이 채워지고, 그 값으로 계약의 정답 단가가 나온다."""
    body = _parse(client, xlsx=_build_workbook(), text=REAL_PDF_TEXT)
    assert body["errors"] == [], body["errors"]

    h = body["header"]
    assert h["hbl_no"] == "SETR2608170216"
    assert h["fx_rate"] == "209.88"
    assert h["carton_count"] == 19
    assert h["invoice_no"] == "SO-WSOH-116-115-114"

    assert len(body["invoice_lines"]) == 15
    assert len(body["packing_lines"]) == 16
    assert len(body["cost_lines"]) == 10

    # PL의 중량·부피가 CI 라인에 배분돼 실렸다 — weight/volume 배부기준의 원료다
    weights = [D(r["gross_weight_kg"]) for r in body["invoice_lines"] if r["gross_weight_kg"]]
    assert len(weights) == 15
    assert abs(sum(weights) - D("287.8")) <= D("0.01")

    # 폼 값 그대로 배부하면 합격기준 숫자가 나온다
    cost = [
        CostLine(c["item_name"], D(c["supply_amount"]), D(c["tax_amount"]), c["is_costing"])
        for c in body["cost_lines"]
    ]
    assert costing_pool(cost) == D("661620")
    inv = [
        InvoiceLine(
            r["seq"], r["item_name"], D(r["quantity"]), D(r["unit_price_foreign"]),
            D(r["gross_weight_kg"]) if r["gross_weight_kg"] else None,
            D(r["cbm"]) if r["cbm"] else None,
        )
        for r in body["invoice_lines"]
    ]
    res = allocate(inv, cost, D(h["fx_rate"]), "amount")
    assert res.unallocated_krw == D("0")
    ip17 = next(x for x in res.lines if x.item_name == "Glass_Ip17Pro")
    kits = next(x for x in res.lines if x.item_name == "cleaning kits")
    assert abs(ip17.unit_cost_ex_vat - D("2910")) <= D("1")
    assert abs(ip17.unit_cost_inc_vat - D("3201")) <= D("1")
    assert abs(kits.allocated_cost_krw - D("54992")) <= D("1")


def test_parse_does_not_guess_line_type(client):
    """분류를 자동으로 찍지 않는다 — cleaning kits가 조용히 «판매 상품»이 되면 안 된다."""
    body = _parse(client, xlsx=_build_workbook())
    assert {r["line_type"] for r in body["invoice_lines"]} == {"unknown"}
    assert all(r["internal_sku"] is None for r in body["invoice_lines"])


def test_parse_is_partial_and_reports_failures(client):
    """하나가 깨져도 나머지는 채워지고, **깨진 것이 errors에 남는다**(조용한 빈 폼 금지)."""
    body = _parse(client, xlsx=b"not an excel file", text=REAL_PDF_TEXT)
    assert len(body["cost_lines"]) == 10          # 경비서는 정상
    assert body["invoice_lines"] == []            # 엑셀은 실패
    assert len(body["errors"]) >= 1
    assert any("Invoice" in e or "Packing" in e for e in body["errors"])


def test_parse_saves_nothing(client):
    """파싱은 저장하지 않는다 — 정본은 사람이 확인한 폼이다."""
    _parse(client, xlsx=_build_workbook(), text=REAL_PDF_TEXT)
    assert client.get("/api/import-cost/shipments").json()["count"] == 0


def test_parse_requires_at_least_one_document(client):
    assert client.post("/api/import-cost/parse").status_code == 400


# ──────────────────────────────────────────────
# 열 밀림 — 같은 공급사인데 양식마다 열이 한 칸 다르다 (2026-08-22 SO-WSOH-114 실측)
# ──────────────────────────────────────────────
def _build_shifted_workbook() -> bytes:
    """`SO-WSOH-114` 양식 — CI·PL이 **각각 별도 파일**이고 열이 한 칸씩 왼쪽이다.

    초판 파서는 열을 D·E·F로 하드코딩해서 이 양식을 **예외 없이** 잘못 읽었다:
    품명 자리에 `100.0`(수량), 수량 자리에 `12.2`(단가)가 실렸다.
    """
    wb = Workbook()
    ci = wb.active
    ci.title = "CI"
    ci["B8"] = "COMMERCIAL  INVOICE"
    ci["E9"], ci["F9"] = "Invoice Date:", 46218.0
    ci["E10"], ci["F10"] = "Invoice No.:", "SO-WSOH-114"
    ci["B14"], ci["C14"] = "Order No.", "Item "
    ci["D14"], ci["E14"], ci["F14"] = "Quantity   (pcs)", "Unit Price     (CNY)", "Total(CNY)"
    rows = [("114\n（20260528-1）", "Glass_Ip16", 100, 12.2), (None, "cleaning kits", 12000, 0.8)]
    r = 15
    for order, item, qty, price in rows:
        if order:
            ci.cell(row=r, column=2, value=order)
        ci.cell(row=r, column=3, value=item)
        ci.cell(row=r, column=4, value=qty)
        ci.cell(row=r, column=5, value=price)
        ci.cell(row=r, column=6, value=round(qty * price, 2))
        r += 1
    ci.cell(row=r, column=2, value="Total QTY:")
    ci.cell(row=r, column=4, value=12100)
    ci.cell(row=r, column=5, value="Total(RMB):")
    ci.cell(row=r, column=6, value=10820.0)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_shifted_pl() -> bytes:
    """같은 건의 PL — **별도 파일**이고 역시 열이 한 칸 왼쪽이다."""
    wb = Workbook()
    pl = wb.active
    pl.title = "PL"
    pl["H11"], pl["J11"] = "Invoice No.:", "SO-WSOH-114"
    pl["A13"], pl["B13"], pl["C13"] = "Ctn No.:/箱号", "Description of Goods", " Shipment Qty"
    pl["D13"], pl["E13"], pl["F13"] = "Quantity", "CTN QTY", "G.W"
    pl["G13"], pl["H13"], pl["I13"] = "Total\nG.W", "Measure", "Total volume"
    data = [("1", "Glass_Ip16", 100, 100, 1, 14.8, 14.8, "50.5*44*24", 0.053328),
            ("2-11", "cleaning kits", 12000, 1200, 10, 16.5, 165.0, "50.5*44*24", 0.53328)]
    r = 14
    for ctn, item, qty, per, cnt, gw1, gw, meas, cbm in data:
        for c, v in enumerate([ctn, item, qty, per, cnt, gw1, gw, meas, cbm], start=1):
            pl.cell(row=r, column=c, value=v)
        r += 1
    pl.cell(row=r, column=2, value="TOTAL:       ")
    pl.cell(row=r, column=3, value=12100)
    pl.cell(row=r, column=5, value=11)
    pl.cell(row=r, column=7, value=179.8)
    pl.cell(row=r, column=9, value=0.586608)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_column_shifted_format_is_read_correctly(client):
    """★열이 밀린 양식을 «정확히» 읽는다 — 조용한 오독이 이 파일의 존재 이유다."""
    body = _parse(client, xlsx=_build_shifted_workbook())
    assert body["errors"] == [], body["errors"]
    # PL 시트가 없는 것은 «실패»가 아니라 «안 준 것» — 경고로 나오고 오류로는 안 센다
    assert any("PL 시트가 없다" in w for w in body["warnings"]), body["warnings"]
    lines = body["invoice_lines"]
    assert len(lines) == 2
    assert [l["item_name"] for l in lines] == ["Glass_Ip16", "cleaning kits"]
    assert [D(l["quantity"]) for l in lines] == [D("100"), D("12000")]
    assert [D(l["unit_price_foreign"]) for l in lines] == [D("12.2"), D("0.8")]
    assert body["header"]["invoice_no"] == "SO-WSOH-114"
    # 품명 자리에 숫자가 실리면 그게 초판의 결함이다
    assert not any(l["item_name"].replace(".", "").isdigit() for l in lines)


def test_separate_pl_file_is_accepted(client):
    """CI와 PL이 **별도 파일**로 와도 둘 다 읽는다."""
    r = client.post(
        "/api/import-cost/parse",
        files={
            "ci_pl_file": ("ci.xlsx", _build_shifted_workbook(), "application/vnd.ms-excel"),
            "pl_file": ("pl.xlsx", _build_shifted_pl(), "application/vnd.ms-excel"),
        },
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["errors"] == [], body["errors"]
    assert len(body["invoice_lines"]) == 2
    assert len(body["packing_lines"]) == 2
    assert body["header"]["carton_count"] == 11
    # PL 중량이 CI 라인에 배분돼 실렸다
    weights = [D(l["gross_weight_kg"]) for l in body["invoice_lines"] if l["gross_weight_kg"]]
    assert len(weights) == 2
    assert abs(sum(weights) - D("179.8")) <= D("0.01")


def test_ci_self_check_catches_column_map_drift(client):
    """총계 행과 안 맞으면 **예외로 막는다** — 초판엔 이 검산이 CI에 없었다."""
    from app.services.import_cost.parser import CustomsDocParseError, parse_commercial_invoice

    wb = _build_shifted_workbook()
    # 총계만 틀리게 바꾼 워크북을 만든다(라인은 그대로)
    from openpyxl import load_workbook as _lw

    book = _lw(io.BytesIO(wb))
    book["CI"].cell(row=17, column=6, value=99999.0)
    buf = io.BytesIO()
    book.save(buf)
    with pytest.raises(CustomsDocParseError, match="자기검산"):
        parse_commercial_invoice(buf.getvalue())


def test_numeric_item_name_is_rejected(client):
    """품명 자리에 숫자가 오면 거부한다 — 열 지도가 어긋났다는 뜻이다."""
    from app.services.import_cost.parser import CustomsDocParseError, parse_commercial_invoice
    from openpyxl import load_workbook as _lw

    book = _lw(io.BytesIO(_build_shifted_workbook()))
    book["CI"].cell(row=15, column=3, value=100.0)  # 품명 자리에 숫자
    buf = io.BytesIO()
    book.save(buf)
    with pytest.raises(CustomsDocParseError, match="열 지도"):
        parse_commercial_invoice(buf.getvalue())


# ──────────────────────────────────────────────
# 13개월 전수(2026-08-22, 세아트랜스 메일 14건 + Drive 30건)에서 나온 두 양식 변종
# ──────────────────────────────────────────────
def test_sheet_name_variant_ci_2_is_accepted():
    """시트명이 `CI (2)`여도 읽는다 — 엑셀에서 시트를 복제하면 붙는 접미사다.

    실측: 2026-03-18 선적건의 워크북 시트가 `['CI (2)', 'PL']`이었고, 정확 일치만 보던
    초판은 그 선적건을 통째로 못 읽었다(14건 중 1건).
    """
    from openpyxl import load_workbook as _lw
    from app.services.import_cost.parser import parse_commercial_invoice

    book = _lw(io.BytesIO(_build_workbook()))
    book["CI"].title = "CI (2)"
    buf = io.BytesIO()
    book.save(buf)
    ci = parse_commercial_invoice(buf.getvalue())
    assert len(ci.lines) == 15
    assert ci.lines[2].item_name == "Glass_Ip17Pro"


def test_item_column_is_resolved_from_data_not_label():
    """헤더의 `Item` 라벨이 한 칸 왼쪽이어도 실제 품목 열을 찾는다.

    실측: 2026-04-01 선적건은 헤더 `Item`이 C열인데 품목명은 D열이고, C열엔 분류어
    (`screen protector`)가 첫 행에만 있었다. 같은 공급사의 다른 파일은 라벨이 D였다.
    ⇒ 라벨을 그대로 믿으면 「품명이 비었는데 수량은 있다」로 죽는다(14건 중 1건).
    ★헤더 라벨은 «힌트»이고 판정은 **데이터**가 한다.
    """
    from app.services.import_cost.parser import parse_commercial_invoice

    wb = Workbook()
    ci = wb.active
    ci.title = "CI"
    ci["F10"], ci["G10"] = "Invoice No.:", "SO-WSOH-TEST"
    # ★라벨은 C에 두고 실제 품목은 D에 둔다(실물 그대로)
    ci["B14"], ci["C14"] = "Order No.", "Item "
    ci["E14"], ci["F14"], ci["G14"] = "Quantity   (pcs)", "Unit Price     (CNY)", "Total(CNY)"
    rows = [("Order 111", "screen protector", "Glass_Ip16 Pro", 200, 12.2),
            (None, None, "Glass_iP15 promax", 100, 12.2),
            (None, None, "cleaning kits", 7000, 0.8)]
    r = 15
    for order, cat, item, qty, price in rows:
        if order:
            ci.cell(row=r, column=2, value=order)
        if cat:
            ci.cell(row=r, column=3, value=cat)
        ci.cell(row=r, column=4, value=item)
        ci.cell(row=r, column=5, value=qty)
        ci.cell(row=r, column=6, value=price)
        ci.cell(row=r, column=7, value=round(qty * price, 2))
        r += 1
    ci.cell(row=r, column=2, value="Total QTY:")
    ci.cell(row=r, column=5, value=7300)
    ci.cell(row=r, column=6, value="Total(RMB):")
    # 200×12.2 + 100×12.2 + 7000×0.8 = 9,260 (처음에 9,270으로 썼다가 자기검산에 걸렸다 —
    # 그 가드가 실제로 «내 산수 오류»를 잡은 것이라 기록해 둔다)
    ci.cell(row=r, column=7, value=9260.0)
    buf = io.BytesIO()
    wb.save(buf)

    res = parse_commercial_invoice(buf.getvalue())
    assert [l.item_name for l in res.lines] == [
        "Glass_Ip16 Pro", "Glass_iP15 promax", "cleaning kits"
    ]
    assert [D(str(l.quantity)) for l in res.lines] == [D("200"), D("100"), D("7000")]
    # 분류어가 품목명 자리에 들어오면 안 된다
    assert "screen protector" not in [l.item_name for l in res.lines]


def test_xls_without_xlrd_says_what_to_do(client, monkeypatch):
    """`xlrd`가 없으면 **무엇을 하면 되는지** 말한다 — 조용히 빈 폼을 주지 않는다."""
    import builtins

    real_import = builtins.__import__

    def _no_xlrd(name, *a, **kw):
        if name == "xlrd":
            raise ImportError("no xlrd")
        return real_import(name, *a, **kw)

    monkeypatch.setattr(builtins, "__import__", _no_xlrd)
    ole2 = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64
    body = _parse(client, xlsx=ole2)
    assert body["invoice_lines"] == []
    assert any("xlrd" in e and (".xlsx" in e or "수기" in e) for e in body["errors"]), body["errors"]
