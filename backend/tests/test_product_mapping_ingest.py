# test_product_mapping_ingest.py — 상품 연관맵 S1 매핑 적재 Harness 테스트
# 트랙 docs/tracks/active/track_product-connection-map.md D-1~D-6.
# 인메모리 SQLite + 실제 openpyxl Worksheet(라이브 실측 헤더 구조 재현)로 검증.
from __future__ import annotations

import io
from decimal import Decimal

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base, get_db
from app.main import app
from app.models import Channel, ProductChannelMapping, ProductMaster
from app.services.coupang.product_sync import _maybe_upsert_mapping
from app.services.product_mapping_ingest import (
    _next_sku_counter,
    check_integrity,
    ingest_master_sheet,
    parse_master_sheet,
    resolve_channel_label,
    upsert_mapping,
    upsert_product_by_name,
)


@pytest.fixture
def db():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    session = sessionmaker(bind=engine)()
    yield session
    session.close()


def _seed_channels(db):
    channels = [
        Channel(name="자사몰(cafe24)", code="CAFE24", platform="cafe24", channel_type="own"),
        Channel(name="네이버", code="NAVER", platform="naver", channel_type="marketplace"),
        Channel(name="쿠팡 로켓배송", code="COUPANG_ROCKET", platform="coupang", channel_type="consignment"),
        Channel(name="쿠팡 Wing2", code="COUPANG_WING2", platform="coupang", channel_type="marketplace"),
        Channel(name="쿠팡 Wing1", code="COUPANG_WING1", platform="coupang", channel_type="marketplace"),
    ]
    db.add_all(channels)
    db.commit()
    return {c.code: c for c in channels}


def _build_master_ws():
    """라이브 실측(2026-07-03) 헤더 구조를 그대로 재현: 채널명 마커 + 옵션ID 블록."""
    wb = Workbook()
    ws = wb.active
    ws.title = "원가 매핑"
    ws.append([
        "상품명", "원가",
        "채널명", "카페24 품목코드1", "카페24 품목코드2",
        "채널명", "스스 품목코드1",
        "채널명", "쿠팡 옵션ID 1", "쿠팡 옵션ID 2",
    ])
    ws.append([
        "필름 A", 1000,
        "자사몰 (cafe24)", "P001", "P002",
        "네이버 스마트스토어", "N001",
        "COUPANG_ohitech_판매자", "91576482085", None,
    ])
    ws.append([
        "필름 B", 2000,
        "자사몰 (cafe24)", "P003", None,
        "네이버 스마트스토어", "N002",
        "COUPANG_ohitech_판매자", "91576482086", None,
    ])
    return ws


# ── SA: 엑셀 파서 ──
def test_parse_master_sheet_blocks_and_labels():
    ws = _build_master_ws()
    rows = parse_master_sheet(ws)
    assert len(rows) == 2
    r1 = rows[0]
    assert r1.product_name == "필름 A"
    assert r1.cost_price == Decimal("1000")
    assert r1.channel_ids["자사몰 (cafe24)"] == ["P001", "P002"]
    assert r1.channel_ids["네이버 스마트스토어"] == ["N001"]
    assert r1.channel_ids["COUPANG_ohitech_판매자"] == ["91576482085"]


def test_parse_master_sheet_empty_id_cells_skipped():
    ws = _build_master_ws()
    rows = parse_master_sheet(ws)
    # 필름 B는 카페24 2번째 컬럼이 비어있음 → 1개만 담김
    assert rows[1].channel_ids["자사몰 (cafe24)"] == ["P003"]


# ── SA: 채널 라벨 리졸버 ──
def test_resolve_channel_label_known():
    assert resolve_channel_label("자사몰 (cafe24)") == "CAFE24"
    assert resolve_channel_label("COUPANG_ohitech_로켓배송") == "COUPANG_ROCKET"


def test_resolve_channel_label_unknown_returns_none():
    assert resolve_channel_label("존재하지않는라벨") is None


# ── SA: 무결성 검사 ──
def test_check_integrity_detects_unknown_label():
    # 블록 라벨은 그 블록의 "첫 데이터 행"에서 고정되므로(파서 설계), 미등록 라벨을
    # 재현하려면 첫 행부터 미등록 라벨이어야 한다.
    wb = Workbook()
    ws = wb.active
    ws.title = "원가 매핑"
    ws.append(["상품명", "원가", "채널명", "옵션ID"])
    ws.append(["필름 C", 500, "미확인채널", "X001"])
    rows = parse_master_sheet(ws)
    report = check_integrity(rows)
    assert any("미확인채널" in x for x in report.unknown_labels)


def test_parse_master_sheet_label_mismatch_not_silently_misattributed():
    # codex P1: 블록 고정라벨과 다른 라벨이 나오는 행은 그 블록 ID를 고정라벨로 귀속시키지 않고
    # label_mismatches에 표면화해야 한다(잘못된 채널로 조용히 매핑되는 것 방지).
    wb = Workbook()
    ws = wb.active
    ws.title = "원가 매핑"
    ws.append(["상품명", "원가", "채널명", "옵션ID"])
    ws.append(["필름 A", 1000, "자사몰 (cafe24)", "P001"])
    ws.append(["필름 D", 700, "미확인채널", "X999"])  # 같은 마커컬럼, 다른 라벨

    rows = parse_master_sheet(ws)
    assert rows[0].channel_ids == {"자사몰 (cafe24)": ["P001"]}
    # 필름 D 행은 X999를 "자사몰 (cafe24)"로 귀속시키지 않음(불일치 행은 통째로 스킵)
    assert rows[1].channel_ids == {}
    assert rows[1].label_mismatches == [(2, "자사몰 (cafe24)", "미확인채널")]

    report = check_integrity(rows)
    assert any("불일치" in x and "미확인채널" in x for x in report.label_mismatches)


def test_check_integrity_detects_duplicate_product_name_and_channel_id():
    ws = _build_master_ws()
    ws.append(["필름 A", 999, "자사몰 (cafe24)", "P001", None, None, None, None, None, None])
    rows = parse_master_sheet(ws)
    report = check_integrity(rows)
    assert any("필름 A" in x for x in report.duplicate_product_names)
    assert any("P001" in x for x in report.duplicate_channel_ids)


# ── SA: 상품/매핑 upsert ──
def test_upsert_product_by_name_creates_then_updates(db):
    counter = [1]
    p1, action1 = upsert_product_by_name(db, "필름 A", Decimal("1000"), counter)
    assert action1 == "created"
    assert p1.internal_sku == "OHI-0001"

    p2, action2 = upsert_product_by_name(db, "필름 A", Decimal("1200"), counter)
    assert action2 == "updated"
    assert p2.id == p1.id
    assert p2.cost_price == Decimal("1200")


def test_sku_counter_ignores_non_numeric_skus(db):
    """수동 명명 SKU가 섞여 있어도 채번은 숫자형 최댓값+1이어야 한다.

    2026-08-05 라이브 500 재현: 옛 코드는 문자열 정렬 최댓값('OHI-Z-PRIVACY-FOLD8-WIDE')의
    가운데 토막을 int()로 바꾸다 실패하면 조용히 1부터 다시 세어, 이미 있는 'OHI-0001'과
    UNIQUE 충돌을 냈다 → 신규 상품이 든 업로드가 전부 500.
    """
    db.add_all([
        ProductMaster(internal_sku="OHI-0001", product_name="기존 A", cost_price=Decimal("100")),
        ProductMaster(internal_sku="OHI-0895", product_name="기존 B", cost_price=Decimal("100")),
        ProductMaster(internal_sku="OHI-Z-PRIVACY-FOLD8-WIDE", product_name="수동 C", cost_price=Decimal("100")),
    ])
    db.commit()

    counter = _next_sku_counter(db)
    assert counter == [896]  # 'Z-PRIVACY…'가 아니라 0895 기준

    product, action = upsert_product_by_name(db, "신규 D", Decimal("500"), counter)
    db.commit()
    assert action == "created"
    assert product.internal_sku == "OHI-0896"


def test_allocate_sku_skips_taken_numbers(db):
    """카운터가 이미 쓰인 번호를 가리켜도 건너뛴다 — 한 건의 충돌이 적재 전체를 죽이지 않게."""
    db.add_all([
        ProductMaster(internal_sku="OHI-0002", product_name="기존 A", cost_price=Decimal("100")),
        ProductMaster(internal_sku="OHI-0003", product_name="기존 B", cost_price=Decimal("100")),
    ])
    db.commit()

    counter = [2]  # 이미 쓰인 번호를 가리키는 상태
    product, action = upsert_product_by_name(db, "신규 C", Decimal("500"), counter)
    db.commit()
    assert action == "created"
    assert product.internal_sku == "OHI-0004"


def test_upsert_mapping_created_updated_conflict(db):
    channels = _seed_channels(db)
    counter = [1]
    p1, _ = upsert_product_by_name(db, "필름 A", Decimal("1000"), counter)
    p2, _ = upsert_product_by_name(db, "필름 B", Decimal("2000"), counter)
    db.commit()

    outcome1 = upsert_mapping(db, p1.id, channels["CAFE24"], "P001")
    assert outcome1 == "created"

    outcome2 = upsert_mapping(db, p1.id, channels["CAFE24"], "P001")
    assert outcome2 == "updated"

    outcome3 = upsert_mapping(db, p2.id, channels["CAFE24"], "P001")
    assert outcome3 == "conflict"
    # 충돌 시 기존 매핑은 p1을 그대로 유지(덮어쓰지 않음)
    row = db.query(ProductChannelMapping).filter_by(channel_id=channels["CAFE24"].id, channel_product_id="P001").one()
    assert row.product_id == p1.id
    assert row.mapping_source == "excel_master"


# ── Harness 통합 ──
def test_ingest_master_sheet_end_to_end(db):
    _seed_channels(db)
    ws = _build_master_ws()

    result = ingest_master_sheet(db, ws)

    assert result.products_created == 2
    assert result.mappings_created == 7  # 행1(cafe24x2+naver+coupang=4) + 행2(cafe24x1+naver+coupang=3)
    assert result.mappings_conflicted == 0
    assert not result.integrity.unknown_labels

    products = db.query(ProductMaster).all()
    assert {p.product_name for p in products} == {"필름 A", "필름 B"}

    mappings = db.query(ProductChannelMapping).all()
    assert all(m.mapping_source == "excel_master" for m in mappings)


def test_ingest_master_sheet_idempotent_rerun(db):
    _seed_channels(db)
    ws = _build_master_ws()
    ingest_master_sheet(db, ws)

    ws2 = _build_master_ws()
    result2 = ingest_master_sheet(db, ws2)

    assert result2.products_created == 0
    assert result2.products_updated == 2
    assert result2.mappings_created == 0
    assert result2.mappings_updated == 7
    assert db.query(ProductMaster).count() == 2
    assert db.query(ProductChannelMapping).count() == 7


def test_ingest_master_sheet_reports_conflict_without_overwrite(db):
    channels = _seed_channels(db)
    counter = [1]
    other_product, _ = upsert_product_by_name(db, "다른상품", Decimal("1"), counter)
    db.add(ProductChannelMapping(
        product_id=other_product.id,
        channel_id=channels["CAFE24"].id,
        channel_product_id="P001",
        is_active=True,
        mapping_source="excel_master",
    ))
    db.commit()

    ws = _build_master_ws()
    result = ingest_master_sheet(db, ws)

    assert result.mappings_conflicted == 1
    assert result.integrity.mapping_conflicts
    row = db.query(ProductChannelMapping).filter_by(channel_id=channels["CAFE24"].id, channel_product_id="P001").one()
    assert row.product_id == other_product.id  # 충돌 상태 유지, 덮어쓰지 않음


# ── product_sync.py provenance 가드 ──
def test_product_sync_does_not_overwrite_excel_master_mapping(db):
    channels = _seed_channels(db)
    channel = channels["COUPANG_WING2"]

    master_a = ProductMaster(internal_sku="OHI-0001", product_name="A", cost_price=Decimal("100"))
    master_b = ProductMaster(internal_sku="SKU-B", product_name="B", cost_price=Decimal("200"))
    db.add_all([master_a, master_b])
    db.flush()

    db.add(ProductChannelMapping(
        product_id=master_a.id, channel_id=channel.id, channel_product_id="91576482085",
        is_active=True, mapping_source="excel_master",
    ))
    db.commit()

    detail = {"sellerProductId": 1, "sellerProductName": "상품", "statusName": "ON_SALE"}
    item = {"vendorItemId": 91576482085, "itemName": "옵션", "externalVendorSku": "SKU-B", "salePrice": 1000}

    changed = _maybe_upsert_mapping(db, channel, detail, item)
    db.commit()

    assert changed is False
    row = db.query(ProductChannelMapping).filter_by(
        channel_id=channel.id, channel_product_id="91576482085"
    ).one()
    assert row.product_id == master_a.id  # 자동동기화가 덮어쓰지 않음


def test_product_sync_still_creates_new_mapping_when_absent(db):
    channels = _seed_channels(db)
    channel = channels["COUPANG_WING2"]
    master_b = ProductMaster(internal_sku="SKU-B", product_name="B", cost_price=Decimal("200"))
    db.add(master_b)
    db.commit()

    detail = {"sellerProductId": 1, "sellerProductName": "상품", "statusName": "ON_SALE"}
    item = {"vendorItemId": 999, "itemName": "옵션", "externalVendorSku": "SKU-B", "salePrice": 1000}

    changed = _maybe_upsert_mapping(db, channel, detail, item)
    db.commit()

    assert changed is True
    row = db.query(ProductChannelMapping).filter_by(channel_id=channel.id, channel_product_id="999").one()
    assert row.product_id == master_b.id
    assert row.mapping_source == "auto_sync"


# ── 라우터 통합 (HTTP 레이어) ──
@pytest.fixture
def client():
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    TestingSession = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    _seed_channels(TestingSession())

    def _override_get_db():
        session = TestingSession()
        try:
            yield session
        finally:
            session.close()

    app.dependency_overrides[get_db] = _override_get_db
    yield TestClient(app)
    app.dependency_overrides.clear()


def _ws_to_xlsx_bytes(ws) -> bytes:
    buf = io.BytesIO()
    ws.parent.save(buf)
    buf.seek(0)
    return buf.read()


def test_upload_by_name_route_ingests_master_sheet(client):
    ws = _build_master_ws()
    xlsx_bytes = _ws_to_xlsx_bytes(ws)

    resp = client.post(
        "/api/products/upload-by-name",
        files={"file": ("master.xlsx", xlsx_bytes,
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert resp.status_code == 200
    body = resp.json()
    assert body["products_created"] == 2
    assert body["mappings_created"] == 7
    assert body["mappings_conflicted"] == 0
    assert body["unknown_labels"] == []


def test_upload_by_name_route_rejects_non_xlsx(client):
    resp = client.post(
        "/api/products/upload-by-name",
        files={"file": ("master.txt", b"not an excel file", "text/plain")},
    )
    assert resp.status_code == 400
