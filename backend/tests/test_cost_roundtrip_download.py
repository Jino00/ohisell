# test_cost_roundtrip_download.py — S3 「DB-생성 다운로드」 (계약 D-CPP-62 §4:122)
#
# 합격기준 원문: *"원가 메뉴 [다운로드] 클릭 → 받은 파일이 S2에서 확인한 표와 같은 행·열이고,
# 스냅샷 ID가 파일 안에 보인다"*
#
# ★이 파일이 지키려는 것은 **「같은 행·열」의 «같은»**이다. 그래서 단순히 「파일이 나온다」를
#   보지 않고, **같은 요청 안에서 `GET /api/cost/materials`(화면이 먹는 페이로드)를 함께 불러
#   파일 셀과 대조**한다. 이게 없으면 「다운로드가 `list_materials`를 버리고 자체 쿼리로
#   바뀌는」 변이가 살아남는다 — 헤더 파리티만 지키면 값이 갈라져도 아무도 안 죽는다.
#
# ★두 번째로 지키는 것은 **전건**이다. 부분집합 파일이 나가면 그 파일의 재업로드에서 빠진 종이
#   전부 S4 「사라짐」 묶음에 서고, 확인 클릭 한 번이 백여 종을 비활성화한다.
from __future__ import annotations

import json
from datetime import date
from decimal import Decimal as D
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base, get_db
from app.main import app
from app.models import (
    CostMaterial,
    CostMaterialPrice,
    CostRoundTripSnapshot,
    CostSetting,
)
from app.services.cost_menu import round_trip as RT

DOWNLOAD = "/api/cost/roundtrip/download"


@pytest.fixture
def client():
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    Base.metadata.create_all(engine)
    TestingSession = sessionmaker(bind=engine, autoflush=False, autocommit=False)

    def _override_get_db():
        db = TestingSession()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = _override_get_db
    tc = TestClient(app)
    tc.testing_session = TestingSession

    with TestingSession() as s:
        s.add_all(
            [
                CostSetting(key="standard_price_rule", value="latest", confirmed=True),
                # ① 원장 파생 단가 보유 — 출처 라벨 「원장」
                CostMaterial(
                    name="cleaning kit",
                    unit="ea",
                    category="부자재",
                    status="approved",
                    form_factor="bar",
                    part=None,
                    excel_ref_price=D("168.00"),
                    note="원가 정본 엑셀에서 구성 파싱으로 생성",
                ),
                # ② 사람이 넣은 단가(VAT 제외만) — inc는 ×1.1 파생
                CostMaterial(
                    name="패키지 (flip)",
                    unit="ea",
                    category="패키지",
                    status="approved",
                    form_factor="flip",
                    part=None,
                    excel_ref_price=D("171.00"),
                    note=None,
                ),
                # ③ ★단가가 하나도 없는 종 — 파일에서 **빈 칸**이지 0이 아니다
                CostMaterial(
                    name="6H 강화유리코팅 BLC_기본 (tablet)",
                    unit="매",
                    category="부자재",
                    status="unconfirmed",
                    form_factor="tablet",
                    part="외부",
                    excel_ref_price=D("1500.00"),
                    note="원가 정본 엑셀에서 구성 파싱으로 생성",
                ),
            ]
        )
        s.flush()
        ids = {m.name: m.id for m in s.query(CostMaterial).all()}
        s.add_all(
            [
                CostMaterialPrice(
                    material_id=ids["cleaning kit"],
                    source="ledger",
                    unit_price_ex_vat=D("190.82"),
                    unit_price_inc_vat=D("209.90"),
                    effective_date=date(2026, 8, 18),
                    linked_item_name="cleaning kits",
                ),
                CostMaterialPrice(
                    material_id=ids["패키지 (flip)"],
                    source="manual",
                    unit_price_ex_vat=D("171.00"),
                    unit_price_inc_vat=None,  # ← ×1.1 파생 경로
                    effective_date=date(2026, 8, 27),
                    note="엑셀 참고값 채택",
                ),
            ]
        )
        s.commit()

    yield tc
    app.dependency_overrides.clear()


def _sheets(resp):
    wb = load_workbook(BytesIO(resp.content))
    return wb


def _data_rows(ws):
    """헤더를 뺀 데이터 행. 셀 값 그대로(빈 칸은 None)."""
    return [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]


# ──────────────────────────────────────────────
# 열 — 스펙이 정본이고 파일이 그걸 따른다
# ──────────────────────────────────────────────
def test_열_스펙은_12개이고_순서가_고정이다():
    cols = RT.load_columns()
    assert [c["key"] for c in cols] == [
        "id",
        "name",
        "form_factor",
        "part",
        "unit",
        "price_ex",
        "price_inc",
        "vat_derived",
        "price_source",
        "effective_date",
        "excel_ref",
        "status_note",
    ]
    assert {c["key"] for c in cols if c["editable"]} == {
        "name",
        "form_factor",
        "part",
        "unit",
        "price_ex",
        "effective_date",
        "status_note",
    }


def test_읽기전용_열은_파일_헤더에_자백을_단다(client):
    resp = client.post(DOWNLOAD)
    assert resp.status_code == 200
    ws = _sheets(resp)[RT.SHEET_DATA]
    header = [c.value for c in ws[1]]

    # 읽기 전용 열 전부에 접미가 붙는다 — 시트 잠금이 아니라 «말하기»가 방어다(설계 Q3).
    assert header[6] == "단가 (VAT 포함) «읽기전용»"
    assert header[7] == "VAT 파생 «읽기전용»"
    assert header[8] == "단가 출처 «읽기전용»"
    assert header[10] == "엑셀 참고값 «읽기전용»"
    assert header[0] == "ID «읽기전용»"
    # 수정 가능 열엔 안 붙는다.
    assert header[1] == "이름"
    # ★12번째는 화면과 **다른 라벨**을 쓴다 — 그 사실을 스펙이 말하고 헤더가 자백한다.
    assert header[11] == "비고 (상태는 화면에서만)"


def test_헤더가_스펙에서_온다_스펙에서_열을_지우면_파일이_줄어든다(client, monkeypatch):
    """★변이 방어: 헤더를 코드에 박아 두면 스펙이 낡아도 파일은 멀쩡해 보인다."""
    full = RT.load_columns()
    monkeypatch.setattr(RT, "load_columns", lambda: full[:5])

    resp = client.post(DOWNLOAD)
    ws = _sheets(resp)[RT.SHEET_DATA]
    assert len([c.value for c in ws[1]]) == 5


# ──────────────────────────────────────────────
# 행 — 화면과 같은 페이로드에서 온다
# ──────────────────────────────────────────────
def test_파일_값이_화면_페이로드와_같다(client):
    """★이 테스트가 「자체 쿼리로 갈아타는」 변이를 죽인다.

    화면이 먹는 것과 **같은 응답**을 불러 파일 셀과 1:1로 맞춘다.
    """
    payload = client.get("/api/cost/materials").json()["items"]
    resp = client.post(DOWNLOAD)
    rows = _data_rows(_sheets(resp)[RT.SHEET_DATA])

    assert len(rows) == len(payload)
    for cells, m in zip(rows, payload):
        assert cells[0] == m["id"]
        assert cells[1] == m["name"]
        assert cells[2] == m["form_factor"]
        assert cells[4] == m["unit"]
        # 돈은 숫자 셀이다 — 문자열이면 사람이 엑셀에서 계산·정렬을 못 한다.
        if m["latest_price_ex_vat"] is None:
            assert cells[5] is None
        else:
            assert cells[5] == pytest.approx(float(m["latest_price_ex_vat"]))
        if m["excel_ref_price"] is None:
            assert cells[10] is None
        else:
            assert cells[10] == pytest.approx(float(m["excel_ref_price"]))
        # ★출처 낱말도 화면과 같은 값에서 온다 — 파일이 독자적으로 판정하지 않는다.
        expected_src = {"ledger": "원장", "manual": "등록가"}.get(
            m["latest_price_source"]
        )
        assert cells[8] == expected_src
        # ★비고만 싣는다 — 상태·배지는 안 싣는다(합성 셀은 왕복이 불가능하다).
        assert cells[11] == m["note"]


def test_단가_없는_종은_빈_칸이지_0이_아니다(client):
    resp = client.post(DOWNLOAD)
    rows = _data_rows(_sheets(resp)[RT.SHEET_DATA])
    target = [r for r in rows if r[1] == "6H 강화유리코팅 BLC_기본 (tablet)"][0]

    assert target[5] is None  # 단가 (VAT 제외)
    assert target[6] is None  # 단가 (VAT 포함)
    assert target[8] is None  # 단가 출처
    assert target[9] is None  # 발효일
    # 그런데 참고값은 있다 — 「참고값은 있고 단가는 없다」가 이 행의 사실이다.
    assert target[10] == pytest.approx(1500.0)


def test_출처_라벨이_화면과_같은_낱말이다():
    """★설계 문서 Q3는 「원장 / 수동」이라 적었지만 라이브 화면은 「원장 / **등록가**」다
    (`CostPage.tsx:127 priceSourceLabel` · D-CPP-56). 파일은 문서가 아니라 **화면**을 따른다
    (계약 §2 「왕복 파일의 모양은 항상 화면의 표를 따른다」).

    ★DB 픽스처가 아니라 payload로 직접 잰다 — `ledger` 단가는 조회 시점에 원장과 재대조되어
    (`materials.ledger_check`) 원장 행이 없으면 최신 자리를 못 차지한다. 그 규칙은 여기서
    잴 것이 아니고, 여기서 잴 것은 **낱말이 화면과 같은가**뿐이다.
    """
    cols = RT.load_columns()
    src = [c for c in cols if c["key"] == "price_source"][0]

    assert RT._label_of(src, "ledger") == "원장"
    assert RT._label_of(src, "manual") == "등록가"
    assert RT._label_of(src, None) is None


def test_vat_파생_표기(client):
    resp = client.post(DOWNLOAD)
    rows = _data_rows(_sheets(resp)[RT.SHEET_DATA])
    by_name = {r[1]: r for r in rows}

    # ex만 있는 종은 ×1.1로 만든 값이라고 파일이 말한다.
    # ★`True`가 아니라 `×1.1`이어야 한다 — 사람이 파일을 열었을 때 `True`는 아무 뜻도 없다.
    assert by_name["패키지 (flip)"][7] == "×1.1"
    # 단가가 아예 없는 종은 파생 여부를 말할 것도 없다.
    assert by_name["6H 강화유리코팅 BLC_기본 (tablet)"][7] is None


# ──────────────────────────────────────────────
# ★전건 — 필터가 파일에 닿지 못한다
# ──────────────────────────────────────────────
def test_필터_쿼리를_붙여도_전건이_내려간다(client):
    """★부분집합 파일이 나가면 재업로드에서 빠진 종이 전부 「사라짐」에 선다.

    엔드포인트는 필터 인자를 **갖지 않는다** — 붙여 보내도 행 수가 안 줄어야 한다.
    """
    total = len(client.get("/api/cost/materials").json()["items"])
    assert total == 3

    resp = client.post(f"{DOWNLOAD}?form_factor=tablet&no_price_only=true")
    rows = _data_rows(_sheets(resp)[RT.SHEET_DATA])
    assert len(rows) == total
    assert resp.headers["X-Snapshot-Rows"] == str(total)


# ──────────────────────────────────────────────
# 스냅샷 — 파일 안에 보이고, 같은 상태는 같은 ID다
# ──────────────────────────────────────────────
def test_스냅샷_ID가_파일_안에_보인다(client):
    resp = client.post(DOWNLOAD)
    wb = _sheets(resp)
    assert RT.SHEET_META in wb.sheetnames

    meta = {r[0]: r[1] for r in wb[RT.SHEET_META].iter_rows(min_row=2, values_only=True)}
    assert meta["스냅샷 ID"] == resp.headers["X-Snapshot-Id"]
    assert meta["스냅샷 ID"].startswith("CRT-")
    assert meta["행 수"] == 3
    assert "안내" in meta


def test_시각은_KST다_UTC가_아니다(client):
    """`server_default=func.now()`는 이 저장소에서 UTC다 — 파일에 9시간 어긋난 시각이 찍히면
    사람이 「어느 게 최신인가」를 바로 그 자리에서 다시 묻게 된다."""
    from app.utils.kst import kst_now

    before = kst_now()
    client.post(DOWNLOAD)
    after = kst_now()

    with client.testing_session() as s:
        snap = s.query(CostRoundTripSnapshot).one()
    assert before <= snap.created_at <= after


def test_같은_상태를_두_번_받으면_같은_스냅샷이다(client):
    a = client.post(DOWNLOAD)
    b = client.post(DOWNLOAD)
    assert a.headers["X-Snapshot-Id"] == b.headers["X-Snapshot-Id"]

    with client.testing_session() as s:
        assert s.query(CostRoundTripSnapshot).count() == 1


def test_값이_바뀌면_새_스냅샷이_선다(client):
    first = client.post(DOWNLOAD).headers["X-Snapshot-Id"]

    with client.testing_session() as s:
        mid = s.query(CostMaterial).filter_by(name="패키지 (flip)").one().id
        s.add(
            CostMaterialPrice(
                material_id=mid,
                source="manual",
                unit_price_ex_vat=D("180.00"),
                effective_date=date(2026, 8, 28),
            )
        )
        s.commit()

    second = client.post(DOWNLOAD).headers["X-Snapshot-Id"]
    assert first != second

    with client.testing_session() as s:
        assert s.query(CostRoundTripSnapshot).count() == 2


def test_스냅샷이_자기_열_스펙을_품는다(client):
    """구 파일을 업로드했을 때 라벨→키 매핑의 정본은 **그 파일의 스냅샷**이다."""
    client.post(DOWNLOAD)
    with client.testing_session() as s:
        snap = s.query(CostRoundTripSnapshot).one()
    spec = json.loads(snap.column_spec)
    assert [c["key"] for c in spec] == [c["key"] for c in RT.load_columns()]
    assert json.loads(snap.rows)[0].keys() == {c["key"] for c in spec}


def test_파일은_스냅샷에서만_만들어진다(client):
    """★DB를 다시 읽으면 그 사이의 변경이 조용히 끼어들어 유령 diff가 난다."""
    with client.testing_session() as s:
        snap = RT.build_snapshot(s)
        s.commit()
        snap_id = snap.id
        # 스냅샷을 찍은 «뒤» 상태를 바꾼다.
        s.query(CostMaterial).filter_by(name="cleaning kit").one().note = "바뀐 비고"
        s.commit()
        snap = s.query(CostRoundTripSnapshot).filter_by(id=snap_id).one()
        wb = load_workbook(RT.build_workbook(snap))

    rows = _data_rows(wb[RT.SHEET_DATA])
    by_name = {r[1]: r for r in rows}
    assert by_name["cleaning kit"][11] == "원가 정본 엑셀에서 구성 파싱으로 생성"
