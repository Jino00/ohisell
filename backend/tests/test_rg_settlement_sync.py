# test_rg_settlement_sync.py — RG 정산 수수료 파싱 fixture 테스트 (D-12)
# D-12: 머니코드 예외 — 파싱·부호·집계·dedup 라이브 self-verify 대신 committed fixture 테스트.
# 라이브 API 호출 없음. 응답 fixture는 S0 실측(2026-06-09) 기반.
import pytest
from decimal import Decimal
from datetime import date

from app.services.coupang.rg_settlement_sync import (
    _parse_status_response,
    _parse_date,
    _dec,
)
from app.clients.coupang.inbound import WingReadError
from app.clients.coupang.rg_settlement import CoupangWingRgSettlementClient


# ─── 헬퍼 테스트 ───────────────────────────────────────

def test_dec_normal():
    assert _dec(1234.56) == Decimal("1234.56")

def test_dec_none():
    assert _dec(None) == Decimal(0)

def test_dec_invalid():
    assert _dec("not_a_number") == Decimal(0)

def test_parse_date_utc_iso():
    d = _parse_date("2026-04-05T15:00:00Z")
    assert d == date(2026, 4, 6)  # UTC T15:00 = KST 다음날 00:00

def test_parse_date_none():
    assert _parse_date(None) is None

def test_parse_date_invalid():
    assert _parse_date("not-a-date") is None


# ─── KST→UTC 변환 테스트 ──────────────────────────────

def test_kst_date_to_utc_iso():
    result = CoupangWingRgSettlementClient._kst_date_to_utc_iso("2026-06-01")
    assert result == "2026-06-01T15:00:00.000Z"


# ─── 응답 파싱 테스트 ─────────────────────────────────

# S0 실측 fixture (2026-06-09 캡처)
_FIXTURE_SINGLE = {
    "settlementStatusReports": [
        {
            "settlementCycle": "WEEKLY",
            "settlementRatio": 70,
            "finalSettlementAmount": 0.00,
            "settlementGroupKey": "A01564720-2026-04-06-2026-04-12",
            "settlementDate": "2026-05-11T15:00:00Z",
            "settlementPeriodStartDate": "2026-04-05T15:00:00Z",  # → KST 2026-04-06
            "settlementPeriodEndDate": "2026-04-11T15:00:00Z",    # → KST 2026-04-12
            "settlementStatusReportDetail": {
                "totalTakeRateAmountWithVat": 2214,
                "totalFulfillmentFeeDeductionAmount": 2090,
                "totalStorageFeeDeductionAmount": 210,
                "totalWarehousingFeeDeductionAmount": 1210,
                "totalCreturnReverseShippingFeeDeductionAmount": 0,
                "totalVreturnHandlingFeeDeductionAmount": 0,
                "totalAdSalesDeductionAmount": 16510,
            }
        }
    ]
}


def test_parse_single_report():
    rows = _parse_status_response(_FIXTURE_SINGLE, "COUPANG_WING1")
    assert len(rows) == 7  # _FEE_FIELD_MAP 항목 수

    fee_map = {r.fee_type: r.amount for r in rows}
    assert fee_map["sale_fee"] == Decimal("2214")
    assert fee_map["delivery"] == Decimal("2090")   # 배송비(D-10: 풀필먼트 합계 아님)
    assert fee_map["storage"] == Decimal("210")
    assert fee_map["warehousing"] == Decimal("1210")
    assert fee_map["ad_sales"] == Decimal("16510")

    # 날짜 확인 (UTC T15:00 → KST 다음날)
    assert rows[0].recognition_date_from == date(2026, 4, 6)
    assert rows[0].recognition_date_to == date(2026, 4, 12)
    assert rows[0].account_key == "COUPANG_WING1"


def test_parse_empty_reports():
    rows = _parse_status_response({"settlementStatusReports": []}, "COUPANG_WING1")
    assert rows == []


def test_parse_missing_key_raises():
    with pytest.raises(WingReadError, match="스키마 드리프트"):
        _parse_status_response({"unexpected_key": []}, "COUPANG_WING1")


def test_parse_non_list_raises():
    with pytest.raises(WingReadError):
        _parse_status_response({"settlementStatusReports": "not_a_list"}, "COUPANG_WING1")


# ─── 부호 테스트 (취소/환급 음수) ──────────────────────

_FIXTURE_NEGATIVE = {
    "settlementStatusReports": [
        {
            "settlementPeriodStartDate": "2026-05-03T15:00:00Z",
            "settlementPeriodEndDate": "2026-05-09T15:00:00Z",
            "settlementStatusReportDetail": {
                "totalTakeRateAmountWithVat": -500,  # 환급(음수)
                "totalFulfillmentFeeDeductionAmount": 1000,
                "totalStorageFeeDeductionAmount": 0,
                "totalWarehousingFeeDeductionAmount": 0,
                "totalCreturnReverseShippingFeeDeductionAmount": 0,
                "totalVreturnHandlingFeeDeductionAmount": 0,
                "totalAdSalesDeductionAmount": 0,
            }
        }
    ]
}


def test_negative_amounts_preserved():
    rows = _parse_status_response(_FIXTURE_NEGATIVE, "COUPANG_WING1")
    fee_map = {r.fee_type: r.amount for r in rows}
    assert fee_map["sale_fee"] == Decimal("-500")  # 음수 그대로 저장
    assert fee_map["delivery"] == Decimal("1000")


# ─── 설치분(가지급70%+확정30%) dedup 테스트 ───────────────
# ★라이브 실측(2026-06-10, WING1·WING2 양 계정, 원칙22): 한 기간에 settlementRatio 70/30
#   설치분 리포트가 2개 온다. settlementRatio는 '지급액' 분할일 뿐, fee_type별 노출이 다르다:
#   take_rate만 기간 총액을 양쪽에 '동일값' 반복(102,950=102,950), 나머지(풀필먼트·보관·입출고·
#   반출처리)는 가지급분 full·확정분 0, ad_sales는 70/30 분할. 따라서 take_rate를 합산하면
#   이중계상(2배=205,900 버그). 정답: sale_fee 한정으로 동일값이면 1회만 계상(dedupe), 그 외는 합산.
#   공식 정산명세서 판매수수료=102,950과 일치.
_FIXTURE_INSTALLMENTS = {
    "settlementStatusReports": [
        {
            "settlementRatio": 70,  # 가지급
            "settlementPeriodStartDate": "2026-05-31T15:00:00Z",  # → KST 2026-06-01
            "settlementPeriodEndDate": "2026-06-06T15:00:00Z",    # → KST 2026-06-07
            "settlementStatusReportDetail": {
                "totalTakeRateAmountWithVat": 102950,             # 기간 총액
                "totalFulfillmentFeeDeductionAmount": 130599,      # full(이 설치분에 귀속)
                "totalStorageFeeDeductionAmount": 168,
                "totalWarehousingFeeDeductionAmount": 75489,
                "totalCreturnReverseShippingFeeDeductionAmount": 0,
                "totalVreturnHandlingFeeDeductionAmount": 0,
                "totalAdSalesDeductionAmount": 0,
            }
        },
        {
            "settlementRatio": 30,  # 확정 — take_rate 반복, 풀필먼트 0
            "settlementPeriodStartDate": "2026-05-31T15:00:00Z",
            "settlementPeriodEndDate": "2026-06-06T15:00:00Z",
            "settlementStatusReportDetail": {
                "totalTakeRateAmountWithVat": 102950,             # 동일값 반복(분할 아님)
                "totalFulfillmentFeeDeductionAmount": 0,
                "totalStorageFeeDeductionAmount": 0,
                "totalWarehousingFeeDeductionAmount": 0,
                "totalCreturnReverseShippingFeeDeductionAmount": 0,
                "totalVreturnHandlingFeeDeductionAmount": 0,
                "totalAdSalesDeductionAmount": 0,
            }
        }
    ]
}


def test_installment_reports_deduped_not_summed():
    """가지급70%+확정30% 설치분은 dedupe(합산 금지). take_rate 2배 버그 회귀 가드."""
    rows = _parse_status_response(_FIXTURE_INSTALLMENTS, "COUPANG_WING1")
    fee_map = {r.fee_type: r.amount for r in rows}
    # ★합산(205,900) 아님 — 공식 명세서 판매수수료 102,950과 일치
    assert fee_map["sale_fee"] == Decimal("102950")
    assert fee_map["sale_fee"] != Decimal("205900")  # 2배 버그 명시적 가드
    # 풀필먼트: full(한쪽)+0(다른쪽) → full 채택
    assert fee_map["delivery"] == Decimal("130599")
    assert fee_map["warehousing"] == Decimal("75489")
    assert fee_map["storage"] == Decimal("168")


def test_installment_dedup_preserves_negative_refund():
    """설치분 dedup이 음수 환급 부호를 보존(절대값 최대 채택)."""
    fixture = {
        "settlementStatusReports": [
            {
                "settlementPeriodStartDate": "2026-05-03T15:00:00Z",
                "settlementPeriodEndDate": "2026-05-09T15:00:00Z",
                "settlementStatusReportDetail": {
                    "totalTakeRateAmountWithVat": -500,  # 환급(음수)
                    "totalFulfillmentFeeDeductionAmount": 0,
                    "totalStorageFeeDeductionAmount": 0,
                    "totalWarehousingFeeDeductionAmount": 0,
                    "totalCreturnReverseShippingFeeDeductionAmount": 0,
                    "totalVreturnHandlingFeeDeductionAmount": 0,
                    "totalAdSalesDeductionAmount": 0,
                },
            },
            {
                "settlementPeriodStartDate": "2026-05-03T15:00:00Z",  # 같은 기간 설치분
                "settlementPeriodEndDate": "2026-05-09T15:00:00Z",
                "settlementStatusReportDetail": {
                    "totalTakeRateAmountWithVat": -500,  # 동일 음수 반복
                    "totalFulfillmentFeeDeductionAmount": 0,
                    "totalStorageFeeDeductionAmount": 0,
                    "totalWarehousingFeeDeductionAmount": 0,
                    "totalCreturnReverseShippingFeeDeductionAmount": 0,
                    "totalVreturnHandlingFeeDeductionAmount": 0,
                    "totalAdSalesDeductionAmount": 0,
                },
            },
        ]
    }
    rows = _parse_status_response(fixture, "COUPANG_WING1")
    fee_map = {r.fee_type: r.amount for r in rows}
    assert fee_map["sale_fee"] == Decimal("-500")  # 합산(-1000) 아님, 부호 보존


def test_installment_ad_sales_split_summed():
    """광고비(ad_sales)는 설치분 간 '다른값'으로 70/30 분할 노출 → 합산해야 기간 총액(라이브 실측).

    ★take_rate(동일값=반복총액)와 달리 ad_sales는 설치분마다 값이 다름(16,510 vs 7,076).
    동일값만 dedupe하고 다른값은 합산하므로 ad는 23,586으로 정확히 합산된다.
    같은 fixture에서 take_rate(동일값)는 dedupe(1회)됨을 함께 검증.
    """
    fixture = {
        "settlementStatusReports": [
            {
                "settlementRatio": 70,
                "settlementPeriodStartDate": "2026-03-22T15:00:00Z",
                "settlementPeriodEndDate": "2026-03-28T15:00:00Z",
                "settlementStatusReportDetail": {
                    "totalTakeRateAmountWithVat": 3321,            # 동일값(반복 총액)
                    "totalFulfillmentFeeDeductionAmount": 6270,     # full
                    "totalStorageFeeDeductionAmount": 219,
                    "totalWarehousingFeeDeductionAmount": 3630,
                    "totalCreturnReverseShippingFeeDeductionAmount": 0,
                    "totalVreturnHandlingFeeDeductionAmount": 0,
                    "totalAdSalesDeductionAmount": 16510,           # 70% 분할
                },
            },
            {
                "settlementRatio": 30,
                "settlementPeriodStartDate": "2026-03-22T15:00:00Z",
                "settlementPeriodEndDate": "2026-03-28T15:00:00Z",
                "settlementStatusReportDetail": {
                    "totalTakeRateAmountWithVat": 3321,            # 동일값 반복
                    "totalFulfillmentFeeDeductionAmount": 0,
                    "totalStorageFeeDeductionAmount": 0,
                    "totalWarehousingFeeDeductionAmount": 0,
                    "totalCreturnReverseShippingFeeDeductionAmount": 0,
                    "totalVreturnHandlingFeeDeductionAmount": 0,
                    "totalAdSalesDeductionAmount": 7076,            # 30% 분할
                },
            },
        ]
    }
    rows = _parse_status_response(fixture, "COUPANG_WING1")
    fee_map = {r.fee_type: r.amount for r in rows}
    assert fee_map["sale_fee"] == Decimal("3321")     # 동일값 → dedupe(합산 6642 아님)
    assert fee_map["ad_sales"] == Decimal("23586")    # 다른값 → 합산(16510+7076)
    assert fee_map["delivery"] == Decimal("6270")     # full+0 → 합산=full
    assert fee_map["warehousing"] == Decimal("3630")
    assert fee_map["storage"] == Decimal("219")


def test_installment_nonsalefee_equal_values_summed_not_deduped():
    """dedupe는 sale_fee에만 적용(codex P2). 비-sale_fee가 우연히 동일값이어도 합산(과소계상 방지).

    ★sale_fee만 '반복 총액'이 라이브로 입증됨. ad_sales 등이 우연히 같은 값(예: 50/50 분할 1000+1000)
    이면 전역 dedupe는 1000으로 과소계상하지만, sale_fee 한정 규칙은 2000으로 정확히 합산한다.
    """
    fixture = {
        "settlementStatusReports": [
            {
                "settlementPeriodStartDate": "2026-03-22T15:00:00Z",
                "settlementPeriodEndDate": "2026-03-28T15:00:00Z",
                "settlementStatusReportDetail": {
                    "totalTakeRateAmountWithVat": 0,
                    "totalFulfillmentFeeDeductionAmount": 0,
                    "totalStorageFeeDeductionAmount": 0,
                    "totalWarehousingFeeDeductionAmount": 0,
                    "totalCreturnReverseShippingFeeDeductionAmount": 0,
                    "totalVreturnHandlingFeeDeductionAmount": 0,
                    "totalAdSalesDeductionAmount": 1000,   # 우연히 동일값(50/50 가정)
                },
            },
            {
                "settlementPeriodStartDate": "2026-03-22T15:00:00Z",
                "settlementPeriodEndDate": "2026-03-28T15:00:00Z",
                "settlementStatusReportDetail": {
                    "totalTakeRateAmountWithVat": 0,
                    "totalFulfillmentFeeDeductionAmount": 0,
                    "totalStorageFeeDeductionAmount": 0,
                    "totalWarehousingFeeDeductionAmount": 0,
                    "totalCreturnReverseShippingFeeDeductionAmount": 0,
                    "totalVreturnHandlingFeeDeductionAmount": 0,
                    "totalAdSalesDeductionAmount": 1000,   # 동일값
                },
            },
        ]
    }
    rows = _parse_status_response(fixture, "COUPANG_WING1")
    fee_map = {r.fee_type: r.amount for r in rows}
    assert fee_map["ad_sales"] == Decimal("2000")  # 비-sale_fee 동일값 → 합산(dedupe 1000 아님)


# ─── 누락 필드 방어 테스트 (D-13) ──────────────────────

_FIXTURE_MISSING_FIELDS = {
    "settlementStatusReports": [
        {
            "settlementPeriodStartDate": "2026-05-03T15:00:00Z",
            "settlementPeriodEndDate": "2026-05-09T15:00:00Z",
            "settlementStatusReportDetail": {
                "totalTakeRateAmountWithVat": 1000,
                # 나머지 필드 누락 (스키마 변동 시뮬레이션)
            }
        }
    ]
}


def test_missing_fields_defaults_to_zero():
    rows = _parse_status_response(_FIXTURE_MISSING_FIELDS, "COUPANG_WING1")
    fee_map = {r.fee_type: r.amount for r in rows}
    assert fee_map["sale_fee"] == Decimal("1000")
    assert fee_map["delivery"] == Decimal("0")   # 누락 → 0
    assert fee_map["storage"] == Decimal("0")


# ─── D-10 basis 잠금 테스트 (라이브 확정 2026-06-09) ──────

def test_search_date_type_default_is_SALES():
    """D-10: status/api 기본 searchDateType=SALES(매출인식일). 클라이언트 시그니처 잠금."""
    import inspect
    from app.clients.coupang.rg_settlement import _SEARCH_DATE_SALES
    sig = inspect.signature(CoupangWingRgSettlementClient.get_settlement_status)
    assert sig.parameters["search_date_type"].default == _SEARCH_DATE_SALES == "SALES"


# 라이브 실측 fixture (2026-06-09 WING1 04-06 리포트, 50필드 중 핵심 + 이월필드 포함)
_FIXTURE_LIVE_WITH_CARRYOVER = {
    "settlementStatusReports": [
        {
            "settlementPeriodStartDate": "2026-04-05T15:00:00Z",
            "settlementPeriodEndDate": "2026-04-11T15:00:00Z",
            "settlementStatusReportDetail": {
                # 발생비용(f) 컴포넌트 — 우리가 적재하는 7개
                "totalTakeRateAmountWithVat": 2214,
                "totalFulfillmentFeeDeductionAmount": 2090,   # 배송비
                "totalStorageFeeDeductionAmount": 210,
                "totalWarehousingFeeDeductionAmount": 1210,
                "totalCreturnReverseShippingFeeDeductionAmount": 0,
                "totalVreturnHandlingFeeDeductionAmount": 0,
                "totalAdSalesDeductionAmount": 16510,
                # 이월(g)/과거차감 — D-10: 발생 f에 섞이면 안 됨. 적재 제외 확인용.
                "totalCarryOverSettlementDeductionAmount": 114.0,
                "totalPastCfsDeductionAmount": -3396.0,
                "totalFinalSettlementAmount": 0.0,
            }
        }
    ]
}


def test_carryover_fields_excluded_from_accrual():
    """D-10: amount=발생비용(f). 이월(g)·과거차감·최종지급액 필드는 적재 안 함(7개 컴포넌트만)."""
    rows = _parse_status_response(_FIXTURE_LIVE_WITH_CARRYOVER, "COUPANG_WING1")
    assert len(rows) == 7  # 이월필드 3개는 무시
    fee_types = {r.fee_type for r in rows}
    # 이월/과거/최종 관련 fee_type 없음
    assert "carryover" not in fee_types
    assert all("final" not in ft and "past" not in ft for ft in fee_types)
    fee_map = {r.fee_type: r.amount for r in rows}
    assert fee_map["delivery"] == Decimal("2090")   # 발생 배송비(이월 미반영)


def test_fulfillment_components_distinct_no_double_count():
    """D-10 라이브 검증: 풀필먼트 J = 배송(delivery)+입출고(warehousing)+보관(storage), 세 값 독립.
    레퍼런스 17 §7(06-01~07): 배송 130,599 + 입출고 75,489 + 보관 168 = J 206,256."""
    detail = {
        "totalFulfillmentFeeDeductionAmount": 130599,  # 배송비
        "totalWarehousingFeeDeductionAmount": 75489,   # 입출고비
        "totalStorageFeeDeductionAmount": 168,         # 보관비
        "totalTakeRateAmountWithVat": 0,
        "totalCreturnReverseShippingFeeDeductionAmount": 0,
        "totalVreturnHandlingFeeDeductionAmount": 0,
        "totalAdSalesDeductionAmount": 0,
    }
    fixture = {"settlementStatusReports": [{
        "settlementPeriodStartDate": "2026-05-31T15:00:00Z",
        "settlementPeriodEndDate": "2026-06-06T15:00:00Z",
        "settlementStatusReportDetail": detail,
    }]}
    rows = _parse_status_response(fixture, "COUPANG_WING1")
    fee_map = {r.fee_type: r.amount for r in rows}
    j = fee_map["delivery"] + fee_map["warehousing"] + fee_map["storage"]
    assert j == Decimal("206256")  # 레퍼런스 §7 검산 일치


# ─── D-11 광고비 dedup 규칙 테스트 ────────────────────────

def test_rg_ad_dedup_pure_rule():
    """D-11: rg_ad_spend_to_exclude — 2P(RG)분만 합산, 3P/Retail 제외."""
    from app.services.coupang.intelligence import rg_ad_spend_to_exclude, RG_AD_SELL_TYPE
    assert RG_AD_SELL_TYPE == "2P"
    rows = [
        ("3P", Decimal("10000")),   # 윙 — 제외 안 함
        ("2P", Decimal("3000")),    # RG — 제외 대상
        ("2P", Decimal("500")),     # RG — 제외 대상
        ("Retail", Decimal("700")), # 로켓배송 — 제외 안 함
    ]
    assert rg_ad_spend_to_exclude(rows) == Decimal("3500")


def test_rg_ad_dedup_empty():
    """D-11: 2P 행이 없으면 0(현재 prod 상태 — 겹침 없음)."""
    from app.services.coupang.intelligence import rg_ad_spend_to_exclude
    assert rg_ad_spend_to_exclude([("3P", Decimal("585670"))]) == Decimal("0")


def test_rg_ad_dedup_normalizes_whitespace_and_none():
    """D-11 견고성(Codex 지적4a): 공백/None sell_type 방어."""
    from app.services.coupang.intelligence import rg_ad_spend_to_exclude
    rows = [(" 2P ", Decimal("100")), ("2P", Decimal("50")), (None, Decimal("999"))]
    assert rg_ad_spend_to_exclude(rows) == Decimal("150")


# ─── 대조 카드 reconcile guard 테스트 (Codex S5 지적1) ────

def test_rg_breakdown_reconciles_clean():
    """정상 entry: sale_fee + 풀필먼트 + return + ad_sales = total, other=0."""
    from app.services.coupang.intelligence import _rg_account_breakdown
    v = {
        "sale_fee": Decimal("2214"), "delivery": Decimal("2090"),
        "warehousing": Decimal("1210"), "storage": Decimal("210"),
        "return_shipping": Decimal("0"), "return_handling": Decimal("0"),
        "ad_sales": Decimal("16510"),
        "total": Decimal("22234"),  # 2214+2090+1210+210+16510
    }
    b = _rg_account_breakdown("COUPANG_WING1", v)
    assert b["fulfillment"] == Decimal("3510")  # 2090+1210+210
    assert b["other"] == Decimal("0")
    assert b["sale_fee"] + b["fulfillment"] + b["return_fee"] + b["ad_sales"] + b["other"] == b["total"]


def test_rg_breakdown_legacy_fulfillment_surfaces_in_other():
    """legacy 'fulfillment' fee_type(미매핑)이 total에 섞여도 'other'로 노출되어 reconcile 유지."""
    from app.services.coupang.intelligence import _rg_account_breakdown
    v = {
        "sale_fee": Decimal("1000"), "delivery": Decimal("500"),
        "fulfillment": Decimal("777"),  # ★legacy 키(미매핑) — total엔 포함
        "total": Decimal("2277"),       # 1000+500+777
    }
    b = _rg_account_breakdown("COUPANG_WING1", v)
    assert b["fulfillment"] == Decimal("500")  # delivery만(legacy 'fulfillment' 키는 미반영)
    assert b["other"] == Decimal("777")        # legacy 잔액 가시화(silent drop·중복 없음)
    # 라인합 + other == total
    assert b["sale_fee"] + b["fulfillment"] + b["return_fee"] + b["ad_sales"] + b["other"] == b["total"]


# ═══════════════════════════════════════════════════════════════════════
# S6 — 옵션 단위 엑셀 파서 테스트 (D-12 머니코드 fixture)
# ═══════════════════════════════════════════════════════════════════════
# 합성 엑셀로 파싱·집계·검산·시트매핑·정규화·컬럼위치독립성을 검증.
# 실제 샘플 엑셀(A01564720-WAREHOUSING_SHIPPING)은 §8-1 라이브 self-verify로 별도 확인(원칙22).
import io
import openpyxl
from app.services.coupang.rg_settlement_sync import (
    parse_settlement_xlsx,
    ingest_settlement_xlsx,
    _norm_option_id,
    _parse_excel_date,
)
from app.models import Base, CoupangRgSettlementFee
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker


def _write_sheet(ws, title, fee_label, rows, *, cost_col=25, summary_sum=None, summary_period=None):
    """실제 종류별 엑셀 구조 재현.
    row2 제목 / row3 요약헤더 / row4 요약데이터 / row7 메인헤더 / row8 서브헤더 / row9+ 상세.
    rows: list of (oid, rec_date_str, period_end_str, A, B, AB). period_end_str=""면 상세 종료일 빈칸.
    cost_col: '할인적용가(A-B)' 컬럼 위치(시트별 다름 → 위치 독립성 테스트용).
    summary_sum: 요약합계 강제값(None=Σ상세 일치). mismatch 케이스용.
    summary_period: 요약 종료일 강제값(None=rows[0] 종료일). 상세 종료일 fallback 테스트용.
    """
    ws.cell(2, 1, title)
    # 요약 (row3 헤더, row4 데이터)
    ws.cell(3, 1, "정산주기(종료일)"); ws.cell(3, 2, f"{fee_label} 합계")
    ws.cell(3, 3, "세액"); ws.cell(3, 4, "최종비용")
    sum_ab = sum(r[5] for r in rows) if summary_sum is None else summary_sum
    tax = round(sum_ab * 0.1)
    pend = summary_period if summary_period is not None else (rows[0][2] if rows else "2026-06-07")
    ws.cell(4, 1, pend); ws.cell(4, 2, sum_ab); ws.cell(4, 3, tax); ws.cell(4, 4, sum_ab + tax)
    # 상세 메인헤더 (row7)
    ws.cell(7, 2, "정산주기(종료일)"); ws.cell(7, 5, "매출인식일"); ws.cell(7, 11, "옵션ID")
    ws.cell(7, cost_col - 2, f"{fee_label} (VAT 별도)")  # 병합 메인 라벨
    # 서브헤더 (row8): 발생비용A / 할인가B / 할인적용가(A-B)
    ws.cell(8, cost_col - 2, "발생비용(A)"); ws.cell(8, cost_col - 1, "할인가(B)")
    ws.cell(8, cost_col, "할인적용가(A-B)")
    # 상세 데이터 (row9+)
    for i, (oid, rec, pe, A, B, AB) in enumerate(rows):
        rr = 9 + i
        ws.cell(rr, 2, pe); ws.cell(rr, 5, rec); ws.cell(rr, 11, oid)
        ws.cell(rr, cost_col - 2, A); ws.cell(rr, cost_col - 1, B); ws.cell(rr, cost_col, AB)


def _build_xlsx(sheet_specs):
    """sheet_specs: list of (sheet_name, fee_label, rows[, cost_col, summary_sum, summary_period]). → bytes."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for spec in sheet_specs:
        name, fee_label, rows = spec[0], spec[1], spec[2]
        cost_col = spec[3] if len(spec) > 3 else 25
        summary_sum = spec[4] if len(spec) > 4 else None
        summary_period = spec[5] if len(spec) > 5 else None
        ws = wb.create_sheet(name)
        _write_sheet(ws, f"{fee_label} 정산 내역", fee_label, rows,
                     cost_col=cost_col, summary_sum=summary_sum, summary_period=summary_period)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# 입출고비: 같은 옵션 2행(합산 검증) + 다른 옵션 1행. 배송비: 컬럼 위치 다름(cost_col=24).
_WH_ROWS = [
    ("95521944483", "2026-06-04", "2026-06-07", 3300, 1050, 2250),
    ("95521944483", "2026-06-04", "2026-06-07", 1650, 525, 1125),  # 같은 옵션 → 합산 3375
    ("95521944484", "2026-06-05", "2026-06-07", 1650, 525, 1125),
]
_DEL_ROWS = [
    ("95521944483", "2026-06-04", "2026-06-07", 2200, 0, 2200),
    ("95521944484", "2026-06-05", "2026-06-07", 2200, 225, 1975),
]


def _db():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    return sessionmaker(bind=engine)()


# ─── 정규화 헬퍼 ──────────────────────────────────────
def test_norm_option_id_float():
    assert _norm_option_id(95521944483.0) == "95521944483"

def test_norm_option_id_dash_empty():
    assert _norm_option_id("-") == ""
    assert _norm_option_id("") == ""
    assert _norm_option_id(None) == ""

def test_norm_option_id_str():
    assert _norm_option_id(" 95521944483 ") == "95521944483"

def test_parse_excel_date_str():
    assert _parse_excel_date("2026-06-07") == date(2026, 6, 7)

def test_parse_excel_date_bad():
    assert _parse_excel_date("not-a-date") is None
    assert _parse_excel_date(None) is None


# ─── 파서 집계·검산 ───────────────────────────────────
def test_parse_xlsx_two_sheets_fee_type_map():
    content = _build_xlsx([("입출고비", "쿠팡풀필먼트서비스(CFS) 입출고비", _WH_ROWS),
                           ("배송비", "쿠팡풀필먼트서비스(CFS) 배송비", _DEL_ROWS, 24)])
    parsed = parse_settlement_xlsx(content)
    by_type = {s["fee_type"]: s for s in parsed["sheets"]}
    assert set(by_type) == {"warehousing", "delivery"}
    assert parsed["sheets_skipped"] == []

def test_parse_xlsx_same_option_summed():
    """같은 옵션 2행 → 합산(3375). 집계 grain=(옵션ID, 정산주기끝)."""
    content = _build_xlsx([("입출고비", "입출고비", _WH_ROWS)])
    s = parse_settlement_xlsx(content)["sheets"][0]
    assert s["options"][("95521944483", date(2026, 6, 7))] == Decimal("3375")
    assert s["options"][("95521944484", date(2026, 6, 7))] == Decimal("1125")
    assert len(s["options"]) == 2  # 고유 옵션 2개

def test_parse_xlsx_reconcile_match():
    content = _build_xlsx([("입출고비", "입출고비", _WH_ROWS)])
    s = parse_settlement_xlsx(content)["sheets"][0]
    assert s["sum_detail"] == Decimal("4500")  # 2250+1125+1125
    assert s["reconcile"]["match"] is True
    assert s["reconcile"]["diff"] == Decimal("0")

def test_parse_xlsx_reconcile_mismatch_flagged():
    """요약합계를 Σ상세와 다르게 → match=False (검산 가드, 원칙22)."""
    content = _build_xlsx([("입출고비", "입출고비", _WH_ROWS, 25, 9999)])
    s = parse_settlement_xlsx(content)["sheets"][0]
    assert s["reconcile"]["match"] is False
    assert s["reconcile"]["diff"] == Decimal("4500") - Decimal("9999")

def test_parse_xlsx_column_position_independence():
    """배송비처럼 cost 컬럼 위치가 달라도(col24) 헤더명 매핑으로 정상 파싱."""
    content = _build_xlsx([("배송비", "배송비", _DEL_ROWS, 24)])
    s = parse_settlement_xlsx(content)["sheets"][0]
    assert s["fee_type"] == "delivery"
    assert s["options"][("95521944483", date(2026, 6, 7))] == Decimal("2200")
    assert s["sum_detail"] == Decimal("4175")  # 2200+1975

def test_parse_xlsx_unknown_sheet_skipped():
    """미지 시트명 → skip(중단 안 함, D-13)."""
    content = _build_xlsx([("입출고비", "입출고비", _WH_ROWS),
                           ("알수없는시트", "기타", _WH_ROWS)])
    parsed = parse_settlement_xlsx(content)
    assert [s["fee_type"] for s in parsed["sheets"]] == ["warehousing"]
    assert parsed["sheets_skipped"] == ["알수없는시트"]


# ─── ingest DB upsert (in-memory) ─────────────────────
def test_ingest_creates_option_rows():
    db = _db()
    content = _build_xlsx([("입출고비", "입출고비", _WH_ROWS),
                           ("배송비", "배송비", _DEL_ROWS, 24)])
    res = ingest_settlement_xlsx(db, "COUPANG_WING1", content)
    assert res["status"] == "ok"
    # 입출고 옵션 2 + 배송 옵션 2 = 4 upsert
    assert res["upserted"] == 4
    rows = db.query(CoupangRgSettlementFee).all()
    assert len(rows) == 4
    assert all(r.vendor_item_id != "" for r in rows)
    wh = db.query(CoupangRgSettlementFee).filter_by(
        fee_type="warehousing", vendor_item_id="95521944483").one()
    assert wh.amount == Decimal("3375")  # 합산값(VAT前)

def test_ingest_coexists_with_account_row():
    """status/api 계정 row(vendor_item_id='')와 옵션 row 공존 — 서로 안 덮어씀.
    + _resolve_period_start가 계정 row의 from을 차용(grain 정합)."""
    db = _db()
    db.add(CoupangRgSettlementFee(
        account_key="COUPANG_WING1", recognition_date_from=date(2026, 6, 1),
        recognition_date_to=date(2026, 6, 7), fee_type="warehousing",
        vendor_item_id="", raw_type="totalWarehousingFeeDeductionAmount",
        amount=Decimal("75489")))  # status/api VAT후
    db.commit()
    content = _build_xlsx([("입출고비", "입출고비", _WH_ROWS)])
    ingest_settlement_xlsx(db, "COUPANG_WING1", content)
    # 계정 row 불변
    acct = db.query(CoupangRgSettlementFee).filter_by(
        fee_type="warehousing", vendor_item_id="").one()
    assert acct.amount == Decimal("75489")
    # 옵션 row from = 계정 row from(2026-06-01) 차용 (폴백 -6d 아님)
    opt = db.query(CoupangRgSettlementFee).filter_by(
        fee_type="warehousing", vendor_item_id="95521944483").one()
    assert opt.recognition_date_from == date(2026, 6, 1)
    assert opt.recognition_date_to == date(2026, 6, 7)

def test_ingest_idempotent_upsert():
    """같은 엑셀 두 번 ingest → 중복 행 없음(upsert)."""
    db = _db()
    content = _build_xlsx([("입출고비", "입출고비", _WH_ROWS)])
    ingest_settlement_xlsx(db, "COUPANG_WING1", content)
    ingest_settlement_xlsx(db, "COUPANG_WING1", content)
    rows = db.query(CoupangRgSettlementFee).filter_by(fee_type="warehousing").all()
    assert len(rows) == 2  # 고유 옵션 2개만(중복 없음)

def test_ingest_period_start_fallback():
    """status/api 계정 row 없으면 달력 규칙 폴백(월~일+월경계분할).
    06-07은 온전한 주(월 06-01~일 06-07)라 달력 규칙=구 -6d와 값이 일치(06-01)."""
    db = _db()
    content = _build_xlsx([("입출고비", "입출고비", _WH_ROWS)])
    ingest_settlement_xlsx(db, "COUPANG_WING1", content)
    opt = db.query(CoupangRgSettlementFee).filter_by(
        fee_type="warehousing", vendor_item_id="95521944484").one()
    assert opt.recognition_date_to == date(2026, 6, 7)
    assert opt.recognition_date_from == date(2026, 6, 1)  # 6/7=일요일, 월~일 주 = 6/1

def test_ingest_empty_when_all_unknown():
    db = _db()
    content = _build_xlsx([("알수없는시트", "기타", _WH_ROWS)])
    res = ingest_settlement_xlsx(db, "COUPANG_WING1", content)
    assert res["status"] == "empty"
    assert res["upserted"] == 0


# ═══════════════════════════════════════════════════════════════════════
# 층1 — ingest_status_payload (Mac 상주 브라우저 status/api push, 쿠키 무관 경로)
# ═══════════════════════════════════════════════════════════════════════
# 기존 sync_rg_settlement upsert 루프를 _upsert_account_rows로 추출한 뒤 쿠키 없이 raw dict를
# 그대로 적재하는 경로. 계정 단위(vendor_item_id="")만 건드리고 옵션 row는 불변.
import copy as _copy
from app.services.coupang.rg_settlement_sync import ingest_status_payload, _upsert_account_rows


def test_ingest_status_payload_creates_account_rows():
    """raw status/api dict → 계정 단위(vendor_item_id="") 7행 적재 + 반환 형식."""
    db = _db()
    res = ingest_status_payload(db, "COUPANG_WING1", _FIXTURE_SINGLE)
    assert res == {"synced": 7, "account_key": "COUPANG_WING1", "status": "ok"}
    rows = db.query(CoupangRgSettlementFee).all()
    assert len(rows) == 7                                   # _FEE_FIELD_MAP 7개
    assert all(r.vendor_item_id == "" for r in rows)        # 계정 단위 sentinel
    fee_map = {r.fee_type: r.amount for r in rows}
    assert fee_map["sale_fee"] == Decimal("2214")
    assert fee_map["ad_sales"] == Decimal("16510")
    assert rows[0].recognition_date_from == date(2026, 4, 6)
    assert rows[0].recognition_date_to == date(2026, 4, 12)


def test_ingest_status_payload_idempotent_update():
    """재호출(값 변경) → 중복 행 생성 없이 update(upsert grain 동일)."""
    db = _db()
    ingest_status_payload(db, "COUPANG_WING1", _FIXTURE_SINGLE)
    modified = _copy.deepcopy(_FIXTURE_SINGLE)
    modified["settlementStatusReports"][0]["settlementStatusReportDetail"][
        "totalTakeRateAmountWithVat"] = 9999
    res = ingest_status_payload(db, "COUPANG_WING1", modified)
    assert res["synced"] == 7
    rows = db.query(CoupangRgSettlementFee).all()
    assert len(rows) == 7                                   # 중복 생성 없음
    sale = db.query(CoupangRgSettlementFee).filter_by(
        fee_type="sale_fee", vendor_item_id="").one()
    assert sale.amount == Decimal("9999")                   # update 반영


def test_ingest_status_payload_parse_error_propagates():
    """스키마 드리프트(WingReadError)는 전파(라우터가 422로 변환)."""
    db = _db()
    with pytest.raises(WingReadError, match="스키마 드리프트"):
        ingest_status_payload(db, "COUPANG_WING1", {"unexpected_key": []})


def test_ingest_status_payload_leaves_option_rows_untouched():
    """계정 row(vendor_item_id='')만 upsert — 옵션 row(vendor_item_id=옵션ID)는 안 덮어씀(S6 가드)."""
    db = _db()
    db.add(CoupangRgSettlementFee(
        account_key="COUPANG_WING1", recognition_date_from=date(2026, 4, 6),
        recognition_date_to=date(2026, 4, 12), fee_type="sale_fee",
        vendor_item_id="95521944483", raw_type="xlsx:sale_fee", amount=Decimal("111")))
    db.commit()
    ingest_status_payload(db, "COUPANG_WING1", _FIXTURE_SINGLE)
    opt = db.query(CoupangRgSettlementFee).filter_by(
        fee_type="sale_fee", vendor_item_id="95521944483").one()
    assert opt.amount == Decimal("111")                    # 옵션 row 불변
    acct = db.query(CoupangRgSettlementFee).filter_by(
        fee_type="sale_fee", vendor_item_id="").one()
    assert acct.amount == Decimal("2214")                  # 계정 row 신규 적재


def test_upsert_account_rows_no_commit_returns_count():
    """_upsert_account_rows는 commit하지 않고 건수만 반환(commit은 호출자 책임)."""
    db = _db()
    rows = _parse_status_response(_FIXTURE_SINGLE, "COUPANG_WING1")
    n = _upsert_account_rows(db, rows)
    assert n == 7
    db.rollback()                                          # 미commit이면 롤백으로 사라짐
    assert db.query(CoupangRgSettlementFee).count() == 0


# ═══════════════════════════════════════════════════════════════════════
# 층3 (§3, PLAN_pipeline-freshness-3layer) — 정산주기 시작일 달력 규칙 (머니코드)
# ═══════════════════════════════════════════════════════════════════════
# _expected_period_start: 월~일 주별 + 달력 월 경계 분할. prod 계정 row 18개 전수 18/18 검증
# (2026-07-17 라이브). 구 -6d 폴백은 월 경계 주간(예: 06-30, 07-05)에서 반드시 틀려 6월 RG
# 비용을 +55% 과대계상한 사고의 원인이었다. 이 테스트가 규칙을 못박는다.
from app.services.coupang.rg_settlement_sync import _expected_period_start, _resolve_period_start

# prod 계정 row 18개 실주기(2026) — (period_end → period_from) 재현 ground truth (전수)
_VERIFIED_CYCLES = [
    (date(2026, 3, 8), date(2026, 3, 2)),
    (date(2026, 3, 15), date(2026, 3, 9)),
    (date(2026, 3, 22), date(2026, 3, 16)),
    (date(2026, 3, 29), date(2026, 3, 23)),
    (date(2026, 3, 31), date(2026, 3, 30)),   # 월말 부분주(월~화)
    (date(2026, 4, 5), date(2026, 4, 1)),     # 월초 부분주(수~일) — 월 경계 분할
    (date(2026, 4, 12), date(2026, 4, 6)),
    (date(2026, 4, 19), date(2026, 4, 13)),
    (date(2026, 4, 26), date(2026, 4, 20)),
    (date(2026, 4, 30), date(2026, 4, 27)),   # 월말 부분주(월~목)
    (date(2026, 5, 3), date(2026, 5, 1)),     # 월초 부분주(금~일)
    (date(2026, 5, 10), date(2026, 5, 4)),
    (date(2026, 5, 17), date(2026, 5, 11)),
    (date(2026, 5, 24), date(2026, 5, 18)),
    (date(2026, 5, 31), date(2026, 5, 25)),
    (date(2026, 6, 7), date(2026, 6, 1)),
    (date(2026, 6, 14), date(2026, 6, 8)),
    (date(2026, 6, 21), date(2026, 6, 15)),
]

# 추가 경계 케이스(월 경계·요일별)
_BOUNDARY_CYCLES = [
    (date(2026, 6, 30), date(2026, 6, 29)),    # 화요일 월말 → 06-29(월)
    (date(2026, 7, 5), date(2026, 7, 1)),      # 일요일이나 월초 분할 → 07-01
    (date(2026, 7, 12), date(2026, 7, 6)),     # 온전한 주(월~일)
    (date(2026, 8, 1), date(2026, 8, 1)),      # 토요일 1일 = 그 자체(월 경계)
    (date(2026, 9, 1), date(2026, 9, 1)),      # 화요일 1일 = 그 자체
    (date(2026, 12, 31), date(2026, 12, 28)),  # 목요일 월말 → 12-28(월)
]


@pytest.mark.parametrize("period_end,expected_from", _VERIFIED_CYCLES + _BOUNDARY_CYCLES)
def test_expected_period_start_calendar_rule(period_end, expected_from):
    """월~일 주별 + 달력 월 경계 분할. prod 18/18 + 경계 케이스."""
    assert _expected_period_start(period_end) == expected_from


def test_resolve_period_start_calendar_fallback_at_month_boundary():
    """계정 row 없을 때 달력 규칙 폴백. 06-30 → 06-29 (구 -6d였으면 06-24 오류)."""
    db = _db()
    got = _resolve_period_start(db, "COUPANG_WING1", date(2026, 6, 30))
    assert got == date(2026, 6, 29)
    assert got != date(2026, 6, 24)  # 구 -6d 회귀 가드


def test_resolve_period_start_borrows_account_row_unchanged():
    """계정 row 있으면 그 from을 차용(1순위 불변) — 달력 규칙보다 우선."""
    db = _db()
    db.add(CoupangRgSettlementFee(
        account_key="COUPANG_WING1", recognition_date_from=date(2026, 6, 25),
        recognition_date_to=date(2026, 6, 30), fee_type="warehousing", vendor_item_id="",
        raw_type="status/api", amount=Decimal("100")))
    db.commit()
    # 계정 row from=06-25 (달력 규칙 06-29와 일부러 다른 값) → 차용값이 우선함을 확인
    assert _resolve_period_start(db, "COUPANG_WING1", date(2026, 6, 30)) == date(2026, 6, 25)


# ─── 파싱 성능 가드 (2026-07-17 라이브 사고, 병행 세션 PR #39) ──────────────
def test_parse_xlsx_scales_linearly_not_quadratically():
    """대형 시트 파싱이 O(n²)로 퇴행하지 않는지 가드.

    ★사고(2026-07-17 21:07~21:12): read_only=True 워크북에 랜덤접근 ws.cell(row,col)을
    행 루프 안에서 쓰면, openpyxl은 cell() 호출마다 시트 XML을 1행부터 **다시 파싱**한다
    (ReadOnlyWorksheet._get_cell → _cells_by_row → 새 WorkSheetParser). 행당 cell() 3회 =
    O(n²) → 실측 100행 1.3초 / 200행 4.5초 / 400행 16.4초(2.5시간 아니고 배가 3.5배씩).
    실제 정산 엑셀(≈800행/시트)이 60초를 넘겨 Mac 페처 POST가 타임아웃(timeout=60)했고,
    단일 워커(uvicorn --workers 1)라 그동안 다른 API까지 멈췄다.

    → 파서는 시트당 **1회 스트리밍 순회**만 해야 한다(iter_rows). 수정 후 1000행 ≈ 0.06초.
    경계 5초 = 수정본 대비 80배 여유(느린 CI 대비) · 퇴행본(≈43초) 대비 8배 아래 → 둘을 확실히 가름.
    """
    import time

    n_rows = 1000
    rows = [(str(95521944483 + (i % 45)), "2026-06-04", "2026-06-07", 3300, 1050, 2250)
            for i in range(n_rows)]
    content = _build_xlsx([("입출고비", "입출고비", rows, 25, None, "2026-06-07")])

    t0 = time.perf_counter()
    parsed = parse_settlement_xlsx(content)
    elapsed = time.perf_counter() - t0

    # 정확성도 같이 확인 — 빨라졌다고 결과가 달라지면 안 된다.
    sheet = parsed["sheets"][0]
    assert len(sheet["options"]) == 45          # 고유 옵션(1000행이 45개로 합산)
    assert sheet["sum_detail"] == Decimal("2250") * n_rows
    assert elapsed < 5.0, (
        f"{n_rows}행 파싱에 {elapsed:.1f}초 — O(n²) 퇴행 의심"
        " (read_only 워크북에 랜덤 ws.cell() 쓰면 재발)"
    )


# ─── codex 리뷰 반영(원칙19) ──────────────────────────
def test_ingest_snapshot_replace_removes_stale_option():
    """재업로드 시 이번 엑셀에 없는 옵션 row 삭제(codex P2 snapshot replace)."""
    db = _db()
    ingest_settlement_xlsx(db, "COUPANG_WING1", _build_xlsx([("입출고비", "입출고비", _WH_ROWS)]))
    # 둘째 업로드: ...483만 남고 ...484 사라짐
    rows2 = [("95521944483", "2026-06-04", "2026-06-07", 1650, 525, 1125)]
    ingest_settlement_xlsx(db, "COUPANG_WING1", _build_xlsx([("입출고비", "입출고비", rows2)]))
    opts = db.query(CoupangRgSettlementFee).filter(
        CoupangRgSettlementFee.fee_type == "warehousing",
        CoupangRgSettlementFee.vendor_item_id != "").all()
    assert {o.vendor_item_id for o in opts} == {"95521944483"}  # 484 제거됨
    assert db.query(CoupangRgSettlementFee).filter_by(
        fee_type="warehousing", vendor_item_id="95521944483").one().amount == Decimal("1125")

def test_agg_rg_settlement_excludes_option_rows():
    """대조뷰 집계는 계정 row(vendor_item_id='')만 — 옵션 row 이중계상 차단(codex P1)."""
    from app.services.coupang.intelligence import _agg_rg_settlement_fees
    db = _db()
    db.add(CoupangRgSettlementFee(
        account_key="COUPANG_WING1", recognition_date_from=date(2026, 6, 1),
        recognition_date_to=date(2026, 6, 7), fee_type="warehousing", vendor_item_id="",
        raw_type="totalWarehousingFeeDeductionAmount", amount=Decimal("75489")))  # 계정 VAT후
    db.add(CoupangRgSettlementFee(
        account_key="COUPANG_WING1", recognition_date_from=date(2026, 6, 1),
        recognition_date_to=date(2026, 6, 7), fee_type="warehousing", vendor_item_id="95521944483",
        raw_type="xlsx:입출고비", amount=Decimal("21375")))  # 옵션 VAT前
    db.commit()
    agg = _agg_rg_settlement_fees(db, date(2026, 6, 1), date(2026, 6, 7))
    # 계정 row만(75489), 옵션 row(21375) 미합산
    assert agg["COUPANG_WING1"]["warehousing"] == Decimal("75489")
    assert agg["COUPANG_WING1"]["total"] == Decimal("75489")


# ─── S3 트랙 D3(grain 파라미터화) — IRON RULE 회귀 + 옵션 grain 신규 ────
def test_agg_rg_settlement_default_grain_is_account_unchanged():
    """IRON RULE: grain 미지정 시 기존 account 동작과 완전 동일(회귀 가드)."""
    from app.services.coupang.intelligence import _agg_rg_settlement_fees
    db = _db()
    db.add(CoupangRgSettlementFee(
        account_key="COUPANG_WING1", recognition_date_from=date(2026, 6, 1),
        recognition_date_to=date(2026, 6, 7), fee_type="warehousing", vendor_item_id="",
        raw_type="totalWarehousingFeeDeductionAmount", amount=Decimal("75489")))
    db.add(CoupangRgSettlementFee(
        account_key="COUPANG_WING1", recognition_date_from=date(2026, 6, 1),
        recognition_date_to=date(2026, 6, 7), fee_type="warehousing", vendor_item_id="95521944483",
        raw_type="xlsx:입출고비", amount=Decimal("21375")))
    db.commit()
    default_call = _agg_rg_settlement_fees(db, date(2026, 6, 1), date(2026, 6, 7))
    explicit_account = _agg_rg_settlement_fees(db, date(2026, 6, 1), date(2026, 6, 7), grain="account")
    assert default_call == explicit_account == {
        "COUPANG_WING1": {"warehousing": Decimal("75489"), "total": Decimal("75489")}
    }


def test_agg_rg_settlement_option_grain_excludes_account_rows():
    """grain='option'은 옵션 row(vendor_item_id≠'')만 — 계정 row(sentinel '') 미포함."""
    from app.services.coupang.intelligence import _agg_rg_settlement_fees
    db = _db()
    db.add(CoupangRgSettlementFee(
        account_key="COUPANG_WING1", recognition_date_from=date(2026, 6, 1),
        recognition_date_to=date(2026, 6, 7), fee_type="warehousing", vendor_item_id="",
        raw_type="totalWarehousingFeeDeductionAmount", amount=Decimal("75489")))
    db.add(CoupangRgSettlementFee(
        account_key="COUPANG_WING1", recognition_date_from=date(2026, 6, 1),
        recognition_date_to=date(2026, 6, 7), fee_type="warehousing", vendor_item_id="95521944483",
        raw_type="xlsx:입출고비", amount=Decimal("21375")))
    db.add(CoupangRgSettlementFee(
        account_key="COUPANG_WING1", recognition_date_from=date(2026, 6, 1),
        recognition_date_to=date(2026, 6, 7), fee_type="sale_fee", vendor_item_id="95521944483",
        raw_type="xlsx:판매수수료", amount=Decimal("5000")))
    db.commit()
    agg = _agg_rg_settlement_fees(db, date(2026, 6, 1), date(2026, 6, 7), grain="option")
    assert agg == {
        "COUPANG_WING1": {
            "95521944483": {
                "warehousing": Decimal("21375"),
                "sale_fee": Decimal("5000"),
                "total": Decimal("26375"),
            }
        }
    }


def test_agg_rg_settlement_option_grain_account_filter():
    """grain='option' + account_key 필터 조합(계정 분리 불변식 유지)."""
    from app.services.coupang.intelligence import _agg_rg_settlement_fees
    db = _db()
    db.add(CoupangRgSettlementFee(
        account_key="COUPANG_WING1", recognition_date_from=date(2026, 6, 1),
        recognition_date_to=date(2026, 6, 7), fee_type="warehousing", vendor_item_id="A1",
        raw_type="xlsx:입출고비", amount=Decimal("100")))
    db.add(CoupangRgSettlementFee(
        account_key="COUPANG_WING2", recognition_date_from=date(2026, 6, 1),
        recognition_date_to=date(2026, 6, 7), fee_type="warehousing", vendor_item_id="B1",
        raw_type="xlsx:입출고비", amount=Decimal("200")))
    db.commit()
    agg = _agg_rg_settlement_fees(db, date(2026, 6, 1), date(2026, 6, 7),
                                  account_key="COUPANG_WING1", grain="option")
    assert agg == {"COUPANG_WING1": {"A1": {"warehousing": Decimal("100"), "total": Decimal("100")}}}


def test_agg_rg_settlement_invalid_grain_raises():
    from app.services.coupang.intelligence import _agg_rg_settlement_fees
    db = _db()
    with pytest.raises(ValueError):
        _agg_rg_settlement_fees(db, date(2026, 6, 1), date(2026, 6, 7), grain="bogus")


# ─── codex 2R 반영(원칙19) ────────────────────────────
def test_ingest_same_fee_type_multiple_sheets_merged():
    """같은 fee_type 시트 2개(반품 회수비+재입고비→return_shipping) → 병합 합산, 삭제 충돌 없음(codex 2R-P1)."""
    db = _db()
    rows_a = [("111", "2026-06-04", "2026-06-07", 1000, 0, 1000)]
    rows_b = [("111", "2026-06-04", "2026-06-07", 500, 0, 500)]  # 같은 옵션·기간
    content = _build_xlsx([("반품 회수비", "반품 회수비", rows_a),
                           ("반품 재입고비", "반품 재입고비", rows_b)])
    ingest_settlement_xlsx(db, "COUPANG_WING1", content)
    opt = db.query(CoupangRgSettlementFee).filter_by(
        fee_type="return_shipping", vendor_item_id="111").one()
    assert opt.amount == Decimal("1500")  # 1000+500 병합(둘째 시트가 첫째 삭제 안 함)

def test_ingest_empty_detail_sheet_still_snapshots():
    """상세 0건이지만 요약 period_end 있는 시트 → 기존 옵션 snapshot replace(codex 2R-P2)."""
    db = _db()
    ingest_settlement_xlsx(db, "COUPANG_WING1", _build_xlsx([("입출고비", "입출고비", _WH_ROWS)]))
    assert db.query(CoupangRgSettlementFee).filter(
        CoupangRgSettlementFee.fee_type == "warehousing",
        CoupangRgSettlementFee.vendor_item_id != "").count() > 0
    # 상세 0건 + 요약(period_end 2026-06-07) 시트 재업로드 → 옵션 전삭제
    ingest_settlement_xlsx(db, "COUPANG_WING1", _build_xlsx([("입출고비", "입출고비", [])]))
    assert db.query(CoupangRgSettlementFee).filter(
        CoupangRgSettlementFee.fee_type == "warehousing",
        CoupangRgSettlementFee.vendor_item_id != "").count() == 0


# ─── codex 3R 반영(원칙19) ────────────────────────────
def test_ingest_detail_no_period_uses_summary_fallback():
    """상세 종료일 없으면 요약 period로 fallback insert(데이터 손실 방지, codex 3R-P1)."""
    db = _db()
    rows = [("222", "2026-06-04", "", 1000, 0, 1000)]  # 상세 종료일 빈칸
    content = _build_xlsx([("입출고비", "입출고비", rows, 25, None, "2026-06-07")])  # 요약 종료일만
    ingest_settlement_xlsx(db, "COUPANG_WING1", content)
    opt = db.query(CoupangRgSettlementFee).filter_by(
        fee_type="warehousing", vendor_item_id="222").one()  # insert됨(손실 없음)
    assert opt.recognition_date_to == date(2026, 6, 7)  # 요약으로 fallback
    assert opt.amount == Decimal("1000")

def test_ingest_vs_status_api_merged_by_fee_type():
    """같은 fee_type 여러 시트 → vs_status_api는 fee_type+period 합계 기준(codex 3R-P2, false mismatch 방지)."""
    db = _db()
    # status/api 계정 row: return_shipping 최종 1650 = (1000+500)*1.1
    db.add(CoupangRgSettlementFee(
        account_key="COUPANG_WING1", recognition_date_from=date(2026, 6, 1),
        recognition_date_to=date(2026, 6, 7), fee_type="return_shipping", vendor_item_id="",
        raw_type="status/api", amount=Decimal("1650")))
    db.commit()
    rows_a = [("111", "2026-06-04", "2026-06-07", 1000, 0, 1000)]  # 시트A final 1100
    rows_b = [("111", "2026-06-04", "2026-06-07", 500, 0, 500)]    # 시트B final 550
    content = _build_xlsx([("반품 회수비", "반품 회수비", rows_a),
                           ("반품 재입고비", "반품 재입고비", rows_b)])
    res = ingest_settlement_xlsx(db, "COUPANG_WING1", content)
    vs = [s["vs_status_api"] for s in res["sheets"]]
    assert all(v is not None and v["match"] for v in vs)  # 합계 1650 == 계정 1650
    assert vs[0]["summary_final"] == Decimal("1650")  # 1100+550 합계(시트별 아님)


# ─── 쿠키 저장↔로드 정합 (2026-06-10 xsrf raw저장 버그 회귀) ──────────────────
# save_settlement_cookie(dead code)가 xsrf를 raw로 저장 → _load_client의 decrypt_secret(xsrf)와
# 어긋나 로드 시 CookieCryptoError 크래시. 함수 제거 후 저장은 rg_inbound_sync.save_cookie 단일
# 경로. 이 테스트는 정상 경로가 xsrf를 암호화 라운드트립하는지 + raw 저장이 _load_client에서
# 실패하는지(=암호화 계약)를 못박는다.
from cryptography.fernet import Fernet
from app.services.coupang.rg_inbound_sync import save_cookie
from app.services.coupang.rg_settlement_sync import _load_client
from app.models import CoupangWingCookie
from app.utils.crypto import decrypt_secret, CookieCryptoError

_CURL = (
    "curl 'https://wing.coupang.com/tenants/rfm/v2/settlements/status/api' "
    "-H 'x-xsrf-token: tok-XSRF-9z8y7x' "
    "-H 'cookie: XSRF-TOKEN=tok-XSRF-9z8y7x; SESSION=sess-abc; bm_sv=zzz'"
)


def test_save_cookie_to_load_client_roundtrips_xsrf(monkeypatch):
    """정상 경로: save_cookie 저장 → _load_client 로드가 xsrf/cookie를 원본대로 복원(크래시 없음)."""
    monkeypatch.setenv("COOKIE_ENC_KEY", Fernet.generate_key().decode())
    db = _db()

    save_cookie(db, "COUPANG_WING1", _CURL)

    row = db.query(CoupangWingCookie).filter_by(account_key="COUPANG_WING1").first()
    assert row is not None
    # 저장은 평문 아님(암호문) — raw 저장 회귀 차단
    assert row.xsrf_token != "tok-XSRF-9z8y7x"
    assert row.cookie_blob != "XSRF-TOKEN=tok-XSRF-9z8y7x; SESSION=sess-abc; bm_sv=zzz"
    # 복호화하면 원본
    assert decrypt_secret(row.xsrf_token) == "tok-XSRF-9z8y7x"

    # _load_client가 예외 없이 원본을 복원해 클라이언트에 주입
    client = _load_client(db, "COUPANG_WING1")
    assert client.xsrf == "tok-XSRF-9z8y7x"
    assert "SESSION=sess-abc" in client.cookie


def test_load_client_rejects_raw_unencrypted_xsrf(monkeypatch):
    """회귀: xsrf를 raw(평문)로 저장하면 _load_client가 CookieCryptoError로 표면화(암호화 계약).

    과거 save_settlement_cookie가 정확히 이렇게 raw 저장 → 로드 크래시였다. 평문 저장 경로가
    다시 생기면 이 테스트가 깨져 알린다.
    """
    monkeypatch.setenv("COOKIE_ENC_KEY", Fernet.generate_key().decode())
    db = _db()

    save_cookie(db, "COUPANG_WING1", _CURL)  # cookie_blob은 정상 암호화
    row = db.query(CoupangWingCookie).filter_by(account_key="COUPANG_WING1").first()
    row.xsrf_token = "tok-XSRF-9z8y7x"  # ← raw로 덮어쓰기(과거 버그 재현)
    db.commit()

    with pytest.raises(CookieCryptoError):
        _load_client(db, "COUPANG_WING1")


# ═══════════════════════════════════════════════════════════════════════
# S4-P2 — RG 정산 자동 다운로드 갱신 트리거·heartbeat (D-8)
# ═══════════════════════════════════════════════════════════════════════
from app.services.coupang.rg_settlement_sync import (
    rg_request_refresh,
    rg_refresh_status,
    rg_claim_refresh,
    rg_mark_heartbeat,
    _RG_STATE_ACCOUNT,
)


def test_rg_refresh_status_initial_none():
    """상태행 없으면 requested=False·status='none'(빈 DB 안전)."""
    db = _db()
    st = rg_refresh_status(db)
    assert st["requested"] is False
    assert st["status"] == "none"
    assert st["last_success_at"] is None
    assert st["stale"] is False


def test_rg_request_refresh_sets_flag():
    """request_refresh → 플래그 set → status가 requested=True 반영."""
    db = _db()
    r = rg_request_refresh(db)
    assert r["requested"] is True and r["requested_at"]
    st = rg_refresh_status(db)
    assert st["requested"] is True
    assert st["requested_at"] is not None


def test_rg_claim_refresh_atomic_once():
    """claim은 요청 있을 때 1회만 claimed=True, 이후 False(중복 fetch 방지).

    ★2026-07-27: claim은 요청을 소비하지 않고 **임대**한다(재시도 보장) — 두 번째 claim이
    False인 이유가 "이미 clear돼서"에서 "lease가 살아있어서"로 바뀌었다. 요청은 성공하거나
    3회 실패/로그인필요일 때 소멸한다.
    """
    db = _db()
    rg_request_refresh(db)
    assert rg_claim_refresh(db)["claimed"] is True
    assert rg_claim_refresh(db)["claimed"] is False   # lease 보유 중 → 중복 claim 차단
    assert rg_refresh_status(db)["requested"] is True  # ★요청은 살아있다
    assert rg_refresh_status(db)["in_flight"] is True


def test_rg_claim_without_request_is_false():
    """요청 없이 claim하면 claimed=False(유실 없음)."""
    db = _db()
    assert rg_claim_refresh(db)["claimed"] is False


def test_rg_mark_heartbeat_green_and_success_at():
    """heartbeat → status green + last_success_at 기록(stale=False)."""
    db = _db()
    rg_mark_heartbeat(db)
    st = rg_refresh_status(db)
    assert st["status"] == "green"
    assert st["last_success_at"] is not None
    assert st["stale"] is False
    assert st["age_hours"] is not None and st["age_hours"] >= 0


def test_rg_state_isolated_from_vendor_summary():
    """RG 상태행은 vendor-summary(COUPANG_WING_VS)와 다른 account_key로 분리(D-5)."""
    from app.services.coupang import vendor_summary_sync as vss
    db = _db()
    rg_request_refresh(db)
    # vendor-summary refresh_status는 RG 요청에 오염되지 않아야 함
    assert vss.refresh_status(db)["requested"] is False
    assert _RG_STATE_ACCOUNT != vss._VS_ACCOUNT


# ═══════════════════════════════════════════════════════════════════════
# 버튼 큐 계정 분리 (2026-07-27 — WING2 데몬 편입)
#   상태행 1개를 두 계정이 공유하면 WING2 데몬이 WING1 버튼 요청을 claim해 가져간다.
#   아래는 그 도난이 구조적으로 불가능함을 고정하는 회귀 테스트.
# ═══════════════════════════════════════════════════════════════════════
from app.services.coupang.rg_settlement_sync import (  # noqa: E402
    _RG_STATE_ACCOUNT_BY_ACCOUNT,
    _rg_state_key,
    rg_mark_fetch_error,
)

_W1 = "COUPANG_WING1"
_W2 = "COUPANG_WING2"


def test_rg_state_key_maps_and_rejects_unknown():
    """WING1은 기존 행 재사용(하위호환), WING2는 별도 행. 그 밖의 값은 ValueError(라우터가 400)."""
    assert _rg_state_key(_W1) == _RG_STATE_ACCOUNT      # 기존 데이터 그대로 이어씀
    assert _rg_state_key(_W2) == "COUPANG_WING_RG2"
    assert _RG_STATE_ACCOUNT_BY_ACCOUNT[_W1] != _RG_STATE_ACCOUNT_BY_ACCOUNT[_W2]
    with pytest.raises(ValueError):
        _rg_state_key("COUPANG_WING9")


def test_rg_default_account_is_wing1_backcompat():
    """account_key 생략 = WING1 — 파라미터 없는 구버전 페처가 기존과 똑같이 동작(배포 순서 무관)."""
    db = _db()
    rg_request_refresh(db)                       # 생략
    assert rg_refresh_status(db, _W1)["requested"] is True
    assert rg_claim_refresh(db)["claimed"] is True  # 생략 claim이 WING1 요청을 소비


def test_rg_wing2_cannot_claim_wing1_request():
    """WING1 요청을 WING2 claim이 가져가지 못한다(큐 도난 차단 — 이 스프린트의 존재 이유)."""
    db = _db()
    rg_request_refresh(db, _W1)
    assert rg_claim_refresh(db, _W2)["claimed"] is False
    assert rg_refresh_status(db, _W1)["requested"] is True   # 요청은 그대로 살아있다
    assert rg_claim_refresh(db, _W1)["claimed"] is True      # 주인만 소비 가능


def test_rg_wing1_cannot_claim_wing2_request():
    """역방향도 동일 — WING2 요청을 WING1(=기본값) 데몬이 훔쳐가지 못한다."""
    db = _db()
    rg_request_refresh(db, _W2)
    assert rg_claim_refresh(db)["claimed"] is False          # 생략=WING1
    assert rg_refresh_status(db, _W2)["requested"] is True
    assert rg_claim_refresh(db, _W2)["claimed"] is True


def test_rg_heartbeat_is_per_account():
    """heartbeat는 계정별 상태행에 기록 — WING2 push가 WING1을 '갱신 완료'로 위장하면 안 된다."""
    db = _db()
    rg_mark_heartbeat(db, _W2)
    assert rg_refresh_status(db, _W2)["last_success_at"] is not None
    st1 = rg_refresh_status(db, _W1)
    assert st1["last_success_at"] is None and st1["status"] == "none"


def test_rg_fetch_error_is_per_account():
    """실패 보고도 계정별 — 남의 실패가 내 화면에 뜨면 원인 오진단."""
    db = _db()
    rg_mark_fetch_error(db, "chrome crash", _W2)
    assert rg_refresh_status(db, _W2)["last_error"] == "chrome crash"
    assert rg_refresh_status(db, _W1)["last_error"] is None
