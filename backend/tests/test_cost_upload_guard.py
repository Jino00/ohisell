# test_cost_upload_guard.py — 엑셀 업로드가 **버퍼 값을 DB에 쓰지 못하게** 지킨다.
#
# ## 왜 이 파일이 따로 있나 (2026-08-10, D-CPP-35)
#
# `test_cost_drift_wiring.py`는 **탐지**를 지킨다 — 이미 들어간 값이 헬스 배너까지 흐르는가.
# 그런데 탐지는 «되돌아간 뒤에» 켜진다. 옛 매핑 엑셀(v3)을 올리면 버퍼 177건이 통째로
# 복귀하고, 그 복귀는 **에러가 안 난다** — 이익만 조용히 줄어든 채 조회된다.
#
# 그래서 이 파일이 지키는 것은 **예방**이다: 업로드가 쓰기 **전에** 거부하는가.
# D-CPP-34가 버퍼를 «단순 오류·재입력 금지»로 확정했으므로 경고가 아니라 거부다.
#
# ★두 경로의 처분이 **다르다**. 그 차이 자체가 회귀 대상이라 여기서 못 박는다:
#   · `/upload`(상품 원가표 시트)   → 행 전체 건너뜀. 행의 존재 이유가 원가다.
#   · `/upload-by-name`(원가 매핑)  → 원가만 안 씀, **매핑은 정상 처리**. 행의 존재 이유가 매핑이다.
#
# ★그리고 HTTP body를 단언한다. 서비스층 dict에만 있고 응답 스키마에 필드가 없으면
#   FastAPI가 **조용히 지운다** — 2026-08-10에 `cost_drift`가 정확히 그렇게 사라졌고
#   배선 테스트 6건이 하나도 안 죽었다(ref 54 §9 P1-1).
from __future__ import annotations

import io
from decimal import Decimal

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base, get_db
from app.models import Channel, ProductMaster
from app.services.cost_truth_audit import (
    screen_cost,
    screen_reason,
    try_load_truth,
)
from app.services.product_mapping_ingest import ingest_master_sheet, upsert_product_by_name

# 라이브 실측(2026-08-10). 정본 2350.7 + 폰 버퍼 265.3 = 2616 — 되돌아오던 바로 그 값이다.
_TRUTH = 2350.7
_BUFFERED = 2616.0
_UNDETERMINED = 3500.0  # OHI-TGLASS-IP17PRO 실제 값 — 원가표에 없다


@pytest.fixture
def truth():
    t = try_load_truth()
    assert t is not None, "정본 스냅샷이 배포 트리에 없다 — 가드가 통째로 무력화된다"
    return t


@pytest.fixture
def db():
    # StaticPool + check_same_thread=False 필수 — TestClient가 라우트를 다른 스레드에서 돈다.
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    session = sessionmaker(bind=engine)()
    yield session
    session.close()


@pytest.fixture
def client(db):
    from fastapi.testclient import TestClient

    from app.main import app

    app.dependency_overrides[get_db] = lambda: db
    yield TestClient(app)
    app.dependency_overrides.clear()


# ═══ 순수층 — screen_cost가 셋 중 «버퍼»만 잡는가 ═══


def test_screen_blocks_buffered_value(truth):
    hit = screen_cost(_BUFFERED, truth)
    assert hit is not None
    assert hit["truth"] == pytest.approx(_TRUTH)
    assert hit["buffer"] == pytest.approx(265.3)


def test_screen_passes_truth_value(truth):
    """★정본 값은 통과해야 한다 — 안 그러면 올바른 엑셀이 업로드 불가가 된다."""
    assert screen_cost(_TRUTH, truth) is None


def test_screen_passes_undetermined_value(truth):
    """★«판정 불가»(459건 계열)를 막지 않는다.

    이 가드가 답할 수 있는 건 «정본 + 알려진 버퍼인가»뿐이다. 모르는 값을 막으면
    근거 없이 업로드를 깨뜨린다 — 케이스·거치대 계열이 통째로 못 올라간다.
    """
    assert screen_cost(_UNDETERMINED, truth) is None


def test_screen_is_noop_without_snapshot():
    """스냅샷이 없으면 판정하지 않는다(fail-soft) — 가드 부재가 업로드를 막지 않는다."""
    assert screen_cost(_BUFFERED, None) is None


def test_screen_survives_unparseable_cost(truth):
    """값 파싱 실패를 «버퍼»로 만들지 않는다 — 파싱은 부르는 쪽 책임이다."""
    assert screen_cost("이건원가가아님", truth) is None
    assert screen_cost(None, truth) is None


def test_reason_carries_truth_value_so_a_human_can_fix_the_excel(truth):
    """★사유에 정본 값이 없으면 사람은 «뭘로 고쳐야 하는지»를 모른다."""
    msg = screen_reason("행5", screen_cost(_BUFFERED, truth))
    assert "2350.7" in msg
    assert "265.3" in msg


# ═══ upsert 층 — cost_blocked가 실제로 원가를 안 쓰는가 ═══


def test_blocked_update_keeps_existing_cost(db):
    """★기존 상품: 버퍼가 와도 **DB의 옛 값이 그대로 남는다**(이게 복귀 차단의 핵심)."""
    db.add(ProductMaster(internal_sku="OHI-1", product_name="필름", cost_price=Decimal("2350.7")))
    db.commit()
    upsert_product_by_name(db, "필름", Decimal("2616"), [2], cost_blocked=True)
    db.commit()
    assert db.query(ProductMaster).one().cost_price == Decimal("2350.7")


def test_unblocked_update_still_writes(db):
    """가드가 없을 땐 종전 동작 그대로 — 하위호환."""
    db.add(ProductMaster(internal_sku="OHI-1", product_name="필름", cost_price=Decimal("100")))
    db.commit()
    upsert_product_by_name(db, "필름", Decimal("2350.7"), [2], cost_blocked=False)
    db.commit()
    assert db.query(ProductMaster).one().cost_price == Decimal("2350.7")


def test_blocked_create_uses_zero_not_the_buffered_value(db):
    """★신규 상품: 0(모르는 값)으로 만든다. 버퍼(틀린 값)를 넣지 않는다."""
    product, action = upsert_product_by_name(db, "새필름", Decimal("2616"), [1], cost_blocked=True)
    db.commit()
    assert action == "created"
    assert product.cost_price == Decimal("0")


# ═══ 배선층 — 실제 엑셀 시트가 ingest를 통과할 때 ═══


def _master_sheet(rows):
    """"원가 매핑" 시트 형태(채널명 마커 + 옵션ID 블록)의 워크시트를 만든다."""
    wb = Workbook()
    ws = wb.active
    ws.append(["상품명", "원가", "채널명", "옵션ID"])
    for name, cost, label, oid in rows:
        ws.append([name, cost, label, oid])
    return wb, ws


def test_ingest_blocks_buffer_but_keeps_the_mapping(db):
    """★이 경로의 처분: 원가는 안 쓰고 **매핑은 산다.**

    행 전체를 버리면 매핑이 사라져 주문이 미연결로 남는다 — 원가 하나 때문에 잃기엔 크다.
    """
    db.add(Channel(code="NAVER", name="네이버", platform="naver"))
    db.add(ProductMaster(internal_sku="OHI-1", product_name="필름", cost_price=Decimal("2350.7")))
    db.commit()

    _, ws = _master_sheet([("필름", _BUFFERED, "네이버 스마트스토어", "OPT-1")])
    result = ingest_master_sheet(db, ws)

    assert db.query(ProductMaster).one().cost_price == Decimal("2350.7")  # 원가 불변
    assert result.mappings_created == 1  # 매핑은 살았다
    assert len(result.integrity.cost_buffers) == 1
    assert "2350.7" in result.integrity.cost_buffers[0]


def test_ingest_lets_a_correct_cost_through(db):
    """정상 엑셀은 종전대로 통과한다 — 가드가 정상 경로를 막지 않는다."""
    db.add(Channel(code="NAVER", name="네이버", platform="naver"))
    db.add(ProductMaster(internal_sku="OHI-1", product_name="필름", cost_price=Decimal("100")))
    db.commit()

    _, ws = _master_sheet([("필름", _TRUTH, "네이버 스마트스토어", "OPT-1")])
    result = ingest_master_sheet(db, ws)

    assert db.query(ProductMaster).one().cost_price == Decimal("2350.7")
    assert result.integrity.cost_buffers == []
    assert result.integrity.cost_guard_unavailable is None


# ═══ HTTP 경계 — 응답 스키마가 사유를 지우지 않는가 ═══


def test_upload_by_name_reports_blocked_rows_in_http_body(client, db):
    """★서비스층에만 있고 응답 스키마에 필드가 없으면 FastAPI가 조용히 지운다.

    2026-08-10에 `cost_drift`가 정확히 그렇게 사라졌고, dict까지만 본 테스트 6건이
    하나도 안 죽었다. 그래서 여기서는 **HTTP body를 단언한다.**
    """
    db.add(Channel(code="NAVER", name="네이버", platform="naver"))
    db.add(ProductMaster(internal_sku="OHI-1", product_name="필름", cost_price=Decimal("2350.7")))
    db.commit()

    wb, _ = _master_sheet([("필름", _BUFFERED, "네이버 스마트스토어", "OPT-1")])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        "/api/products/upload-by-name",
        files={"file": ("m.xlsx", buf, "application/vnd.ms-excel")},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert "cost_buffers" in body, "응답에서 사유가 지워졌다 — 화면은 «이상 없음»으로 보인다"
    assert len(body["cost_buffers"]) == 1
    assert "2350.7" in body["cost_buffers"][0]
    assert db.query(ProductMaster).one().cost_price == Decimal("2350.7")


def test_upload_cost_sheet_skips_buffered_row_and_says_so(client, db):
    """`/upload`(상품 원가표 시트)는 **행 전체를 건너뛴다** — 음수 가드와 같은 처분."""
    db.add(ProductMaster(internal_sku="OHI-1", product_name="옛이름", cost_price=Decimal("2350.7")))
    db.commit()

    wb = Workbook()
    ws = wb.active
    ws.append(["SKU", "상품명", "원가", "카테고리", "메모"])
    ws.append(["OHI-1", "새이름", _BUFFERED, None, None])
    buf = io.BytesIO()
    ws.title = "상품 원가표"
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        "/api/products/upload",
        files={"file": ("c.xlsx", buf, "application/vnd.ms-excel")},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["cost_buffer_blocked"] == 1
    assert body["cost_guard"] == "active"
    assert body["updated"] == 0
    row = db.query(ProductMaster).one()
    assert row.cost_price == Decimal("2350.7")  # 원가 불변
    assert row.product_name == "옛이름"  # 행 전체를 건너뛰었으므로 이름도 안 바뀐다


def test_upload_cost_sheet_still_accepts_a_correct_row(client, db):
    """정상 행은 통과 — 가드가 붙어도 종전 업로드가 깨지지 않는다."""
    wb = Workbook()
    ws = wb.active
    ws.title = "상품 원가표"
    ws.append(["SKU", "상품명", "원가", "카테고리", "메모"])
    ws.append(["OHI-9", "새상품", _TRUTH, None, None])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(
        "/api/products/upload",
        files={"file": ("c.xlsx", buf, "application/vnd.ms-excel")},
    )
    assert resp.status_code == 200
    assert resp.json()["created"] == 1
    assert resp.json()["cost_buffer_blocked"] == 0
    assert db.query(ProductMaster).filter_by(internal_sku="OHI-9").one().cost_price == Decimal("2350.7")


# ═══ 휴면 가드 — 스냅샷이 사라지면 «검사 못 함»이 보이는가 ═══


def test_missing_snapshot_is_reported_not_silent(db, monkeypatch):
    """★검사 못 함과 이상 없음이 같은 모양이면 안 된다(교훈 #123).

    스냅샷이 없어도 업로드는 계속되되(fail-soft), 그 사실이 응답에 남아야 한다.
    """
    import app.services.product_mapping_ingest as ing

    monkeypatch.setattr(ing, "try_load_truth", lambda: None)
    db.add(Channel(code="NAVER", name="네이버", platform="naver"))
    db.commit()

    _, ws = _master_sheet([("필름", _BUFFERED, "네이버 스마트스토어", "OPT-1")])
    result = ingest_master_sheet(db, ws)

    assert result.integrity.cost_guard_unavailable is not None
    assert result.integrity.cost_buffers == []  # 판정을 안 했으니 «0건»이 아니라 «모름»이다
    assert result.products_created == 1  # 업로드 자체는 살아 있다
