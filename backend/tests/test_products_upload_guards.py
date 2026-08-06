# test_products_upload_guards.py — 구형 엑셀 업로드가 «원가를 모르게 만드는» 데이터를 못 넣는다
#
# ★왜(2026-08-06 적대 리뷰 P1): 이중 활성 매핑(1채널옵션ID→2마스터)이 생기면 손익에서
#   원가 후보가 둘이 되고 어느 쪽이 맞는지 알 수 없다. `_active_option_clash` 가드가
#   add/update 경로엔 있었는데(D-12) **구형 엑셀 업로드만 뚫려 있었다** —
#   `existing_mapping` 조회가 내 product_id로만 걸려 있어, 다른 상품이 이미 그 옵션ID를
#   active로 갖고 있어도 새 활성 매핑을 그냥 추가했다.
#   라이브 실측(2026-08-06): 네이버 이중 활성 매핑 **22건**이 이렇게 쌓여 있었다
#   (그중 원가가 갈리는 건 0건이라 아직 손익이 틀리진 않았다 — 구조만 깔려 있던 상태).
#
# ★음수 원가도 여기서 막는다 — 이 경로는 pydantic 스키마(ge=0)를 거치지 않고 DB에 직접 쓴다.
from __future__ import annotations

import io
from decimal import Decimal

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base, get_db
from app.models import Channel, ProductChannelMapping, ProductMaster
from app.routers import products as products_router

D = Decimal
NAVER_ID = 6


@pytest.fixture
def client_db():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False},
                           poolclass=StaticPool)
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine, autoflush=False, autocommit=False)
    s = Session()
    s.add(Channel(id=NAVER_ID, code="NAVER", name="네이버 스마트스토어",
                  platform="naver", channel_type="own"))
    s.commit()

    app = FastAPI()
    app.include_router(products_router.router)
    app.dependency_overrides[get_db] = lambda: s
    yield TestClient(app), s
    s.close()


def _xlsx(*, costs: list[tuple] | None = None, mappings: list[tuple] | None = None) -> bytes:
    """업로드 워크북 생성. 시트/열 순서는 upload_excel 구현과 같다."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("상품 원가표")
    ws.append(["내부SKU", "상품명", "원가", "카테고리", "메모"])
    for r in costs or []:
        ws.append(list(r))
    ws2 = wb.create_sheet("채널 매핑")
    ws2.append(["내부SKU", "채널코드", "채널상품ID", "채널상품명", "채널SKU", "판매가", "활성"])
    for r in mappings or []:
        ws2.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _post(client, content: bytes):
    return client.post(
        "/api/products/upload",
        files={"file": ("m.xlsx", content,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


# ── 이중 활성 매핑 차단 ──────────────────────────────────────────────

def test_upload_rejects_second_active_mapping_for_same_option_id(client_db):
    """★핵심 — 다른 상품이 이미 active로 쥔 옵션ID에 두 번째 활성 매핑을 만들지 않는다.

    종전엔 그냥 추가돼 손익의 원가 후보가 둘이 됐다(그리고 코드는 product_master.id가
    작은 쪽을 조용히 골랐다 — 원가 기준이 아니라).
    """
    client, db = client_db
    db.add(ProductMaster(id=10, internal_sku="SKU-A", product_name="A", cost_price=D("9000")))
    db.add(ProductChannelMapping(product_id=10, channel_id=NAVER_ID,
                                 channel_product_id="OPT-1", is_active=True))
    db.add(ProductMaster(id=55, internal_sku="SKU-B", product_name="B", cost_price=D("15000")))
    db.commit()

    res = _post(client, _xlsx(mappings=[("SKU-B", "NAVER", "OPT-1", "B", "", 30000, "Y")]))

    assert res.status_code == 200
    body = res.json()
    assert body["mappings_created"] == 0                    # 만들어지지 않았다
    assert any("이미 다른 상품" in e for e in body["errors"])
    # 라이브 DB에도 두 번째 활성 매핑이 없어야 한다.
    n_active = (
        db.query(ProductChannelMapping)
        .filter_by(channel_id=NAVER_ID, channel_product_id="OPT-1", is_active=True)
        .count()
    )
    assert n_active == 1


def test_upload_allows_inactive_mapping_for_same_option_id(client_db):
    """비활성(N)으로 올리는 것은 막지 않는다 — 이중 «활성»만 문제다.

    이 가드가 없으면 이력 보존용 비활성 매핑까지 못 올려 업로드가 과잉 차단된다.
    """
    client, db = client_db
    db.add(ProductMaster(id=10, internal_sku="SKU-A", product_name="A", cost_price=D("9000")))
    db.add(ProductChannelMapping(product_id=10, channel_id=NAVER_ID,
                                 channel_product_id="OPT-1", is_active=True))
    db.add(ProductMaster(id=55, internal_sku="SKU-B", product_name="B", cost_price=D("15000")))
    db.commit()

    res = _post(client, _xlsx(mappings=[("SKU-B", "NAVER", "OPT-1", "B", "", 30000, "N")]))

    assert res.status_code == 200
    assert res.json()["mappings_created"] == 1
    assert res.json()["errors"] == []


def test_upload_updates_own_existing_mapping_without_clash(client_db):
    """자기 매핑을 다시 올리는 것은 충돌이 아니다(가장 흔한 정상 경로 — 회귀 가드)."""
    client, db = client_db
    db.add(ProductMaster(id=10, internal_sku="SKU-A", product_name="A", cost_price=D("9000")))
    db.add(ProductChannelMapping(product_id=10, channel_id=NAVER_ID,
                                 channel_product_id="OPT-1", is_active=True,
                                 selling_price=D("20000")))
    db.commit()

    res = _post(client, _xlsx(mappings=[("SKU-A", "NAVER", "OPT-1", "A", "", 30000, "Y")]))

    assert res.status_code == 200
    assert res.json()["errors"] == []
    m = (db.query(ProductChannelMapping)
           .filter_by(channel_id=NAVER_ID, channel_product_id="OPT-1").first())
    assert m.selling_price == D("30000")   # 갱신은 정상 동작


# ── 음수 원가 차단 ───────────────────────────────────────────────────

def test_upload_rejects_negative_cost(client_db):
    """★음수 원가는 안 들어간다 — 이 경로는 스키마(ge=0)를 안 거친다.

    원가가 음수면 손익에서 빼는 게 아니라 **더한다**(과대의 방향이 뒤집힌 채 커진다).
    """
    client, db = client_db

    res = _post(client, _xlsx(costs=[("SKU-NEG", "음수원가", -500, "", "")]))

    assert res.status_code == 200
    body = res.json()
    assert body["created"] == 0
    assert any("음수" in e for e in body["errors"])
    assert db.query(ProductMaster).filter_by(internal_sku="SKU-NEG").first() is None


def test_upload_accepts_zero_and_positive_cost(client_db):
    """0원·양수는 그대로 통과한다(0은 «모름»으로 손익에서 처리되지 손익 밖에서 막지 않는다)."""
    client, db = client_db

    res = _post(client, _xlsx(costs=[("SKU-Z", "원가0", 0, "", ""),
                                     ("SKU-P", "원가양수", 12000, "", "")]))

    assert res.status_code == 200
    assert res.json()["created"] == 2
    assert res.json()["errors"] == []
