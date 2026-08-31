"""원가 메뉴 API — D-CPP-53 / 계약 `docs/PLAN_cost-menu-standard-cost.md` (S1: 부자재 층).

★**응답에 `response_model`을 쓰지 않는다.** 교훈 #321(2026-08-19): `response_model`이 선언 안 된
키를 HTTP body에서 조용히 지워, 서비스층 테스트 9건이 전부 초록인데 화면엔 배너가 통째로 안 뜨는
사고가 났다. 이 층은 「미매칭」·「미확인」·「제안 이유」처럼 **있으면 반드시 보여야 하는** 자백
필드가 많아 같은 함정의 표면이 넓다. 요청(입력)만 Pydantic으로 검증하고, 응답은
`services/cost_menu/materials.py`의 payload 함수가 만든 dict를 그대로 낸다.
대신 테스트가 **HTTP body를 단언**한다(서비스층 dict만 보면 못 잡는다).

★요청 스키마를 `app/schemas.py`가 아니라 이 파일에 둔다 — 병행 세션이 같은 파일을 건드릴 때
충돌하는 자리라서다(계약 B 라우터와 같은 이유). 이 라우터 밖에서 쓰이지 않는다.

★**`product_master.cost_price`를 쓰지 않는다** — 계약 C(D-CPP-64, 2026-08-31 승인)가 A′ §3의
이 금지선을 **조건부로** 풀었지만, 열린 것은 「컷오버 경로 한 벌 + Jino 클릭 + 이력 + 근거
좌표」일 때뿐이고 그 경로는 **S3에서 생긴다.** S1인 지금 이 라우터의 쓰기는 여전히 0이다.

★단 **이력은 읽는다**(2026-08-31, 계약 §4 S1-①): `GET /api/cost/price-history`는
`cost_price_history` 표를 읽어 「누가·언제·어느 문으로·무엇에서 무엇으로」를 화면에 낸다.
값 자체(`product_master.cost_price`)를 읽는 것이 아니라 **그 값이 움직인 사건**을 읽는 것이고,
쓰기는 아니다. 그 구별이 흐려지면 이 주석이 지키던 경계가 말로만 남는다.
"""

from __future__ import annotations

from datetime import date
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from typing import Literal, Optional

from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.database import get_db
from app.services import cost_price_history as CPH
from app.services.cost_menu import auto_refresh as AR
from app.services.cost_menu import materials as M
from app.services.cost_menu import recipes as R
from app.services.cost_menu import purchased_price as PP
from app.services.cost_menu import round_trip as RT
from app.services.cost_menu.purchased_price_parser import (
    PriceSheetError,
    parse_price_sheet,
)

router = APIRouter(prefix="/api/cost", tags=["cost-menu"])


# ──────────────────────────────────────────────
# 요청 스키마 (입력 검증 전용)
# ──────────────────────────────────────────────
class MaterialIn(BaseModel):
    name: str = Field(min_length=1, max_length=200)
    unit: Optional[str] = None
    category: Optional[str] = None
    # ★기본값은 `unconfirmed`다 — 새로 만든 종이 곧바로 승인분 행세를 하지 않는다(계약 §2-2).
    status: Literal["unconfirmed", "approved"] = "unconfirmed"
    excel_label: Optional[str] = None
    match_rule: Optional[str] = None
    form_factor: Optional[str] = None
    part: Optional[str] = None
    note: Optional[str] = None


class MaterialPatch(BaseModel):
    """부분 갱신. **주지 않은 필드는 안 건드린다**(라우터가 `exclude_unset=True`로 거른다).

    ★정확히 말하면(적대 리뷰 1R P2-7): 「보내지 않은 필드」와 「명시적으로 `null`을 보낸
    필드」는 다르다 — 앞은 무시되고, **뒤는 실제로 그 칸을 비운다.** 초판 주석은 이 둘을 뭉쳐
    *"None을 「비워라」로 읽지 않는다"*고 적었는데 그건 사실이 아니었다. 화면은 지금 명시적
    `null`을 보내지 않으므로 동작은 그대로고, 고친 것은 **주석의 거짓말**이다.
    """

    name: Optional[str] = None
    unit: Optional[str] = None
    category: Optional[str] = None
    status: Optional[Literal["unconfirmed", "approved"]] = None
    excel_label: Optional[str] = None
    match_rule: Optional[str] = None
    form_factor: Optional[str] = None
    part: Optional[str] = None
    note: Optional[str] = None


class LinkIn(BaseModel):
    """원장 라인 → 부자재 종 **사람이 하는 확정**(계약 §5-2)."""

    import_invoice_line_id: int
    note: Optional[str] = None


class AdoptIn(BaseModel):
    """엑셀 참고값 채택 시 남길 한 줄 — 근거 보존용(선택)."""

    note: Optional[str] = None


class PickIn(BaseModel):
    """사람이 고른 **원가표 항목** (계약 §0-E-2, D-CPP-59).

    ★이 스키마가 존재한다는 것 자체가 개정 4의 요점이다 — 화면은 개정 전에도 「사람이
    고른다」고 말했는데 고른 것을 받아 줄 자리가 없었다.
    """

    item_id: int


class LineMaterialIn(BaseModel):
    """구성 한 줄이 가리킬 **부자재 종** (설계 Q6).

    ★단가를 받지 않는다 — 종을 바꾸면 그 종의 단가가 따라올 뿐이다. 여기서 단가를 받으면
    「이 레시피에서만 다른 값」이 생기고, 그게 계약 §0-A‴이 «축 부족»이라 진단한 병을
    다시 만드는 길이다(같은 부자재가 다른 값 = 실은 다른 부자재).
    """

    material_id: int


class AbsentIn(BaseModel):
    """「원가표에 없음」 확인에 남기는 사유 (계약 합격 19).

    ★사유를 받는 이유: 비필름 48건이 같은 목록에 뜨므로(계약 §0-E-4) 「애초에 필름이 아니다」와
    「필름인데 원가표에 없다」가 갈려 기록돼야 다음 사람이 같은 판단을 다시 안 한다.
    """

    note: Optional[str] = None


class ManualPriceIn(BaseModel):
    """국내 구매 부자재 등 원장 파생이 불가한 종의 단가 (계약 §4 하이브리드 ②)."""

    unit_price_ex_vat: Optional[Decimal] = None
    unit_price_inc_vat: Optional[Decimal] = None
    supplier: Optional[str] = None
    effective_date: Optional[date] = None
    note: Optional[str] = None


# ──────────────────────────────────────────────
# 헬퍼
# ──────────────────────────────────────────────
def _guard(fn, *args, **kwargs):
    try:
        return fn(*args, **kwargs)
    except M.CostMenuConflict as exc:
        raise HTTPException(status_code=409, detail=str(exc))
    except M.CostMenuError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


# ──────────────────────────────────────────────
# 엔드포인트
# ──────────────────────────────────────────────
@router.get("/materials")
def list_materials(db: Session = Depends(get_db)):
    """부자재 종 전건 + 단가 이력(로트별).

    화면 탭1의 원료다 — 「승인 상태」·「로트 수」·「최신 단가」가 전부 여기서 온다.
    """
    return {"items": M.list_materials(db)}


@router.post("/materials", status_code=201)
def create_material(body: MaterialIn, db: Session = Depends(get_db)):
    m = _guard(M.create_material, db, **body.model_dump())
    db.commit()
    return M.payload_with_usage(db, M.get_material(db, m.id))


@router.get("/materials/{material_id}")
def get_material(material_id: int, db: Session = Depends(get_db)):
    try:
        m = M.get_material(db, material_id)
    except M.CostMenuError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    return M.payload_with_usage(db, m)


@router.patch("/materials/{material_id}")
def patch_material(material_id: int, body: MaterialPatch, db: Session = Depends(get_db)):
    fields = body.model_dump(exclude_unset=True)
    if not fields:
        raise HTTPException(status_code=400, detail="바꿀 필드가 없다.")
    m = _guard(M.update_material, db, material_id, **fields)
    db.commit()
    m = M.get_material(db, material_id)
    return M.payload_with_usage(db, m)


@router.get("/ledger-material-lines")
def ledger_material_lines(include_products: bool = False, db: Session = Depends(get_db)):
    """확정 수입건의 부자재 라인 전건 + 링크 상태 + **제안**.

    ★제안은 제안이다 — 이 GET은 아무것도 연결하지 않는다. 연결은 아래 POST를 화면의
    「연결」 버튼이 부를 때만 생긴다(계약 §5-2: 확정은 사람).
    ★미매칭 라인도 빠짐없이 실린다 — 안 보이면 단가 이력이 조용히 비어 있게 된다.
    ★`include_products=true`면 수입 완제품(`product`) 라인도 함께 온다(계약 D-CPP-61) —
    기본이 false인 이유·이미 연결된 라인이 항상 실리는 이유는 서비스 docstring 참조.
    """
    return {"items": M.ledger_material_lines(db, include_products=include_products)}


@router.post("/materials/{material_id}/prices/link", status_code=201)
def link_price(material_id: int, body: LinkIn, db: Session = Depends(get_db)):
    p = _guard(
        M.link_ledger_line, db, material_id, body.import_invoice_line_id, body.note
    )
    db.commit()
    m = M.get_material(db, material_id)
    return {"linked_price_id": p.id, "material": M.payload_with_usage(db, m)}


@router.post("/materials/{material_id}/prices/{price_id}/refresh")
def refresh_price(material_id: int, price_id: int, db: Session = Depends(get_db)):
    """어긋난 `ledger` 단가 행을 **원장 현재값으로 다시 맞춘다** (적대 리뷰 1R P1-2).

    ★이게 없으면 원장이 스스로 고칠 길이 없다: 환율 정정 후 재확정으로 원장 값이 바뀌어도
    저장 행은 옛 값 그대로이고, 다시 연결하려 하면 유일 제약 때문에 409였다.
    ★품목이 달라진 행(rowid 재사용)은 **여기서 거부한다** — 갱신은 「같은 것의 새 값」을
    옮기는 일이지 「다른 것」을 삼키는 일이 아니다. 그건 사람이 해제·재연결한다.
    """
    p, before = _guard(M.refresh_ledger_price, db, material_id, price_id)
    db.commit()
    m = M.get_material(db, material_id)
    return {
        "price_id": p.id,
        # 갱신 «전» 판정을 함께 낸다 — 화면이 「무엇이 어긋나 있었나」를 말할 수 있어야 한다.
        "was": M.check_payload(before),
        "material": M.payload_with_usage(db, m),
    }


@router.post("/materials/{material_id}/prices", status_code=201)
def add_manual_price(material_id: int, body: ManualPriceIn, db: Session = Depends(get_db)):
    p = _guard(M.add_manual_price, db, material_id, **body.model_dump())
    db.commit()
    m = M.get_material(db, material_id)
    return {"price_id": p.id, "material": M.payload_with_usage(db, m)}


@router.delete("/materials/{material_id}/prices/{price_id}")
def delete_price(material_id: int, price_id: int, db: Session = Depends(get_db)):
    _guard(M.delete_price, db, material_id, price_id)
    db.commit()
    m = M.get_material(db, material_id)
    return {"deleted": True, "id": price_id, "material": M.payload_with_usage(db, m)}


@router.get("/settings")
def list_settings(db: Session = Depends(get_db)):
    """설정 전건 — `confirmed=false`가 화면의 자백 배지 원료다(계약 §9-1·합격 8)."""
    return {"items": M.list_settings(db)}


# ──────────────────────────────────────────────
# 평가방법 확인·변경 + 이력 (D-CPP-60 갈래① · 합격 ②)
# ──────────────────────────────────────────────
class SettingUpdateIn(BaseModel):
    value: Optional[str] = None
    confirmed: Optional[bool] = None
    actor: Optional[str] = Field(default=None, max_length=50)
    note: Optional[str] = None


@router.get("/settings/history")
def setting_history(limit: int = 50, db: Session = Depends(get_db)):
    """설정 변경 이력. 비어 있으면 「아직 바꾼 적 없음」이고 그건 사실이다."""
    return {"items": M.list_setting_history(db, limit=limit)}


# ──────────────────────────────────────────────
# `product_master.cost_price` 변경 이력 (계약 D-CPP-64 §4 S1-①)
#
# ★읽기 전용이다. 이 라우터는 `cost_price`를 쓰지 않는다(모듈 헤더) — 쓰는 문은 S3의 컷오버
#   경로 한 벌뿐이고 아직 없다. 여기 있는 이유는 **Jino가 원가를 보는 자리가 원가 메뉴**이기
#   때문이다(계약 §2-0). 이력이 products API 밑에 있으면 원가를 보러 온 사람이 못 찾는다.
# ──────────────────────────────────────────────
@router.get("/price-history")
def cost_price_history(
    limit: int = 100,
    internal_sku: Optional[str] = None,
    path: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """`cost_price`가 움직인 사건 목록.

    ★0건은 「원가가 안 바뀌었다」가 **아니다** — 이 표는 배포 시점부터 쌓이므로 대개
      「아직 시작 안 됐다」다. 응답이 `empty_reason`·`started_at`으로 그 둘을 갈라 말하고
      화면은 그 문장을 그대로 띄운다(교훈 #123 — 발견 0건과 실행 안 됨은 같은 숫자로 보인다).
    """
    return CPH.list_cost_price_history(
        db, limit=limit, internal_sku=internal_sku, path=path
    )


@router.post("/settings/{key}")
def update_setting(
    key: str, body: SettingUpdateIn, db: Session = Depends(get_db)
):
    """설정 1건 변경 — **값이 안 바뀌어도 이력이 남는다**(계약 §4-②).

    ★「선입선출 재확인」은 값 변경이 아니라 «사람이 확인했다»는 사건이고, §74의 「신고한
    방법」 확인 기록이 곧 그 사건이다. 값 비교로 걸러 버리면 확인 행위가 사라진다.
    """
    try:
        out = M.update_setting(
            db,
            key,
            value=body.value,
            confirmed=body.confirmed,
            actor=body.actor,
            note=body.note,
        )
    except M.CostMenuError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    db.commit()
    return out


# ──────────────────────────────────────────────
# 단가 자동 갱신 (D-CPP-60 갈래② · 합격 ③④⑤)
# ──────────────────────────────────────────────
@router.get("/auto-refresh/runs")
def auto_refresh_runs(limit: int = 20, db: Session = Depends(get_db)):
    """자동 갱신 회전 이력 — **`updated=0`인 회전도 실린다**(합격 ④ 침묵 금지).

    목록이 비어 있으면 「바뀔 게 없었다」가 아니라 **「한 번도 안 돌았다」**다. 화면은 그
    둘을 구별해 말해야 한다.
    """
    return {"items": AR.recent_runs(db, limit=limit)}


@router.get("/auto-refresh/queue")
def auto_refresh_queue(db: Session = Depends(get_db)):
    """「연결 대기」 큐 — 자동이 **안 건드리고 사람에게 올린** 라인(합격 ⑤ · 계약 §7-4)."""
    return {"items": AR.pending_queue(db)}


@router.post("/auto-refresh/run")
def auto_refresh_run(db: Session = Depends(get_db)):
    """지금 1회전 — 화면의 「지금 검사」 버튼. 크론을 기다리지 않고 사람이 확인할 길이다.

    ★이 경로도 «사람이 만든 짝의 반복»만 한다 — 수동 실행이 게이트를 여는 문이 되면 안 된다.
    """
    # ★P2-4(적대 리뷰 1R): 로트 확정 경로와 «같은 방어»를 둔다. `run()`은 라인별 예외를
    #   이미 savepoint 안에서 잡지만, 회전 «층 자체»가 터지면(설정 오류 등) 여기서 세션이
    #   오염된 채 남는다. 일관성이 곧 다음 사람의 예측 가능성이다.
    try:
        result = AR.run(db, trigger=AR.TRIGGER_MANUAL)
        db.commit()
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        raise HTTPException(status_code=500, detail=f"자동 갱신 회전 실패: {exc}")
    return {
        "run_id": result.run_id,
        "trigger": result.trigger,
        "checked": result.checked,
        "updated": result.updated,
        "failed": result.failed,
        "queued": result.queued,
    }


# ──────────────────────────────────────────────
# 레시피·표준원가 (S2)
# ──────────────────────────────────────────────
def _sheet_rows(upload: UploadFile, wanted: str, what: str) -> list[tuple]:
    """업로드된 .xlsx에서 시트 하나를 행 목록으로. **파일을 여는 것은 이 층의 일이다** —
    파서는 DB도 IO도 모르는 순수 SA라 파일을 안 받는다(계약 §2-6).

    ★시트 이름이 안 맞으면 **첫 시트로 넘어가지 않고 거부한다** — 엉뚱한 시트를 조용히 읽어
    「0건 파싱됨」으로 끝나는 것이 가장 나쁜 결말이기 때문이다(사용자는 성공으로 읽는다).
    """

    try:
        wb = load_workbook(BytesIO(upload.file.read()), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(400, detail=f"{what} 파일을 열 수 없다: {exc}")
    try:
        if wanted not in wb.sheetnames:
            raise HTTPException(
                400,
                detail=f"{what}에 「{wanted}」 시트가 없다 (있는 시트: {', '.join(wb.sheetnames)})",
            )
        return list(wb[wanted].iter_rows(values_only=True))
    finally:
        wb.close()


@router.post("/recipes/import")
def import_recipes(
    cost_file: Optional[UploadFile] = File(
        None, description="원가 정본 — 「제품 원가표」 시트"
    ),
    mapping_file: Optional[UploadFile] = File(
        None, description="매핑 정본 — 「원가 매핑」 시트"
    ),
    db: Session = Depends(get_db),
):
    """엑셀 → 레시피·링크 **초안**(계약 §5-3 탭2).

    ★아무것도 승인하지 않는다. 단가도 만들지 않는다 — 만들어지는 것은 `draft` 레시피와
    링크, 그리고 `unconfirmed` 부자재 종뿐이다(계약 §2-2·§3).
    ★이미 `approved`인 레시피는 **건너뛴다** — 재수입이 승인분을 덮지 않는다.
    ★**둘 중 하나만 올려도 된다**(Jino 2026-08-24: *"여기서 둘중에 하나만도 업데이트가
    되게 해줘"*). 어느 절반이 갱신되고 어느 절반이 그대로인지는 `updated_halves`·
    `untouched`로 응답에 실려 나간다 — 소유권 규율은 `R.import_drafts` docstring이 정본이다.
    """

    if cost_file is None and mapping_file is None:
        raise HTTPException(
            status_code=400, detail="원가 정본과 매핑 정본 중 최소 하나는 올려야 한다."
        )

    cost_rows = (
        _sheet_rows(cost_file, "제품 원가표", "원가 정본")
        if cost_file is not None
        else None
    )
    mapping_rows = (
        _sheet_rows(mapping_file, "원가 매핑", "매핑 정본")
        if mapping_file is not None
        else None
    )
    out = _guard(R.import_drafts, db, cost_rows, mapping_rows)
    db.commit()
    return out


@router.get("/recipes")
def list_recipes(form_factor: Optional[str] = None, db: Session = Depends(get_db)):
    """레시피 전건 + 각각의 표준원가(계산 가능하면). 미승인은 `reason`이 왜인지 말한다."""
    return {"items": R.list_recipes(db, form_factor=form_factor)}


@router.get("/recipes/{recipe_id}")
def get_recipe(recipe_id: int, db: Session = Depends(get_db)):
    try:
        r = R.get_recipe(db, recipe_id)
    except M.CostMenuError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    return R.recipe_payload(db, r, with_links=True)


@router.post("/recipes/{recipe_id}/approve")
def approve_recipe(recipe_id: int, db: Session = Depends(get_db)):
    """Jino가 눈으로 보고 누르는 확정(계약 §2-2). 이 순간부터 표준원가가 저장된다."""
    r = _guard(R.approve_recipe, db, recipe_id)
    db.commit()
    return R.recipe_payload(db, R.get_recipe(db, r.id), with_links=True)


@router.post("/recipes/{recipe_id}/unapprove")
def unapprove_recipe(recipe_id: int, db: Session = Depends(get_db)):
    r = _guard(R.unapprove_recipe, db, recipe_id)
    db.commit()
    return R.recipe_payload(db, R.get_recipe(db, r.id), with_links=True)


@router.post("/recipes/{recipe_id}/adopt-excel-prices")
def adopt_excel_prices(recipe_id: int, body: AdoptIn | None = None, db: Session = Depends(get_db)):
    """엑셀 참고값 → `manual` 단가로 **채택**(계약 §3이 허용한 유일한 유입 경로).

    ★이미 단가가 있는 종은 건드리지 않는다 — 원장 파생 단가를 엑셀로 덮지 않는다(§2-1).
    """
    out = _guard(R.adopt_excel_prices, db, recipe_id, body.note if body else None)
    db.commit()
    r = R.get_recipe(db, recipe_id)
    return {**out, "recipe": R.recipe_payload(db, r, with_links=True)}


@router.get("/recipes/{recipe_id}/cost-table-items")
def cost_table_items(recipe_id: int, db: Session = Depends(get_db)):
    """이 레시피에 붙일 수 있는 **원가표 항목 전건 목록** (계약 합격 18 · D-CPP-59).

    ★개정 4 전까지 화면은 「후보 N건 — 사람이 고른다」고 말하면서 **고를 길을 안 줬다**
    (엔드포인트 17개 전수에 0건, 계약 §0-E-1 ③). 이 엔드포인트가 그 길이다.
    ★`suggested`는 제안 라벨이지 확정이 아니다 — 제안 0건이어도 목록은 나온다.
    """
    return _guard(R.list_cost_table_items, db, recipe_id)


@router.post("/recipes/{recipe_id}/pick-cost-table-item")
def pick_cost_table_item(recipe_id: int, body: PickIn, db: Session = Depends(get_db)):
    """사람이 고른 원가표 항목을 구성으로 확정한다 — **재업로드 없이 즉시**(계약 합격 18).

    ★픽은 승인이 아니다(status는 draft 유지) — §2-2의 사람 확정 지점을 한 클릭에 접지 않는다.
    """
    recipe = _guard(R.pick_cost_table_item, db, recipe_id, body.item_id)
    db.commit()
    return {"recipe": R.recipe_payload(db, recipe, with_links=True)}


@router.post("/recipes/{recipe_id}/unpick-cost-table-item")
def unpick_cost_table_item(recipe_id: int, db: Session = Depends(get_db)):
    """픽을 되돌린다 — 되돌릴 길이 없으면 사람이 고르기를 주저한다."""
    recipe = _guard(R.unpick_cost_table_item, db, recipe_id)
    db.commit()
    return {"recipe": R.recipe_payload(db, recipe, with_links=True)}


@router.patch("/recipes/{recipe_id}/lines/{line_id}/material")
def swap_line_material(
    recipe_id: int, line_id: int, body: LineMaterialIn, db: Session = Depends(get_db)
):
    """구성 한 줄이 가리키는 종을 사람이 바꾼다 (설계 Q6).

    ★이 엔드포인트가 없어서 prod 레시피 45·97(SKU 8개)이 **각 196.9원 과대**인 채로
    고칠 길이 없었다 — 구성이 바뀌는 길은 업로드 통짜 재생성과 픽뿐이었고, 둘 다
    「이 한 줄만」을 못 한다.

    ★`PATCH`인 이유: 줄을 지우고 다시 만드는 게 아니라 **한 필드를 고치는 것**이다.
    지우고 만들면 줄 id가 바뀌어 감사 흔적(`note`에 쌓이는 교체 이력)이 끊긴다.
    """
    recipe = _guard(R.swap_line_material, db, recipe_id, line_id, body.material_id)
    db.commit()
    return {"recipe": R.recipe_payload(db, recipe, with_links=True)}


@router.post("/recipes/{recipe_id}/confirm-cost-table-absent")
def confirm_cost_table_absent(
    recipe_id: int, body: AbsentIn | None = None, db: Session = Depends(get_db)
):
    """「원가표에 없음」을 사람이 **명시적으로** 확인한다 (계약 합격 19).

    ★이 칸이 없으면 「다 보고 없다고 판정했다」와 「아직 아무도 안 봤다」가 화면에서 같아진다.
    """
    recipe = _guard(
        R.confirm_cost_table_absent, db, recipe_id, body.note if body else None
    )
    db.commit()
    return {"recipe": R.recipe_payload(db, recipe, with_links=True)}


@router.get("/cost-table-items")
def cost_table_census(db: Session = Depends(get_db)):
    """원가표 항목 **전건** — 홈 탭 「할 일 인박스」의 넷째 묶음 분모 (D-CPP-62 S2).

    ★`/recipes/{id}/cost-table-items`와 **다른 질문**이다: 저건 「이 레시피에 붙일 수 있는
    항목」(폼팩터 버킷)이고, 이건 「지금 사람 손을 기다리는 항목이 전부 몇 건인가」다.
    레시피를 하나도 안 고른 첫 화면에서 나와야 하는 숫자라 레시피별 경로로는 못 얻는다.

    ★읽기 전용이다 — 이 엔드포인트는 아무것도 쓰지 않는다.
    """
    return R.cost_table_census(db)


@router.get("/board")
def standard_cost_board(db: Session = Depends(get_db)):
    """표준원가 보드 — SKU별 표준원가 · 현 `cost_price` 대조 · 격차(계약 §5-3 탭3)."""
    return R.board(db)


@router.post("/roundtrip/download")
def roundtrip_download(db: Session = Depends(get_db)):
    """왕복 표를 **엑셀 파일로 내려보낸다** (계약 D-CPP-62 S3).

    ★**GET이 아니라 POST다.** 이 호출은 스냅샷 행을 만든다 — 즉 상태를 바꾼다. GET으로 두면
    브라우저 프리페치·링크 미리보기·크롤러가 스냅샷을 찍게 되고, 그러면 S4의 「내가 받았을 때」
    축이 사람이 안 받은 순간들로 오염된다. (같은 내용이면 재발급이라 실질은 멱등이지만,
    그건 최적화지 «아무것도 안 쓴다»는 뜻이 아니다.)

    ★★**필터 인자를 받지 않는다 — 언제나 전건이다.**
      화면의 왕복 표에는 폼팩터·「단가 없음만」·「모순만」 필터가 있다. 그 필터를 여기로 넘겨
      **부분집합** 파일이 나가면, 그 파일을 그대로 재업로드했을 때 빠져 있던 종이 전부 S4의
      「사라짐」 묶음에 선다 — **확인 클릭 한 번이 백여 종을 비활성화**한다.
      그래서 인자를 「무시」하는 게 아니라 **아예 갖지 않는다**: 안 받으면 실수로도 못 넘긴다.

    ★파일 안의 값은 스냅샷에서만 만든다(`build_workbook(snap)`) — DB를 다시 읽지 않는다.
      둘이 각자 조회하면 그 사이의 변경이 조용히 끼어들어, 무수정 재업로드가 「변경 N건」으로
      서는 유령 diff가 난다.
    """
    snap = RT.build_snapshot(db)
    db.commit()

    buf = RT.build_workbook(snap)
    name = RT.filename(snap)
    # 파일명이 한글이라 RFC 5987로 싣고 ASCII 폴백을 함께 준다. 그래도 스냅샷 ID의 정본은
    # `_meta` 시트지 파일명이 아니다 — 파일명은 다른 이름으로 저장하는 순간 사라진다.
    disposition = (
        f'attachment; filename="{RT.snapshot_code(snap)}.xlsx"; '
        f"filename*=UTF-8''{quote(name)}"
    )
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": disposition,
            # 화면이 「방금 받은 파일이 어느 스냅샷인가」를 파일을 열지 않고도 말할 수 있게 한다.
            "X-Snapshot-Id": RT.snapshot_code(snap),
            "X-Snapshot-Rows": str(snap.row_count),
            "Access-Control-Expose-Headers": "X-Snapshot-Id, X-Snapshot-Rows",
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# 매입 완제품 단가 — 제안(쓰기 없음) → 묶음 확인 → 보드 (계약 D-CPP-63 S1 3/3)
#
# ★엔드포인트를 «미리보기»와 «확인» 둘로 가르는 것이 계약 §4 S1 첫 항목
#   (*"확인 전까지 아무 값도 안 써진다"*)의 집행이다. 하나로 합치면 업로드가 곧 적재가
#   되고, 그 순간 「사람의 확인 클릭이 곧 분류」라는 이 계약의 축이 사라진다.
# ─────────────────────────────────────────────────────────────────────────────


class PurchasedConfirmIn(BaseModel):
    """묶음 한 클릭. 화면은 «무엇을 눌렀는지»만 보내고 대상 여부는 서버가 다시 판정한다."""

    internal_skus: list[str]
    price: Decimal
    source_file: str
    source_names: Optional[dict[str, str]] = None
    note: Optional[str] = None


def _money(v) -> Optional[str]:
    """Decimal → 문자열. **float로 내보내지 않는다.**

    ★적대 리뷰 P2-7이 잡은 자리: 이 라우터는 Pydantic 모델이 아니라 raw dict를 돌려주므로
    `jsonable_encoder`가 `Decimal`을 **float**로 바꾼다(실측: `2694.5` number). 원가 시스템에서
    돈을 float로 내보내는 것은 그 자체로 결함이고, 이 저장소의 다른 원가 엔드포인트는 전부
    문자열이다(`CostMaterial.latest_price_ex_vat` 등) — 축마다 타입이 다르면 프론트가
    「어느 축은 string, 어느 축은 number」를 외워야 한다.
    """

    if v is None:
        return None
    # ★2자리로 못 박는다 — 저장 컬럼이 `Numeric(14,2)`라 표시와 저장이 갈리면 안 된다
    #   (파일에서 온 `Decimal("922")`가 화면엔 「922」, DB엔 「922.00」으로 서는 상태).
    try:
        return str(Decimal(v).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
    except InvalidOperation:
        # ★★자릿수가 Decimal 컨텍스트를 넘는 비정상 값(예: 1e30)에서 `quantize`가 던진다.
        #   **표시 한 칸 때문에 업로드 전체를 500으로 죽이지 않는다** — 적대 리뷰 2R가 잡은
        #   자리이고, 그 500은 「원가 칸 하나가 이상하다」가 아니라 **이유 없는 실패**로 보인다.
        #   원문을 그대로 보여주고 판단은 사람에게 남긴다(계약 §2 「없음 ≠ 0」과 같은 결).
        return str(v)


def _sku_out(s: PP.SkuProposal) -> dict:
    return {
        "internal_sku": s.internal_sku,
        "product_name": s.product_name,
        "source_product_name": s.source_product_name,
        "file_price": _money(s.file_price),
        "is_placeholder": s.is_placeholder,
        "current_cost_price": _money(s.current_cost_price),
        "diff": _money(s.diff),
        "recipe_id": s.recipe_id,
        "recipe_name": s.recipe_name,
        "excluded_reason": s.excluded_reason,
        "approved_price": _money(s.approved_price),
    }


@router.post("/purchased-prices/preview")
def preview_purchased_prices(
    price_file: UploadFile = File(..., description="원가 열을 가진 매핑 파일(08-07판 계열)"),
    db: Session = Depends(get_db),
):
    """업로드 → 제안. **DB를 한 글자도 바꾸지 않는다.**

    ★`원가` 열이 없는 판(08-22판)은 파서가 `PriceSheetError`로 거부한다 — 400으로
    그대로 올린다. 「0건 파싱됨」으로 조용히 성공하지 않는 것이 요지다(교훈 #123).
    """

    rows = _sheet_rows(price_file, "원가 매핑", "단가 파일")
    try:
        parsed = parse_price_sheet(rows)
    except PriceSheetError as exc:
        raise HTTPException(400, detail=str(exc))

    p = PP.build_proposal(db, parsed, price_file.filename or "(이름 없는 파일)")
    return {
        "source_file": p.source_file,
        # 어느 열을 읽었는지를 화면이 보여준다 — 계약 §3 「위치로 읽지 않는다」의 표면.
        "read_columns": {"name": p.name_label, "price": p.price_label},
        "counts": p.counts(),
        "groups": [
            {
                "recipe_id": g.recipe_id,
                "recipe_name": g.recipe_name,
                "price": _money(g.price),
                "sku_count": g.sku_count,
                "already_approved": g.already_approved,
                "skus": [_sku_out(s) for s in g.skus],
            }
            for g in p.groups
        ],
        "blanks": [_sku_out(s) for s in p.blanks],
        "excluded": [_sku_out(s) for s in p.excluded],
        "unmatched": p.unmatched,
        "anomalies": p.anomalies,
    }


@router.post("/purchased-prices/confirm")
def confirm_purchased_prices(body: PurchasedConfirmIn, db: Session = Depends(get_db)):
    """묶음 확인 — 대상에만 쓴다. 거부는 **세어서 돌려준다.**

    ★`written`과 `skipped`를 함께 돌려주는 이유: 「10건 눌렀는데 7건만 써졌다」가 화면에
    보여야 사람이 금지선이 작동했음을 안다. 조용히 성공으로 응답하면 막은 것과 안 막은 것이
    구분되지 않는다(계약 §3의 표면).
    """

    if not body.internal_skus:
        raise HTTPException(400, detail="확인할 SKU가 비었다")
    res = PP.confirm_group(
        db,
        internal_skus=body.internal_skus,
        price=body.price,
        source_file=body.source_file,
        source_names=body.source_names,
        note=body.note,
    )
    db.commit()
    return {
        "written": res.written,
        "skipped": [{"internal_sku": s, "reason": r} for s, r in res.skipped],
        "board": PP.board_counts(db),
    }


@router.get("/purchased-prices/board")
def purchased_prices_board(db: Session = Depends(get_db)):
    """첫 화면 카운트 — 계약 §4 S1 넷째 항목(「어디까지 왔나」를 세션 없이 읽는다)."""

    return PP.board_counts(db)
