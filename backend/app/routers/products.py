# products.py — 상품 원가표 CRUD + 채널 매핑 + 엑셀 업로드/다운로드
import io
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session, joinedload

from app.database import get_db
from app.models import ProductMaster, ProductChannelMapping, Channel
from app.utils.kst import kst_today
# compute_pnl_reconciliation은 엔드포인트 내부에서 지연 임포트한다 — 앱 로드 시
# product_pnl→profit_calculator/intelligence 임포트 체인이 라우터 등록과 순환을 이뤄 데드락하는 것 회피.
from app.schemas import (
    ProductCreate,
    ProductUpdate,
    ProductOut,
    MappingCreate,
    MappingUpdate,
    MappingOut,
    MappingIngestResult,
    ChannelCoverageOut,
    UnmappedOption,
    ConnCell,
    ConnChannel,
    ConnRow,
    ConnectionMapOut,
)
from app.services.product_mapping_ingest import ingest_master_sheet
from app.services.mapping_coverage import compute_mapping_coverage
from app.services.product_connection_map import build_connection_map

_COVERAGE_UNMAPPED_LIMIT = 50  # 응답 payload 상한(채널당). 나머지는 truncated 카운트로만 표기.

router = APIRouter(prefix="/api/products", tags=["products"])


def _product_to_out(p: ProductMaster) -> ProductOut:
    mappings = []
    for m in p.channel_mappings:
        mappings.append(
            MappingOut(
                id=m.id,
                channel_id=m.channel_id,
                channel_name=m.channel.name if m.channel else None,
                channel_product_id=m.channel_product_id,
                channel_product_name=m.channel_product_name,
                channel_sku=m.channel_sku,
                selling_price=m.selling_price,
                is_active=m.is_active,
                mapping_source=m.mapping_source,
            )
        )
    return ProductOut(
        id=p.id,
        internal_sku=p.internal_sku,
        product_name=p.product_name,
        cost_price=p.cost_price,
        category=p.category,
        memo=p.memo,
        created_at=p.created_at,
        updated_at=p.updated_at,
        mappings=mappings,
    )


@router.get("", response_model=list[ProductOut])
def list_products(db: Session = Depends(get_db)):
    products = (
        db.query(ProductMaster)
        .options(
            joinedload(ProductMaster.channel_mappings).joinedload(
                ProductChannelMapping.channel
            )
        )
        .order_by(ProductMaster.id)
        .all()
    )
    return [_product_to_out(p) for p in products]


@router.post("", response_model=ProductOut, status_code=201)
def create_product(data: ProductCreate, db: Session = Depends(get_db)):
    exists = db.query(ProductMaster).filter_by(internal_sku=data.internal_sku).first()
    if exists:
        raise HTTPException(400, f"SKU '{data.internal_sku}' 이미 존재합니다")
    p = ProductMaster(**data.model_dump())
    db.add(p)
    db.commit()
    db.refresh(p)
    return _product_to_out(p)


@router.put("/{product_id}", response_model=ProductOut)
def update_product(
    product_id: int, data: ProductUpdate, db: Session = Depends(get_db)
):
    p = db.query(ProductMaster).get(product_id)
    if not p:
        raise HTTPException(404, "상품을 찾을 수 없습니다")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(p, k, v)
    db.commit()
    db.refresh(p)
    p = (
        db.query(ProductMaster)
        .options(
            joinedload(ProductMaster.channel_mappings).joinedload(
                ProductChannelMapping.channel
            )
        )
        .filter_by(id=product_id)
        .first()
    )
    return _product_to_out(p)


@router.delete("/{product_id}", status_code=204)
def delete_product(product_id: int, db: Session = Depends(get_db)):
    p = db.query(ProductMaster).get(product_id)
    if not p:
        raise HTTPException(404, "상품을 찾을 수 없습니다")
    db.query(ProductChannelMapping).filter_by(product_id=product_id).delete()
    db.delete(p)
    db.commit()


# ── 채널 매핑 ──
def _active_option_clash(
    db: Session, channel_id: int, channel_product_id: str, exclude_product_id: int
):
    """같은 채널에서 다른 상품이 이미 이 옵션ID를 active로 갖고 있으면 그 매핑을 반환(없으면 None).

    이중귀속(1채널옵션ID→2마스터) 방지 가드. 매핑 생성/편집/재활성화 시 공통 사용(D-12).
    """
    return (
        db.query(ProductChannelMapping)
        .filter(
            ProductChannelMapping.channel_id == channel_id,
            ProductChannelMapping.channel_product_id == channel_product_id,
            ProductChannelMapping.product_id != exclude_product_id,
            ProductChannelMapping.is_active.is_(True),
        )
        .first()
    )


@router.post("/{product_id}/mappings", response_model=MappingOut, status_code=201)
def add_mapping(
    product_id: int, data: MappingCreate, db: Session = Depends(get_db)
):
    p = db.query(ProductMaster).get(product_id)
    if not p:
        raise HTTPException(404, "상품을 찾을 수 없습니다")
    ch = db.query(Channel).get(data.channel_id)
    if not ch:
        raise HTTPException(404, "채널을 찾을 수 없습니다")
    # 유일성 가드(codex P1): 화면에서 수동 추가할 때도 다른 상품이 쓰는 active 옵션ID면 409.
    clash = _active_option_clash(db, data.channel_id, data.channel_product_id, product_id)
    if clash:
        raise HTTPException(
            409,
            f"채널옵션ID '{data.channel_product_id}'는 이미 다른 상품(id={clash.product_id})에 매핑돼 있습니다",
        )
    # 수동 추가 → mapping_source='manual'(codex P1): 자동동기화가 조용히 덮어쓰지 않게(product_sync 가드와 짝).
    m = ProductChannelMapping(
        product_id=product_id, mapping_source="manual", **data.model_dump()
    )
    db.add(m)
    db.commit()
    db.refresh(m)
    return MappingOut(
        id=m.id,
        channel_id=m.channel_id,
        channel_name=ch.name,
        channel_product_id=m.channel_product_id,
        channel_product_name=m.channel_product_name,
        channel_sku=m.channel_sku,
        selling_price=m.selling_price,
        is_active=m.is_active,
        mapping_source=m.mapping_source,
    )


@router.patch("/{product_id}/mappings/{mapping_id}", response_model=MappingOut)
def update_mapping(
    product_id: int,
    mapping_id: int,
    data: MappingUpdate,
    db: Session = Depends(get_db),
):
    """단일 매핑 인라인 편집(상품 연관맵 트랙 S4, D-12).

    유일성 가드: 옵션ID 변경 시 **또는** 비활성→활성 재활성화 시(codex P1), 같은 채널에서
    다른 상품이 이미 그 옵션ID를 active로 갖고 있으면 409(매출·원가 이중귀속 방지). 수동 편집한
    매핑은 mapping_source='manual'로 표시해 쿠팡 상품 자동동기화가 덮어쓰지 않게 한다.
    """
    m = (
        db.query(ProductChannelMapping)
        .filter_by(id=mapping_id, product_id=product_id)
        .first()
    )
    if not m:
        raise HTTPException(404, "매핑을 찾을 수 없습니다")

    fields = data.model_dump(exclude_unset=True)
    new_cpid = fields.get("channel_product_id")
    cpid_changed = new_cpid is not None and new_cpid != m.channel_product_id
    if cpid_changed and not new_cpid.strip():
        raise HTTPException(422, "채널옵션ID는 비울 수 없습니다")

    # 편집 후 유효값(effective) 기준으로 이중귀속 가드.
    eff_cpid = new_cpid if new_cpid is not None else m.channel_product_id
    eff_active = fields.get("is_active", m.is_active)
    reactivating = fields.get("is_active") is True and not m.is_active
    # active 상태로 남거나 되면서, 옵션ID가 바뀌었거나 방금 재활성화된 경우에만 검사
    # (이미 active인 매핑의 가격만 고치는 편집은 기존 충돌을 새로 막지 않음).
    if eff_active and (cpid_changed or reactivating):
        clash = _active_option_clash(db, m.channel_id, eff_cpid, product_id)
        if clash:
            raise HTTPException(
                409,
                f"채널옵션ID '{eff_cpid}'는 이미 다른 상품(id={clash.product_id})에 매핑돼 있습니다",
            )

    for k, v in fields.items():
        setattr(m, k, v)
    m.mapping_source = "manual"  # 수동 편집 → 자동동기화 clobber 방지(D-12)
    db.commit()
    db.refresh(m)
    ch = db.query(Channel).get(m.channel_id)
    return MappingOut(
        id=m.id,
        channel_id=m.channel_id,
        channel_name=ch.name if ch else None,
        channel_product_id=m.channel_product_id,
        channel_product_name=m.channel_product_name,
        channel_sku=m.channel_sku,
        selling_price=m.selling_price,
        is_active=m.is_active,
        mapping_source=m.mapping_source,
    )


@router.delete("/{product_id}/mappings/{mapping_id}", status_code=204)
def delete_mapping(
    product_id: int, mapping_id: int, db: Session = Depends(get_db)
):
    m = (
        db.query(ProductChannelMapping)
        .filter_by(id=mapping_id, product_id=product_id)
        .first()
    )
    if not m:
        raise HTTPException(404, "매핑을 찾을 수 없습니다")
    db.delete(m)
    db.commit()


# ── 매핑 템플릿 다운로드 (주문 데이터 기반) ──
@router.get("/mapping-template")
def download_mapping_template(db: Session = Depends(get_db)):
    """주문 데이터에서 채널별 상품 목록을 추출한 매핑 템플릿 엑셀"""
    from app.models import Order
    from sqlalchemy import func

    # 채널별 고유 상품 집계
    rows = (
        db.query(
            Channel.code,
            Channel.name,
            Order.platform_product_id,
            Order.platform_product_name,
            func.count(Order.id).label("order_count"),
            func.round(func.avg(Order.selling_price), 0).label("avg_price"),
        )
        .join(Channel, Channel.id == Order.channel_id)
        .filter(Order.platform_product_id != "")
        .group_by(Channel.code, Order.platform_product_id)
        .order_by(Channel.code, func.count(Order.id).desc())
        .all()
    )

    channels = db.query(Channel).order_by(Channel.id).all()

    wb = Workbook()

    # 통합 1시트: 상품명 | 원가 | 채널명 | 채널코드 | 상품ID
    ws1 = wb.active
    ws1.title = "원가 매핑"
    ws1.append(["상품명", "원가", "채널명", "채널코드", "채널상품ID"])
    for r in rows:
        ws1.append([
            "",  # 상품명 — 사용자 입력
            "",  # 원가 — 사용자 입력
            r.name,
            r.code,
            r.platform_product_id,
        ])

    # 시트3: 채널 목록 (참고용)
    ws3 = wb.create_sheet("채널 목록")
    ws3.append(["채널코드", "채널명", "플랫폼", "수수료율(%)"])
    for ch in channels:
        ws3.append([ch.code, ch.name, ch.platform, float(ch.commission_rate)])

    # 스타일: 사용자가 채울 셀 강조 (상품명+원가 = 노란색)
    from openpyxl.styles import PatternFill
    yellow = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    for row in ws1.iter_rows(min_row=2, min_col=1, max_col=2):
        for cell in row:
            cell.fill = yellow

    ws1.column_dimensions["A"].width = 50
    ws1.column_dimensions["B"].width = 12
    ws1.column_dimensions["C"].width = 22
    ws1.column_dimensions["D"].width = 22
    ws1.column_dimensions["E"].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ohisell_mapping_template.xlsx"},
    )


# ── 엑셀 다운로드 ──
@router.get("/download")
def download_excel(db: Session = Depends(get_db)):
    products = (
        db.query(ProductMaster)
        .options(
            joinedload(ProductMaster.channel_mappings).joinedload(
                ProductChannelMapping.channel
            )
        )
        .order_by(ProductMaster.id)
        .all()
    )
    channels = db.query(Channel).order_by(Channel.id).all()

    wb = Workbook()
    # 시트1: 상품 원가표
    ws = wb.active
    ws.title = "상품 원가표"
    ws.append(["사내SKU", "상품명", "원가", "카테고리", "메모"])
    for p in products:
        ws.append([p.internal_sku, p.product_name, float(p.cost_price), p.category, p.memo])

    # 시트2: 채널 매핑
    ws2 = wb.create_sheet("채널 매핑")
    ws2.append([
        "사내SKU", "채널코드", "채널상품번호", "채널상품명", "채널SKU", "판매가", "활성"
    ])
    for p in products:
        for m in p.channel_mappings:
            ws2.append([
                p.internal_sku,
                m.channel.code if m.channel else "",
                m.channel_product_id,
                m.channel_product_name,
                m.channel_sku,
                float(m.selling_price),
                "Y" if m.is_active else "N",
            ])

    # 시트3: 채널 목록 (참고용)
    ws3 = wb.create_sheet("채널 목록")
    ws3.append(["채널코드", "채널명", "플랫폼", "수수료율(%)"])
    for ch in channels:
        ws3.append([ch.code, ch.name, ch.platform, float(ch.commission_rate)])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ohisell_products.xlsx"},
    )


# ── 엑셀 업로드 ──
@router.post("/upload")
async def upload_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "xlsx 파일만 업로드 가능합니다")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content))

    results = {"created": 0, "updated": 0, "mappings_created": 0, "errors": []}

    # 시트1: 상품 원가표
    if "상품 원가표" in wb.sheetnames:
        ws = wb["상품 원가표"]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:
                continue
            sku, name, cost, category, memo = (
                str(row[0]).strip(),
                str(row[1]).strip() if row[1] else "",
                row[2],
                str(row[3]).strip() if row[3] else None,
                str(row[4]).strip() if row[4] else None,
            )
            try:
                cost_decimal = Decimal(str(cost)) if cost else Decimal("0")
            except (InvalidOperation, ValueError):
                results["errors"].append(f"행 {row_idx}: 원가 '{cost}' 변환 실패")
                continue

            existing = db.query(ProductMaster).filter_by(internal_sku=sku).first()
            if existing:
                existing.product_name = name
                existing.cost_price = cost_decimal
                existing.category = category
                existing.memo = memo
                results["updated"] += 1
            else:
                db.add(ProductMaster(
                    internal_sku=sku,
                    product_name=name,
                    cost_price=cost_decimal,
                    category=category,
                    memo=memo,
                ))
                results["created"] += 1

    # 시트2: 채널 매핑
    if "채널 매핑" in wb.sheetnames:
        ws2 = wb["채널 매핑"]
        for row_idx, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0] or not row[1]:
                continue
            sku = str(row[0]).strip()
            ch_code = str(row[1]).strip()
            ch_pid = str(row[2]).strip() if row[2] else ""
            ch_pname = str(row[3]).strip() if row[3] else None
            ch_sku = str(row[4]).strip() if row[4] else None
            try:
                price = Decimal(str(row[5])) if row[5] else Decimal("0")
            except (InvalidOperation, ValueError):
                price = Decimal("0")
            is_active = str(row[6]).strip().upper() != "N" if row[6] else True

            product = db.query(ProductMaster).filter_by(internal_sku=sku).first()
            channel = db.query(Channel).filter_by(code=ch_code).first()
            if not product:
                results["errors"].append(f"행 {row_idx}: SKU '{sku}' 없음")
                continue
            if not channel:
                results["errors"].append(f"행 {row_idx}: 채널코드 '{ch_code}' 없음")
                continue

            existing_mapping = (
                db.query(ProductChannelMapping)
                .filter_by(
                    product_id=product.id,
                    channel_id=channel.id,
                    channel_product_id=ch_pid,
                )
                .first()
            )
            if existing_mapping:
                existing_mapping.channel_product_name = ch_pname
                existing_mapping.channel_sku = ch_sku
                existing_mapping.selling_price = price
                existing_mapping.is_active = is_active
            else:
                db.add(ProductChannelMapping(
                    product_id=product.id,
                    channel_id=channel.id,
                    channel_product_id=ch_pid,
                    channel_product_name=ch_pname,
                    channel_sku=ch_sku,
                    selling_price=price,
                    is_active=is_active,
                ))
                results["mappings_created"] += 1

    db.commit()
    return results


# ── 상품 연관맵 마스터 엑셀 업로드 (D-5: 옛 롱포맷 upload-by-name을 대체) ──
@router.post("/upload-by-name", response_model=MappingIngestResult)
async def upload_by_name(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """마스터 엑셀("원가 매핑" 시트, 1행=옵션 1개, 채널명 마커+옵션ID 블록) 업로드.

    상품 연관맵 트랙 S1: docs/tracks/active/track_product-connection-map.md D-1~D-6.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "xlsx 파일만 업로드 가능합니다")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb["원가 매핑"] if "원가 매핑" in wb.sheetnames else wb.active

    result = ingest_master_sheet(db, ws)
    return MappingIngestResult(
        products_created=result.products_created,
        products_updated=result.products_updated,
        mappings_created=result.mappings_created,
        mappings_updated=result.mappings_updated,
        mappings_conflicted=result.mappings_conflicted,
        orders_linked=result.orders_linked,
        unknown_labels=result.integrity.unknown_labels,
        duplicate_product_names=result.integrity.duplicate_product_names,
        duplicate_channel_ids=result.integrity.duplicate_channel_ids,
        mapping_conflicts=result.integrity.mapping_conflicts,
        label_mismatches=result.integrity.label_mismatches,
    )


# ── 상품 연관맵 커버리지 리포트 (상품 연관맵 트랙 S2) ──
@router.get("/mapping-coverage", response_model=list[ChannelCoverageOut])
def mapping_coverage(
    limit: int = _COVERAGE_UNMAPPED_LIMIT, db: Session = Depends(get_db)
):
    """채널별 매핑 보유율 + 주문에는 있으나 연관맵에 없는 옵션ID 목록.

    상품 연관맵 트랙 S2: docs/tracks/active/track_product-connection-map.md.
    주문↔마스터 백필 자체는 S1 업로드 시 relink_unlinked_orders로 이미 수행됨 — 이 엔드포인트는
    현재 커버리지 상태를 조회만 한다(부작용 없음).
    `limit`: 채널당 미매핑 옵션ID 응답 개수(기본 50). truncated 필드로 잘린 만큼을 알 수 있고,
    나머지가 필요하면 limit을 키워 재요청(codex P2: API 계약에서 긴 꼬리를 숨기지 않도록).
    """
    limit = max(0, min(limit, 5000))
    coverages = compute_mapping_coverage(db)
    out = []
    for c in coverages:
        shown = c.unmapped_order_options[:limit]
        out.append(
            ChannelCoverageOut(
                channel_id=c.channel_id,
                channel_code=c.channel_code,
                channel_name=c.channel_name,
                mapped_option_count=c.mapped_option_count,
                order_option_count=c.order_option_count,
                order_option_coverage=c.order_option_coverage,
                unmapped_order_options=[UnmappedOption(**o) for o in shown],
                unmapped_order_options_truncated=len(c.unmapped_order_options) - len(shown),
                total_orders=c.total_orders,
                unlinked_orders=c.unlinked_orders,
                blank_option_id_orders=c.blank_option_id_orders,
            )
        )
    return out


# ── 상품 연결맵 매트릭스 (내부옵션×채널, 트랙 S4 D-12) ──
@router.get("/connection-map", response_model=ConnectionMapOut)
def connection_map(
    q: str | None = Query(None, description="internal_sku·상품명 부분일치 필터"),
    limit: int | None = Query(None, description="반환 행 상한(생략=전체). total_products로 잘림 확인"),
    db: Session = Depends(get_db),
):
    """내부옵션(행) × 채널(열) 매트릭스 + 채널옵션ID 충돌 표면화(읽기전용, 부작용 없음).

    상품 연관맵 트랙 S4: docs/tracks/active/track_product-connection-map.md D-12.
    커버리지 배지는 별도 GET /mapping-coverage(S2)를 함께 소비한다.
    """
    cmap = build_connection_map(db, q=q, limit=limit)
    return ConnectionMapOut(
        channels=[
            ConnChannel(
                channel_id=c.channel_id,
                channel_code=c.channel_code,
                channel_name=c.channel_name,
                platform=c.platform,
                sell_type=c.sell_type,
            )
            for c in cmap.channels
        ],
        rows=[
            ConnRow(
                product_id=r.product_id,
                internal_sku=r.internal_sku,
                product_name=r.product_name,
                cost_price=r.cost_price,
                cells={
                    cid: [
                        ConnCell(
                            mapping_id=cell.mapping_id,
                            channel_product_id=cell.channel_product_id,
                            channel_product_name=cell.channel_product_name,
                            channel_sku=cell.channel_sku,
                            selling_price=cell.selling_price,
                            is_active=cell.is_active,
                            mapping_source=cell.mapping_source,
                            conflict=cell.conflict,
                        )
                        for cell in cells
                    ]
                    for cid, cells in r.cells.items()
                },
                mapped_channel_count=r.mapped_channel_count,
                has_conflict=r.has_conflict,
            )
            for r in cmap.rows
        ],
        total_products=cmap.total_products,
        shown_products=cmap.shown_products,
        conflict_option_count=cmap.conflict_option_count,
    )


# ── 통합 손익 대조원장 (S3 T6, D-11) ──────────────────────────
# ★법인 단위 Wing 계정키만 허용(codex T6 P1, overview.py와 동일 계약): _resolve_account는 company로
#   풀어 그 법인의 3P/RG/1P를 모두 스코프하고, RG 정산/수수료/반품 테이블은 account_key==WING1/2로
#   저장된다(RG1/RG2/ROCKET로는 저장 안 함). RG1/ROCKET을 넘기면 RG 정산이 0으로 누락돼 과소 손익.
#   오픽스=WING1(3P+RG), 오하이테크=WING2(3P+RG+1P). 생략=전 계정 합산(네이버/cafe24 포함).
_PNL_VALID_ACCOUNTS = {"COUPANG_WING1", "COUPANG_WING2"}


def _pnl_parse_date(s: str | None, default):
    if not s:
        return default
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=422, detail=f"잘못된 날짜 형식: {s} (YYYY-MM-DD)")


def _pnl_jsonify(v):
    """Decimal → str(정밀도 보존, 머니), 중첩 dict/list 재귀. 다른 라우터(_jsonify)와 동일 규칙."""
    if isinstance(v, Decimal):
        return str(v)
    if isinstance(v, dict):
        return {k: _pnl_jsonify(x) for k, x in v.items()}
    if isinstance(v, list):
        return [_pnl_jsonify(x) for x in v]
    return v


@router.get("/pnl-reconciliation")
def pnl_reconciliation(
    from_: str | None = Query(None, alias="from"),
    to: str | None = Query(None),
    account: str | None = Query(
        None,
        description=("계정 필터(D-11): COUPANG_WING1(오픽스)·COUPANG_WING2(오하이테크) — 법인 단위 "
                     "키(3P+RG+1P 전부 스코프). 생략=전 계정 합산(네이버/cafe24 포함)."),
    ),
    db: Session = Depends(get_db),
):
    """통합 손익 대조원장(reconciliation-first). 기본 기간=최근 7일(KST).

    ledger: 채널·컴포넌트별 보존 법칙(Σ SKU귀속+Σ잔차==권위 엔진 계정소계). conservation_ok가
    false면 조인/기준 버그로 SKU행 신뢰 불가 → by_sku는 빈 배열(원장 균형 후에만 노출, D-6).
    by_sku: internal_sku 행(채널 분해 + net_profit_allocated_only=SKU 귀속 순익).
    summary: reconciled_net_profit(엔진 총액) / account_adjustment_residual(미매핑+계정조정, 안분 안 함).
    """
    today = kst_today()
    dto = _pnl_parse_date(to, today)
    dfrom = _pnl_parse_date(from_, dto - timedelta(days=6))
    if dfrom > dto:
        raise HTTPException(status_code=422, detail="from이 to보다 늦습니다")
    if account is not None and account not in _PNL_VALID_ACCOUNTS:
        raise HTTPException(
            status_code=422,
            detail=f"잘못된 account: {account} (허용: {', '.join(sorted(_PNL_VALID_ACCOUNTS))} 또는 생략)",
        )
    from app.services.product_pnl import compute_pnl_reconciliation  # 지연 임포트(순환 회피)
    result = compute_pnl_reconciliation(db, dfrom, dto, account)
    return _pnl_jsonify(result)
