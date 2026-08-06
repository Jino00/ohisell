# routers/ad_costs.py — 광고비 조회 + GFA CSV 업로드 + SA/Meta 동기화 + 쿠팡 XLSX 업로드 API
from __future__ import annotations

import csv
import io
import logging
import os
import re
from app.utils.kst import kst_now, kst_today
from datetime import date, timedelta
from decimal import Decimal

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import text

from app.database import get_ad_db, get_db
from app.models import SchedulerState
from app.schemas import AdSpendByOption, AdSpendDaily
from app.services.ad_cost_reader import get_ad_spend_by_option, get_daily_ad_spend

log = logging.getLogger(__name__)

# profit_calculator와 동일한 키워드 목록 (source 생성 기준)
_NAVER_SA_KEYWORDS = [
    "지문방지", "강화유리", "종이질감", "사생활", "갤럭시탭", "아이패드", "아이폰",
    "갤럭시", "셀카봉", "뮤패드", "케이스",
]
_META_KEYWORDS = [
    "지문방지필름", "골프필름", "버디필름", "강화유리", "셀카봉", "문캅스", "일미리케이스",
]
_KEYWORD_ALIAS = {"샐카봉": "셀카봉"}


def _extract_naver_sa_keyword(campaign_name: str) -> str:
    name = campaign_name or ""
    for kw in _NAVER_SA_KEYWORDS:
        if kw in name:
            return kw
    return "기타"


def _extract_meta_keyword(campaign_name: str) -> str:
    name = campaign_name or ""
    for kw in _META_KEYWORDS:
        if kw in name:
            return _KEYWORD_ALIAS.get(kw, kw)
    return "기타"


def _upsert_ad_cost(db: Session, channel_id: int, ad_date: date, spend: Decimal, source: str) -> None:
    """ad_costs 테이블에 특정 날짜/source 레코드를 delete+insert (멱등)."""
    db.execute(
        text("DELETE FROM ad_costs WHERE channel_id = :cid AND source = :src AND ad_date = :dt"),
        {"cid": channel_id, "src": source, "dt": ad_date.isoformat()},
    )
    db.execute(
        text("""
            INSERT INTO ad_costs (channel_id, product_id, ad_date, ad_spend, ad_revenue, source, created_at)
            VALUES (:cid, NULL, :dt, :spend, NULL, :src, datetime('now'))
        """),
        {"cid": channel_id, "dt": ad_date.isoformat(), "spend": str(spend), "src": source},
    )

def _upsert_ad_revenue(db: Session, channel_id: int, ad_date: date, revenue: Decimal, source: str) -> None:
    """전환매출 전용 행 upsert. ad_spend=0(광고비 중복합산 방지), ad_revenue=전환매출."""
    db.execute(
        text("DELETE FROM ad_costs WHERE channel_id = :cid AND source = :src AND ad_date = :dt"),
        {"cid": channel_id, "src": source, "dt": ad_date.isoformat()},
    )
    db.execute(
        text("""
            INSERT INTO ad_costs (channel_id, product_id, ad_date, ad_spend, ad_revenue, source, created_at)
            VALUES (:cid, NULL, :dt, 0, :rev, :src, datetime('now'))
        """),
        {"cid": channel_id, "dt": ad_date.isoformat(), "rev": str(revenue), "src": source},
    )


# 검색광고 전환매출 전용 source (ad_spend=0, ad_revenue=구매 전환매출)
NAVER_SA_CONV_SOURCE = "naver_sa:conv"

router = APIRouter(prefix="/api/ad-costs", tags=["ad-costs"])


@router.get("/daily", response_model=list[AdSpendDaily])
def daily_ad_spend(
    date_from: str = Query(..., description="YYYY-MM-DD"),
    date_to: str = Query(..., description="YYYY-MM-DD"),
    ad_db=Depends(get_ad_db),
):
    """일별 총 광고비 조회"""
    if ad_db is None:
        return []
    return get_daily_ad_spend(
        ad_db,
        date.fromisoformat(date_from),
        date.fromisoformat(date_to),
    )


@router.get("/by-option", response_model=list[AdSpendByOption])
def ad_spend_by_option(
    date_from: str = Query(..., description="YYYY-MM-DD"),
    date_to: str = Query(..., description="YYYY-MM-DD"),
    option_id: str | None = Query(None),
    ad_db=Depends(get_ad_db),
):
    """상품(option_id)별 광고비 집계"""
    if ad_db is None:
        return []
    return get_ad_spend_by_option(
        ad_db,
        date.fromisoformat(date_from),
        date.fromisoformat(date_to),
        option_id,
    )


# 디스플레이 광고비 축의 source 계열.
#   `gfa:쇼핑`   = 수동 CSV(2026-06-04에 멈춤 — 종전 유일 경로)
#   `gfa:advoost`·`gfa:da` = 비즈머니 실차감 API 자동 적재(2026-08-03~, 매일 07:10 KST)
# ★신선도는 **날짜로 판정하지 않는다**(2026-08-06 적대 리뷰). 소진 0인 날은
# naver_display_ad_costs가 행을 만들지 않으므로(음수/0 방지) '행 없음'이 「소진 0」과
# 「수집 실패」를 겸한다 — 소스별로 보면 거짓 빨강, 계열 합집합으로 보면 **거짓 초록**이다.
# 판정은 `_gfa_collection_health`(수집 잡이 도는가)가 한다. 아래 날짜들은 사실 진술이다.
_GFA_AUTO_SOURCES = ("gfa:advoost", "gfa:da")
# 신선도 판정의 실제 대상 — 이 잡이 돌고 있는가(데이터가 아니라 수집기를 본다).
_GFA_SYNC_JOB = "sync_naver_display_ad_costs"   # 매일 07:10 KST
# 하루 1회 잡이라 만 하루(+여유 6h)를 넘겨 성공이 없으면 최소 한 번은 건너뛴 것이다.
_GFA_SYNC_STALE_HOURS = float(os.getenv("GFA_SYNC_STALE_HOURS") or "30")
_GFA_MANUAL_SOURCE = "gfa:쇼핑"


def _naver_channel_id(db: Session) -> int | None:
    row = db.execute(text("SELECT id FROM channels WHERE code = 'NAVER' LIMIT 1")).fetchone()
    return row[0] if row else None


def _gfa_span(db: Session, sources: tuple[str, ...] | None, channel_id: int | None) -> dict:
    """ad_costs의 GFA 계열 적재 구간 요약.

    sources=None이면 `gfa:%` 계열 전체. 그 외에는 명시된 source들만.
    ★SQL 조각이 아니라 **값**을 받는다(적대 리뷰 P2): 호출부가 문자열로 조건을 넘기는
    시그니처면 다음 사람이 거기에 요청 파라미터를 흘려 넣기 쉽다. 지금 안 뚫렸다는 것과
    뚫리기 쉬운 구조라는 것은 다른 얘기다.
    ★channel_id로 좁힌다: profit_calculator._get_gfa_ad_spend_daily는 채널을 거는데 이쪽만
    안 걸면, 다른 채널이 `gfa:` 접두사를 쓰는 순간 이 화면만 조용히 섞인다.
    """
    params: dict = {}
    conds = []
    if sources is None:
        conds.append("source LIKE 'gfa:%'")
    else:
        keys = [f"s{i}" for i in range(len(sources))]
        params.update(dict(zip(keys, sources)))
        conds.append("source IN (%s)" % ", ".join(f":{k}" for k in keys))
    if channel_id is not None:
        conds.append("channel_id = :cid")
        params["cid"] = channel_id
    row = db.execute(
        text(f"""
            SELECT MIN(ad_date), MAX(ad_date), COUNT(DISTINCT ad_date), COALESCE(SUM(ad_spend), 0)
            FROM ad_costs
            WHERE {" AND ".join(conds)}
        """),
        params,
    ).fetchone()
    if not row or row[0] is None:
        return {"has_data": False, "date_from": None, "date_to": None, "days": 0, "total_spend": 0}
    return {
        "has_data": True,
        "date_from": row[0],
        "date_to": row[1],
        "days": int(row[2]),
        "total_spend": int(row[3]),
    }


def _gfa_collection_health(db: Session) -> dict:
    """★신선도 판정 대상을 **데이터가 아니라 수집기**로 바꾼다(2026-08-06 적대 리뷰).

    왜 데이터로 판정하면 안 되나 — `ad_costs`의 **「행 없음」이 두 가지를 뜻하기 때문**이다:
      ① 그날 그 소스의 소진이 0이었다(정상 — 수집기가 0 이하인 날은 행을 안 만든다)
      ② 수집이 실패해 못 넣었다(사고)
    행만 보고는 이 둘을 원리적으로 못 가른다. 그래서
      · 소스별 MAX(ad_date)로 판정하면 → 소진 0인 날을 사고로 오탐(**거짓 빨강**)
      · 계열 합집합 MAX로 판정하면 → 형제 소스 하나가 죽어도 초록(**거짓 초록**)
    앞의 판은 거짓 빨강을 피하려다 거짓 초록을 만들었다. 둘 다 데이터로 판정한 탓이다.

    ★수집기는 **자기가 언제 성공했는지 안다**(`scheduler_state`, 새 테이블 불필요).
      "우리가 물어봤는가"와 "그날 돈을 썼는가"는 독립이므로, 전자로 판정하면 둘 다 안 틀린다.
      → 잡이 오늘 성공했는데 advoost 행이 없다 = **관측했고 소진이 없었다**(정상).
        잡이 며칠째 실패/미실행이다 = 소스에 행이 있든 없든 **사고**.
    """
    row = (
        db.query(SchedulerState)
        .filter(SchedulerState.job_name == _GFA_SYNC_JOB)
        .first()
    )
    if row is None:
        # 잡 자체가 등록돼 있지 않다 — 초록으로 넘기지 않는다(모르면 모른다고 한다).
        return {
            "job_name": _GFA_SYNC_JOB, "registered": False, "enabled": None,
            "last_success_at": None, "last_status": None, "last_status_at": None,
            "last_error": None, "age_hours": None, "stale": True,
            "reason": "수집 잡이 스케줄러에 등록돼 있지 않다 — 자동 수집이 돌지 않는다.",
        }
    last_ok = row.last_run_at          # ★이 컬럼은 '마지막 **성공**' 의미다(D-F)
    age_h = None if last_ok is None else (kst_now().replace(tzinfo=None) - last_ok).total_seconds() / 3600
    # 하루 1회(07:10) 잡이라 만 하루를 넘겨 성공이 없으면 최소 한 번은 건너뛴 것이다.
    stale = (
        last_ok is None
        or (age_h is not None and age_h > _GFA_SYNC_STALE_HOURS)
        or row.is_enabled is False
        or row.last_status == "error"
    )
    if last_ok is None:
        reason = "한 번도 성공한 적이 없다."
    elif row.is_enabled is False:
        reason = "잡이 비활성화돼 있다 — 자동 수집이 돌지 않는다."
    elif row.last_status == "error":
        reason = "마지막 실행이 실패했다."
    elif stale:
        reason = f"마지막 성공이 {age_h:.0f}시간 전이다(하루 1회 잡이라 {_GFA_SYNC_STALE_HOURS}시간 초과는 건너뜀)."
    else:
        reason = "수집 잡이 정상 주기로 성공하고 있다. 소스별 행이 없는 날은 그날 소진이 0이었다는 뜻이다."
    return {
        "job_name": _GFA_SYNC_JOB,
        "registered": True,
        "enabled": bool(row.is_enabled),
        "last_success_at": last_ok.isoformat() if last_ok else None,
        "last_status": row.last_status,
        "last_status_at": row.last_status_at.isoformat() if row.last_status_at else None,
        "last_error": (row.last_error or None) and row.last_error[:500],
        "age_hours": None if age_h is None else round(age_h, 1),
        "stale": bool(stale),
        "reason": reason,
    }


@router.get("/gfa/status")
def get_gfa_status(db: Session = Depends(get_db)):
    """디스플레이(GFA·ADVoost) 광고비 적재 현황.

    ★**배지 판정은 `collection.stale`로 한다** — 데이터(MAX(ad_date))가 아니다.
      `ad_costs`의 '행 없음'은 「소진 0」과 「수집 실패」를 겸해서 그것으로 판정하면
      반드시 한쪽으로 틀린다(자세한 이유는 `_gfa_collection_health` 주석).
    `date_to`·`auto`·`manual`·`by_source`는 **사실 진술**이지 판정 근거가 아니다.
    """
    cid = _naver_channel_id(db)

    out = _gfa_span(db, None, cid)
    out["auto"] = _gfa_span(db, _GFA_AUTO_SOURCES, cid)
    out["manual"] = _gfa_span(db, (_GFA_MANUAL_SOURCE,), cid)

    rows = db.execute(
        text("""
            SELECT source, MIN(ad_date), MAX(ad_date), COUNT(DISTINCT ad_date),
                   COALESCE(SUM(ad_spend), 0)
            FROM ad_costs
            WHERE source LIKE 'gfa:%'
              AND (:cid IS NULL OR channel_id = :cid)
            GROUP BY source
            ORDER BY source
        """),
        {"cid": cid},
    ).fetchall()
    out["by_source"] = [
        {
            "source": r[0],
            "date_from": r[1],
            "date_to": r[2],
            "days": int(r[3]),
            "total_spend": int(r[4]),
        }
        for r in rows
    ]
    # ★판정 근거. 위 날짜들은 사실 진술이고, 초록/빨강은 여기서만 나온다.
    out["collection"] = _gfa_collection_health(db)
    return out


def _recalc_profit_background(date_from: date, date_to: date) -> None:
    """업로드된 날짜 범위의 이익을 백그라운드에서 재계산."""
    from app.database import SessionLocal
    from app.services.profit_calculator import calculate_daily_trend

    db = SessionLocal()
    try:
        calculate_daily_trend(db, None, None, date_from, date_to)
        log.info("GFA 업로드 후 이익 재계산 완료: %s ~ %s", date_from, date_to)
    except Exception as e:
        log.exception("GFA 업로드 후 이익 재계산 실패: %s", e)
    finally:
        db.close()


@router.post("/gfa/upload")
async def upload_gfa_csv(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """GFA(디스플레이) 광고비 CSV 업로드 → ad_costs 저장 (source=gfa:쇼핑, 채널=NAVER).
    파일 형식: theohi11_광고비 보고서_YYYYMMDD_YYYYMMDD.csv
    필수 컬럼: 기간(YYYY.MM.DD.), 총비용(디스플레이 전체=ADVoost 쇼핑+동영상 조회 등).
    총비용 없으면 ADVoost 쇼핑으로 폴백.
    """
    if not file.filename or not file.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="CSV 파일만 업로드 가능합니다.")

    content = await file.read()
    text_content = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text_content))

    naver_row = db.execute(
        text("SELECT id FROM channels WHERE code = 'NAVER' LIMIT 1")
    ).fetchone()
    if not naver_row:
        raise HTTPException(status_code=500, detail="NAVER 채널이 DB에 없습니다.")
    naver_id = naver_row[0]

    inserted = 0
    skipped = 0
    records: list[dict] = []

    for row in reader:
        date_str = row.get("기간", "").strip().rstrip(".")
        if not date_str:
            skipped += 1
            continue
        try:
            ad_date = date.fromisoformat(date_str.replace(".", "-"))
        except ValueError:
            skipped += 1
            continue

        # 총비용 = 디스플레이 전체(ADVoost 쇼핑 + 동영상 조회 등). 동영상 광고비 누락 방지로 총비용 우선.
        spend_str = (row.get("총비용") or row.get("ADVoost 쇼핑") or "0").strip().replace(",", "")
        try:
            spend = Decimal(spend_str)
        except Exception:
            skipped += 1
            continue

        if spend <= 0:
            skipped += 1
            continue

        records.append({"date": ad_date, "spend": spend})

    if not records:
        raise HTTPException(status_code=422, detail="저장할 데이터가 없습니다. 컬럼(기간, ADVoost 쇼핑)을 확인하세요.")

    for rec in records:
        db.execute(
            text("""
                DELETE FROM ad_costs
                WHERE channel_id = :cid AND source = 'gfa:쇼핑' AND ad_date = :ad_date
            """),
            {"cid": naver_id, "ad_date": rec["date"].isoformat()},
        )
        db.execute(
            text("""
                INSERT INTO ad_costs (channel_id, product_id, ad_date, ad_spend, ad_revenue, source, created_at)
                VALUES (:cid, NULL, :ad_date, :spend, NULL, 'gfa:쇼핑', datetime('now'))
            """),
            {"cid": naver_id, "ad_date": rec["date"].isoformat(), "spend": str(rec["spend"])},
        )
        inserted += 1

    db.commit()

    total_spend = sum(r["spend"] for r in records)
    date_from = min(r["date"] for r in records)
    date_to = max(r["date"] for r in records)

    # 업로드된 날짜 범위에 대해 이익 재계산 (백그라운드)
    background_tasks.add_task(_recalc_profit_background, date_from, date_to)

    return {
        "inserted": inserted,
        "skipped": skipped,
        "total_spend": int(total_spend),
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "dates": [r["date"].isoformat() for r in records],
        "recalculation_triggered": True,
    }


@router.post("/naver-sa/sync")
def sync_naver_sa_ad_costs(
    date_from: str = Query(default=None, description="YYYY-MM-DD (기본: 어제)"),
    date_to: str = Query(default=None, description="YYYY-MM-DD (기본: 어제)"),
    db: Session = Depends(get_db),
):
    """네이버 검색광고(SA) 캠페인별 일별 광고비 + 구매 전환매출 → ad_costs 저장.
    source 형식: naver_sa:{키워드}(광고비) / naver_sa:conv(전환매출, ad_spend=0)
    """
    from app.services.naver_sa_ad_fetcher import fetch_campaign_daily_spend, fetch_daily_conversion_revenue

    yesterday = kst_today() - timedelta(days=1)
    d_from = date.fromisoformat(date_from) if date_from else yesterday
    d_to = date.fromisoformat(date_to) if date_to else yesterday

    naver_row = db.execute(
        text("SELECT id FROM channels WHERE code = 'NAVER' LIMIT 1")
    ).fetchone()
    if not naver_row:
        raise HTTPException(status_code=500, detail="NAVER 채널이 DB에 없습니다.")
    naver_id = naver_row[0]

    try:
        campaigns = fetch_campaign_daily_spend(d_from, d_to)
    except Exception as e:
        log.error("Naver SA API 오류: %s", e)
        raise HTTPException(status_code=502, detail=f"Naver SA API 오류: {e}")

    if not campaigns:
        return {"inserted": 0, "skipped": 0, "total_spend": 0, "dates": [], "message": "수집된 캠페인 없음"}

    # 날짜×source 별 집계
    agg: dict[tuple[str, str], Decimal] = {}
    for c in campaigns:
        kw = _extract_naver_sa_keyword(c["campaign_name"])
        source = f"naver_sa:{kw}"
        key = (c["date"], source)
        agg[key] = agg.get(key, Decimal("0")) + c["spend"]

    for (dt_str, source), spend in agg.items():
        _upsert_ad_cost(db, naver_id, date.fromisoformat(dt_str), spend, source)

    # 구매 전환매출(직접+간접) 일별 적재 — RoAS 산출용
    conv_revenue = 0
    try:
        conv_daily = fetch_daily_conversion_revenue(d_from, d_to)
        for dt_str, rev in conv_daily.items():
            _upsert_ad_revenue(db, naver_id, date.fromisoformat(dt_str), rev, NAVER_SA_CONV_SOURCE)
        conv_revenue = int(sum(conv_daily.values()))
    except Exception as e:
        log.warning("Naver SA 전환매출 적재 실패(광고비는 저장됨): %s", e)

    db.commit()

    dates = sorted({k[0] for k in agg})
    total = int(sum(agg.values()))
    log.info("Naver SA 광고비 %d건 + 전환매출 %d원 적재 (%s~%s) 총 %d원", len(agg), conv_revenue, d_from, d_to, total)
    return {
        "inserted": len(agg),
        "skipped": 0,
        "total_spend": total,
        "conv_revenue": conv_revenue,
        "dates": dates,
    }


@router.post("/meta/sync")
def sync_meta_ad_costs(
    date_from: str = Query(default=None, description="YYYY-MM-DD (기본: 어제)"),
    date_to: str = Query(default=None, description="YYYY-MM-DD (기본: 어제)"),
    db: Session = Depends(get_db),
):
    """Meta 캠페인별 일별 광고비 → ad_costs 저장.
    source 형식: meta:{키워드} 또는 meta:기타
    """
    from app.services.meta_ad_fetcher import fetch_campaign_daily_spend

    yesterday = kst_today() - timedelta(days=1)
    d_from = date.fromisoformat(date_from) if date_from else yesterday
    d_to = date.fromisoformat(date_to) if date_to else yesterday

    cafe24_row = db.execute(
        text("SELECT id FROM channels WHERE code = 'CAFE24' LIMIT 1")
    ).fetchone()
    if not cafe24_row:
        raise HTTPException(status_code=500, detail="CAFE24 채널이 DB에 없습니다.")
    cafe24_id = cafe24_row[0]

    try:
        campaigns = fetch_campaign_daily_spend(d_from, d_to)
    except Exception as e:
        log.error("Meta API 오류: %s", e)
        raise HTTPException(status_code=502, detail=f"Meta API 오류: {e}")

    if not campaigns:
        return {"inserted": 0, "skipped": 0, "total_spend": 0, "dates": [], "message": "수집된 캠페인 없음"}

    agg: dict[tuple[str, str], Decimal] = {}
    for c in campaigns:
        kw = _extract_meta_keyword(c["campaign_name"])
        source = f"meta:{kw}"
        key = (c["date"], source)
        agg[key] = agg.get(key, Decimal("0")) + c["spend"]

    for (dt_str, source), spend in agg.items():
        _upsert_ad_cost(db, cafe24_id, date.fromisoformat(dt_str), spend, source)

    db.commit()

    dates = sorted({k[0] for k in agg})
    total = int(sum(agg.values()))
    log.info("Meta 광고비 %d건 적재 (%s~%s) 총 %d원", len(agg), d_from, d_to, total)
    return {
        "inserted": len(agg),
        "skipped": 0,
        "total_spend": total,
        "dates": dates,
    }


# ── 쿠팡 광고비 XLSX 업로드 ──────────────────────────────────────────────────

_SELL_TYPE_TO_CHANNEL_SUFFIX = {
    "3P": "WING",
    "2P": "RG",
    "Retail": "ROCKET",
}


def _build_vendor_channel_map(db: Session) -> dict[tuple[str, str], int]:
    """(vendor_id, channel_suffix) → channel_id 조회 테이블 생성."""
    mapping: dict[tuple[str, str], int] = {}
    for suffix in ("WING", "RG"):
        for num in ("1", "2"):
            code = f"COUPANG_{suffix}{num}"
            vid = os.getenv(f"COUPANG_{suffix}{num}_VENDOR_ID", "").strip()
            if not vid:
                continue
            row = db.execute(text("SELECT id FROM channels WHERE code = :c LIMIT 1"), {"c": code}).fetchone()
            if row:
                mapping[(vid, suffix)] = row[0]
    rocket_row = db.execute(text("SELECT id FROM channels WHERE code = 'COUPANG_ROCKET' LIMIT 1")).fetchone()
    if rocket_row:
        mapping[("*", "ROCKET")] = rocket_row[0]
    return mapping


@router.get("/coupang/status")
def get_coupang_ad_status(db: Session = Depends(get_db)):
    """현재 DB에 적재된 쿠팡 광고비 현황 반환."""
    rows = db.execute(text("""
        SELECT ac.source, c.name, MIN(ac.ad_date), MAX(ac.ad_date), COUNT(*), COALESCE(SUM(ac.ad_spend), 0)
        FROM ad_costs ac
        JOIN channels c ON ac.channel_id = c.id
        WHERE ac.source LIKE 'coupang_%'
        GROUP BY ac.source, c.name
        ORDER BY ac.source
    """)).fetchall()
    if not rows:
        return {"has_data": False, "items": []}
    return {
        "has_data": True,
        "items": [
            {"source": r[0], "channel_name": r[1], "date_from": r[2], "date_to": r[3], "days": r[4], "total_spend": int(r[5])}
            for r in rows
        ],
    }


def _detect_xlsx_format(headers: list) -> dict:
    """헤더 행으로 쿠팡 광고 XLSX 포맷을 감지해 컬럼 인덱스 반환.

    지원 포맷:
      - pa_daily_keyword : 광고비=P(15), 노출수=N(13), 클릭수=O(14)
      - pa_daily_adGroup : 광고비=L(11), 노출수/클릭수는 헤더 탐색
    """
    def _find(keyword: str) -> int:
        for i, h in enumerate(headers):
            if h and keyword in str(h):
                return i
        return -1

    spend_idx = _find("광고비")
    if spend_idx == -1:
        spend_idx = 11  # fallback

    return {
        "spend":   spend_idx,
        "impr":    _find("노출수"),
        "clicks":  _find("클릭수"),
        "orders":  _find("총 주문수(1일)") if _find("총 주문수(1일)") != -1 else _find("전환수"),
        "qty":     _find("총 판매수량(1일)") if _find("총 판매수량(1일)") != -1 else _find("판매수량"),
        "rev":     _find("총 전환매출액(1일)") if _find("총 전환매출액(1일)") != -1 else _find("전환매출액"),
        # 옵션ID 컬럼(keyword 포맷에만 존재 — adGroup 포맷은 -1) → 트랙 D-9 3자 조인 광고축
        "ad_opt":   _find("광고집행 옵션"),       # 비용·노출·클릭 귀속 옵션ID
        "conv_opt": _find("전환매출발생 옵션"),    # 매출·주문 귀속 옵션ID
        # 상품명 — 옵션ID의 사람이 읽는 라벨. 옵션ID 컬럼과 짝(있으면 같이 있다).
        #   "광고집행 상품명"은 "광고집행 옵션ID"와 접두가 같으므로 **전체 어구로** 찾는다.
        "ad_name":   _find("광고집행 상품명"),
        "conv_name": _find("전환매출발생 상품명"),
    }


def ingest_coupang_ad_xlsx_content(
    content: bytes, filename: str, db: Session, *, options_only: bool = False
) -> tuple[dict, date | None, date | None]:
    """쿠팡 광고 XLSX 바이트 → ad_costs + coupang_ad_report + coupang_ad_option_daily 적재.

    수동 업로드 엔드포인트와 토큰 인증 자동 ingest(coupang_ops)가 공유하는 단일 파서(SA).
    반환: (public_result, recalc_from, recalc_to). 이익 재계산 범위는 호출자가 스케줄(여기선 DB 커밋만).
    재계산 불필요(광고비 행 없음) 시 recalc_from/to=None (codex P2 — private 키 누출 방지).
    파일명 형식: {vendor_id}_pa_daily_*.xlsx. 포맷(adGroup/keyword)은 헤더 행에서 자동 감지.

    ★options_only(D-13, 2026-08-03): `coupang_ad_option_daily`만 적재하고 **머니 테이블
      (`ad_costs`·`coupang_ad_report`)은 건드리지 않는다**. 오하이테크(1P 로켓배송)가 이 모드를
      쓴다 — 그 계정의 계정 단위 광고비는 `report/SALES` 페처가 **전체(ALL_DELIVERED) 기준**으로
      쓰는데(D-10), 같은 행을 이 XLSX가 **PA 기준**으로 덮으면 나중에 쓴 쪽이 조용히 이기고
      그 사고는 순이익에서만 드러난다. 정의가 다른 두 writer를 한 행에 붙이지 않는다.
    """
    if not filename or not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="xlsx 파일만 업로드 가능합니다.")

    m = re.match(r"(A\d+)_", filename)
    if not m:
        raise HTTPException(status_code=400, detail="파일명에서 vendor_id를 찾을 수 없습니다. 형식: {vendor_id}_pa_daily_...")
    vendor_id = m.group(1)

    # ★구조 가드(D-13ⓒ): 로켓 벤더의 XLSX가 머니 경로로 들어오는 것을 **코드로** 막는다.
    #   트랙 S1c 배포 체크리스트 ②("A01029796 PA-XLSX 수동업로드 금지")는 문서 규칙이라
    #   수동 업로드 화면에서도, 잘못된 엔드포인트 호출에서도 지켜지지 않을 수 있다.
    #   여기서 막으면 경로가 무엇이든 같은 결론이 난다.
    _rocket_vendor = os.getenv("COUPANG_ROCKET_VENDOR_ID", "").strip()
    if not options_only and _rocket_vendor and vendor_id == _rocket_vendor:
        raise HTTPException(
            status_code=422,
            detail=(
                f"{vendor_id}(1P 로켓배송)의 XLSX는 계정 단위 광고비를 덮어쓸 수 없습니다 — "
                "그 값은 report/SALES 페처가 전체(비-PA 포함) 기준으로 적재합니다(D-10). "
                "옵션 단위만 적재하려면 /api/coupang/ops/rocket/ad-cost/option-ingest 를 쓰세요(D-13)."
            ),
        )

    vendor_channel_map = _build_vendor_channel_map(db)

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 패키지가 설치되지 않았습니다.")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"XLSX 파일을 열 수 없습니다: {e}")

    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col = _detect_xlsx_format(headers)
    log.info("쿠팡 XLSX 포맷 감지: spend=%s impr=%s clicks=%s orders=%s qty=%s rev=%s",
             col["spend"], col["impr"], col["clicks"], col["orders"], col["qty"], col["rev"])

    # ad_costs 집계: (date, channel_id, source) → spend
    agg: dict[tuple[date, int, str], Decimal] = {}
    # coupang_ad_report 집계: (date, sell_type) → metrics
    report_agg: dict[tuple[date, str], dict] = {}
    # coupang_ad_option_daily 집계: (date, vendor_id, sell_type, ad_opt, conv_opt) → metrics (트랙 D-9)
    opt_agg: dict[tuple, dict] = {}
    has_opt = col["ad_opt"] != -1 and col["conv_opt"] != -1
    skipped = 0

    def _cell_int(r: tuple, idx: int) -> int:
        try:
            return int(r[idx] or 0) if idx != -1 and idx < len(r) else 0
        except (TypeError, ValueError):
            return 0

    def _cell_dec(r: tuple, idx: int) -> Decimal:
        try:
            return Decimal(str(r[idx] or 0)) if idx != -1 and idx < len(r) else Decimal("0")
        except Exception:
            return Decimal("0")

    def _cell_name(r: tuple, idx: int) -> str | None:
        """상품명 셀 → 정리된 문자열(빈값·'-'는 None). 컬럼 길이(300) 초과분은 자른다."""
        if idx == -1 or idx >= len(r):
            return None
        v = r[idx]
        if v is None:
            return None
        s = str(v).strip()
        if s in ("", "-"):
            return None
        return s[:300]

    def _norm_opt(v):
        """옵션ID 정규화: 94277472815.0 → '94277472815', 빈값('-'/None)은 None."""
        if v is None:
            return None
        s = str(v).strip()
        if s in ("", "-"):
            return None
        try:
            return str(int(float(s)))
        except (TypeError, ValueError):
            return s

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= col["spend"]:
            skipped += 1
            continue

        date_raw = row[0]
        sell_type = str(row[2] or "").strip() if row[2] else None
        spend_raw = row[col["spend"]]

        if not date_raw or not sell_type:
            skipped += 1
            continue

        try:
            date_str = str(int(date_raw))
            ad_date = date(int(date_str[:4]), int(date_str[4:6]), int(date_str[6:8]))
        except Exception:
            skipped += 1
            continue

        spend = Decimal(str(spend_raw or 0))

        suffix = _SELL_TYPE_TO_CHANNEL_SUFFIX.get(sell_type)
        if suffix is None:
            skipped += 1
            continue

        # ── ad_costs 집계 (spend > 0 인 행만) ──
        if spend > 0:
            if suffix == "ROCKET":
                channel_id = vendor_channel_map.get(("*", "ROCKET"))
            else:
                channel_id = vendor_channel_map.get((vendor_id, suffix))

            if channel_id is not None:
                source = f"coupang_{suffix.lower()}"
                key = (ad_date, channel_id, source)
                agg[key] = agg.get(key, Decimal("0")) + spend

        # ── coupang_ad_report 집계 (노출 지표가 있는 포맷만) ──
        if col["impr"] != -1:
            rkey = (ad_date, sell_type)
            if rkey not in report_agg:
                report_agg[rkey] = {"impr": 0, "clicks": 0, "spend": Decimal("0"),
                                     "orders": 0, "qty": 0, "rev": Decimal("0")}
            g = report_agg[rkey]
            g["impr"]   += _cell_int(row, col["impr"])
            g["clicks"] += _cell_int(row, col["clicks"])
            g["spend"]  += spend
            g["orders"] += _cell_int(row, col["orders"])
            g["qty"]    += _cell_int(row, col["qty"])
            g["rev"]    += _cell_dec(row, col["rev"])

        # ── coupang_ad_option_daily 집계 (옵션ID 컬럼이 있는 keyword 포맷만, 트랙 D-9) ──
        if has_opt:
            ad_opt = _norm_opt(row[col["ad_opt"]] if col["ad_opt"] < len(row) else None)
            conv_opt = _norm_opt(row[col["conv_opt"]] if col["conv_opt"] < len(row) else None)
            if ad_opt or conv_opt:
                # 한쪽만 있으면 동일 옵션으로 폴백(귀속축 결측 방지). 보통 둘은 같음.
                ad_opt = ad_opt or conv_opt
                conv_opt = conv_opt or ad_opt
                okey = (ad_date, vendor_id, sell_type, ad_opt, conv_opt)
                if okey not in opt_agg:
                    opt_agg[okey] = {"impr": 0, "clicks": 0, "spend": Decimal("0"),
                                     "orders": 0, "qty": 0, "rev": Decimal("0"),
                                     "ad_name": None, "conv_name": None}
                o = opt_agg[okey]
                # 상품명: 같은 키의 여러 행(키워드별)이 같은 이름을 반복한다. 빈 행이 섞일 수
                #   있으므로 **처음 만난 non-null을 유지**한다(빈값으로 덮지 않는다).
                if o["ad_name"] is None:
                    o["ad_name"] = _cell_name(row, col["ad_name"])
                if o["conv_name"] is None:
                    o["conv_name"] = _cell_name(row, col["conv_name"])
                o["impr"]   += _cell_int(row, col["impr"])
                o["clicks"] += _cell_int(row, col["clicks"])
                o["spend"]  += spend
                o["orders"] += _cell_int(row, col["orders"])
                o["qty"]    += _cell_int(row, col["qty"])
                o["rev"]    += _cell_dec(row, col["rev"])

    if options_only:
        # 머니 테이블은 쓰지 않는다 → 집계도 버려서 아래 저장·재계산이 구조적으로 못 돌게 한다.
        #   (플래그를 저장 시점마다 다시 확인하는 방식은 한 군데만 빠뜨려도 새는 반면,
        #    원료를 비우면 새는 경로가 남지 않는다.)
        agg = {}
        report_agg = {}
        if not opt_agg:
            raise HTTPException(
                status_code=422,
                detail="옵션 단위 행이 없습니다 — keyword 포맷(옵션ID 컬럼 포함) XLSX가 필요합니다.")
    elif not agg and not report_agg:
        raise HTTPException(status_code=422, detail="저장할 광고 데이터가 없습니다. 파일 형식을 확인하세요.")

    # ad_costs 저장
    for (ad_date, channel_id, source), spend in agg.items():
        _upsert_ad_cost(db, channel_id, ad_date, spend, source)

    # coupang_ad_report 저장
    if report_agg:
        from app.models import CoupangAdReport
        for (ad_date, sell_type), g in report_agg.items():
            existing = db.query(CoupangAdReport).filter(
                CoupangAdReport.report_date == ad_date,
                CoupangAdReport.sell_type == sell_type,
                CoupangAdReport.vendor_id == vendor_id,
            ).first()
            if existing:
                existing.impressions = g["impr"]
                existing.clicks = g["clicks"]
                existing.ad_spend = g["spend"]
                existing.orders = g["orders"]
                existing.sales_qty = g["qty"]
                existing.conversion_revenue = g["rev"]
            else:
                db.add(CoupangAdReport(
                    report_date=ad_date, sell_type=sell_type, vendor_id=vendor_id,
                    impressions=g["impr"], clicks=g["clicks"], ad_spend=g["spend"],
                    orders=g["orders"], sales_qty=g["qty"], conversion_revenue=g["rev"],
                ))

    # coupang_ad_option_daily 저장 (옵션 그레인 — 트랙 D-9 3자 조인 광고축)
    # delete-then-insert: 이 업로드가 커버하는 (vendor_id, 등장 날짜) 범위의 기존 옵션 행을
    # 먼저 제거 후 전량 삽입. 14일 전환 윈도우로 conv_option_id가 나중에 바뀌어 5-튜플 키가
    # 달라질 때 옛 행이 stale로 남아 이중집계되는 것을 차단(codex P2). 같은 파일 재업로드 시엔
    # 같은 날짜를 지우고 동일 재삽입하므로 멱등성 유지.
    if opt_agg:
        from app.models import CoupangAdOptionDaily
        opt_dates = {od for (od, _vid, _st, _ad, _cv) in opt_agg}
        db.query(CoupangAdOptionDaily).filter(
            CoupangAdOptionDaily.vendor_id == vendor_id,
            CoupangAdOptionDaily.report_date.in_(opt_dates),
        ).delete(synchronize_session=False)
        for (od, vid, st, ad_opt, conv_opt), o in opt_agg.items():
            db.add(CoupangAdOptionDaily(
                report_date=od, vendor_id=vid, sell_type=st,
                ad_option_id=ad_opt, conv_option_id=conv_opt,
                ad_product_name=o["ad_name"], conv_product_name=o["conv_name"],
                impressions=o["impr"], clicks=o["clicks"], ad_spend=o["spend"],
                orders=o["orders"], sales_qty=o["qty"], conversion_revenue=o["rev"],
            ))

    db.commit()

    # 날짜 범위 — 머니 집계가 비면(options_only) 옵션 집계에서 뽑는다.
    _date_keys = [k[0] for k in agg] or [k[0] for k in report_agg] or [k[0] for k in opt_agg]
    date_from = min(_date_keys)
    date_to   = max(_date_keys)
    dates = sorted({k[0].isoformat() for k in agg})
    total = int(sum(agg.values()))
    channel_summary: dict[str, int] = {}
    for (_, _, source), spend in agg.items():
        channel_summary[source] = channel_summary.get(source, 0) + int(spend)

    log.info("쿠팡 광고 적재 %s (%s~%s) 광고비=%d원 리포트=%d건 옵션=%d건",
             vendor_id, date_from, date_to, total, len(report_agg), len(opt_agg))
    result = {
        "vendor_id": vendor_id,
        "inserted": len(agg),
        "skipped": skipped,
        "total_spend": total,
        "report_rows": len(report_agg),
        "option_rows": len(opt_agg),
        # ★옵션 합계를 따로 돌려준다: 계정 총액과의 대조가 이 경로의 유일한 자기검증이다
        #   (D-12 실측 기준 0.02% 차이 — 원인 미규명이라 계속 보고 있어야 한다).
        "option_spend": int(sum(o["spend"] for o in opt_agg.values())),
        # 상품명이 실제로 붙은 행 수 — 0이면 헤더가 바뀌었거나 컬럼을 못 찾은 것이다(조용한 실패 방지).
        "option_named_rows": sum(1 for o in opt_agg.values() if o["ad_name"]),
        "options_only": options_only,
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "dates": dates,
        "channel_summary": channel_summary,
        "recalculation_triggered": bool(agg),
    }
    return result, (date_from if agg else None), (date_to if agg else None)


@router.post("/coupang/upload")
async def upload_coupang_ad_xlsx(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """쿠팡 광고 XLSX 수동 업로드 → 파서(SA) 호출 + 이익 재계산 스케줄.

    파일명 형식: {vendor_id}_pa_daily_*.xlsx
    C열(판매방식): Retail=로켓배송, 3P=윙, 2P=로켓그로스
    """
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="xlsx 파일만 업로드 가능합니다.")
    content = await file.read()
    result, recalc_from, recalc_to = ingest_coupang_ad_xlsx_content(content, file.filename, db)
    if recalc_from and recalc_to:
        background_tasks.add_task(_recalc_profit_background, recalc_from, recalc_to)
    return result
