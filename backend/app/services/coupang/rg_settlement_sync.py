# rg_settlement_sync.py — RG 정산 수수료 동기화 Harness (트랙 RG-Fee-Accounting S3)
# 흐름: 계정별 → 쿠키 로드·복호화(재사용) → Wing status/api 호출(SA) → 파싱·검산 → upsert.
# D-8: parser 별도 SA 없음 — 파싱은 이 Harness 책임. D-10: 매출인식일(searchDateType=SALES).
# D-9: 판매수수료(B)+풀필먼트(J) 둘 다 수집. D-13: 방어적 파싱+스키마 드리프트 감지.
# D-12: fixture 기반 테스트(tests/test_rg_settlement_sync.py) — 머니코드라 예외.
# fail-soft: 302/401=status red + 예외 전파(호출자가 catch). 성공=last_success_at 기록.
from __future__ import annotations

import io
import logging
import time
from datetime import date, datetime, timedelta
from decimal import Decimal
from zoneinfo import ZoneInfo

import openpyxl
from sqlalchemy.orm import Session

from app.clients.coupang.inbound import WingAuthError, WingReadError
from app.clients.coupang.rg_settlement import (
    CONFIRMED_SELLER_REPORT_TYPES,
    CoupangWingRgSettlementClient,
)
from app.models import CoupangRgSettlementFee, CoupangWingCookie
from app.utils.crypto import decrypt_secret
from app.utils.kst import kst_now

log = logging.getLogger(__name__)

_KST = ZoneInfo("Asia/Seoul")
RG_ACCOUNTS = ["COUPANG_WING1", "COUPANG_WING2"]

# status/api 응답 → fee_type 매핑 (D-9: 판매수수료+풀필먼트 포함)
# key=응답 필드명, value=(fee_type, 부호: 1=비용[양수저장], -1=환급[음수저장])
#
# ★D-10 라이브 확정(2026-06-09, 원칙22 — WING1 06-01~07 리포트가 레퍼런스 17 §7과 정확 일치):
#   풀필먼트 전체비용(J) = 배송비(fulfillment) + 입출고비(warehousing) + 보관비(storage).
#   → totalFulfillmentFeeDeductionAmount 는 "풀필먼트 합계"가 아니라 **배송비(delivery)뿐**.
#     검산: ful=130,599 + ware=75,489 + stor=168 = J 206,256 (레퍼런스와 ±0). 세 값은 서로
#     독립이라 합산해도 이중계상 아님.
#   비용 인식 basis = **발생비용(f)** — 이월(g)은 totalCarryOverSettlementDeductionAmount /
#     pastDeductedCfsFeeDetails 등 별도 필드로 분리돼 있고 아래 7개 컴포넌트엔 섞이지 않는다.
#     즉 우리가 적재하는 값은 해당 인식기간 gross 발생분이며 f-g(이월조정)·최종지급액이 아님(D-10 충족).
_FEE_FIELD_MAP: dict[str, tuple[str, int]] = {
    "totalTakeRateAmountWithVat": ("sale_fee", 1),           # 판매수수료(B, VAT포함)
    "totalFulfillmentFeeDeductionAmount": ("delivery", 1),    # ★배송비(풀필먼트 J의 배송 컴포넌트)
    "totalStorageFeeDeductionAmount": ("storage", 1),         # 보관비(풀필먼트 J 컴포넌트)
    "totalWarehousingFeeDeductionAmount": ("warehousing", 1), # 입출고비(풀필먼트 J 컴포넌트)
    "totalCreturnReverseShippingFeeDeductionAmount": ("return_shipping", 1),  # 반품배송비
    "totalVreturnHandlingFeeDeductionAmount": ("return_handling", 1),          # 반출처리비
    "totalAdSalesDeductionAmount": ("ad_sales", 1),           # 광고비(d)(D-16: net_profit 전액 차감 포함, RG 광고는 정산 전용)
}

# 정산주기 최대 조회 주 수 (폭주 방지)
_MAX_WEEKS = 52


# ═══════════════════════════════════════════
# 파싱 헬퍼
# ═══════════════════════════════════════════

def _dec(value) -> Decimal:
    """숫자 값 → Decimal. None/변환불가=0. D-13 방어."""
    if value is None:
        return Decimal(0)
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal(0)


def _parse_date(value) -> date | None:
    """UTC ISO "YYYY-MM-DDTHH:mm:ssZ" → KST date. None=파싱실패."""
    if not value:
        return None
    s = str(value).strip()
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt.astimezone(_KST).date()
    except (ValueError, TypeError):
        log.warning("RG settlement date 파싱 실패: %s", value)
        return None


def _extract_fees(detail: dict, period_start: date, period_end: date, account_key: str) -> list[CoupangRgSettlementFee]:
    """settlementStatusReportDetail → CoupangRgSettlementFee 리스트.

    D-9: sale_fee+fulfillment 둘 다. D-10: recognition_date_from/to=매출인식일.
    D-13: 키 없으면 0으로 처리(warn만, 중단 안 함).
    """
    rows = []
    for field, (fee_type, sign) in _FEE_FIELD_MAP.items():
        raw = detail.get(field)
        if raw is None:
            log.debug("RG settlement field 없음(스키마 변동 가능): %s", field)
        amount = _dec(raw) * sign
        rows.append(CoupangRgSettlementFee(
            account_key=account_key,
            recognition_date_from=period_start,
            recognition_date_to=period_end,
            fee_type=fee_type,
            vendor_item_id="",  # 계정 단위(status/api) sentinel — 옵션 row와 grain 구분
            raw_type=field,
            amount=amount,
        ))
    return rows


def _parse_status_response(data: dict, account_key: str) -> list[CoupangRgSettlementFee]:
    """status/api 응답 dict → CoupangRgSettlementFee 리스트.

    D-13: 'settlementStatusReports' 키 없거나 list 아님 → WingReadError(스키마 드리프트).
    정산주기 하나당 여러 row(fee_type별). 같은 기간 중복은 upsert로 처리.
    """
    if "settlementStatusReports" not in data:
        raise WingReadError(
            "status/api 응답에 settlementStatusReports 키 없음(스키마 드리프트 의심)"
        )
    reports = data["settlementStatusReports"]
    if not isinstance(reports, list):
        raise WingReadError(
            f"settlementStatusReports가 list 아님: {type(reports).__name__}"
        )

    all_rows: list[CoupangRgSettlementFee] = []
    seen_keys: set[tuple] = set()

    for report in reports:
        if not isinstance(report, dict):
            log.warning("RG settlement report가 dict 아님, 건너뜀: %s", type(report))
            continue

        period_start = _parse_date(report.get("settlementPeriodStartDate"))
        period_end = _parse_date(report.get("settlementPeriodEndDate"))
        if period_start is None or period_end is None:
            log.warning("RG settlement 기간 파싱 실패, 건너뜀: %s", report.get("settlementGroupKey"))
            continue

        detail = report.get("settlementStatusReportDetail")
        if not isinstance(detail, dict):
            log.warning("settlementStatusReportDetail 없음, 건너뜀: %s / %s", period_start, period_end)
            continue

        rows = _extract_fees(detail, period_start, period_end, account_key)
        for row in rows:
            key = (account_key, period_start, period_end, row.fee_type)
            if key not in seen_keys:
                seen_keys.add(key)
                all_rows.append(row)
            # ★같은 기간에 가지급(70%)+확정(30%) 설치분 리포트가 2개 온다(라이브 확정 2026-06-10,
            #   WING1, 원칙22). settlementRatio는 '지급액' 분할일 뿐, 수수료 필드는 fee_type마다
            #   설치분 노출 방식이 다르다(라이브 실측):
            #     - 판매수수료(take_rate): 기간 총액을 양쪽에 '동일값' 반복(예: 102,950=102,950).
            #       → 합산하면 2배 이중계상(net_profit 과다차감 버그). 동일값은 반복 총액이라 1회만 계상.
            #     - 풀필먼트(배송/입출고/보관): 가지급분에 full·확정분 0 → 합산=full(정상).
            #     - 광고비(ad_sales): 70/30 등으로 '분할' 노출(예: 16,510 vs 7,076) → 합산해야 기간 총액.
            #   원칙: 설치분 간 '동일값'이면 반복 총액 → 1회 계상(dedupe), '다른값'이면 분할/누적 → 합산.
            #   (동일값 dedupe가 필요한 건 현재 take_rate뿐. 나머지는 full+0 또는 분할이라 합산이 옳다.)
            else:
                for existing in all_rows:
                    if (existing.account_key == row.account_key
                            and existing.recognition_date_from == row.recognition_date_from
                            and existing.recognition_date_to == row.recognition_date_to
                            and existing.fee_type == row.fee_type):
                        # dedupe는 라이브로 '반복 총액'이 입증된 sale_fee에만 적용(codex P2 수용).
                        # sale_fee가 설치분에 동일값으로 오면 기간 총액 반복이므로 1회만 계상(2배 버그 차단).
                        # 그 외 fee_type(풀필먼트=full+0, ad_sales=분할)은 항상 합산(기존 동작 유지) —
                        # 증거 없는 전역 dedupe는 우연한 동일분할(예: 50/50)에서 과소계상 위험이 있다.
                        if existing.fee_type == "sale_fee" and row.amount == existing.amount:
                            pass  # 반복 총액 → dedupe(추가 안 함)
                        else:
                            existing.amount += row.amount   # 분할/누적/풀필먼트 → 합산
                        break

    return all_rows


# ═══════════════════════════════════════════
# 쿠키 로드 (저장은 rg_inbound_sync.save_cookie 단일 경로)
# ═══════════════════════════════════════════
# ★쿠키 저장은 rg_inbound_sync.save_cookie 하나로만 한다(같은 CoupangWingCookie 테이블 공유).
#   과거 이 파일에 save_settlement_cookie가 있었으나 xsrf를 raw로 저장 → _load_client의
#   decrypt_secret(xsrf)와 어긋나 로드 시 CookieCryptoError로 크래시(2026-06-10 take_rate 조사 중
#   발견, dead code였음). 두 writer가 같은 테이블에 다르게 쓰던 drift라 함수를 제거했다.
#   여기에 두 번째 쿠키 writer를 다시 추가하지 말 것 — 저장은 save_cookie로 통일(원칙18 중복 금지).

def _load_client(db: Session, account_key: str) -> CoupangWingRgSettlementClient:
    """DB에서 쿠키 로드·복호화 → CoupangWingRgSettlementClient 반환.

    쿠키 없거나 status=red → WingAuthError(호출자가 fail-soft 처리).
    """
    row = db.query(CoupangWingCookie).filter_by(account_key=account_key).first()
    if row is None or not row.cookie_blob:
        raise WingAuthError(f"{account_key} Wing 쿠키 없음 — 쿠키 등록 필요")
    if row.status == "red":
        raise WingAuthError(f"{account_key} Wing 쿠키 만료(status=red) — 재등록 필요")
    cookie = decrypt_secret(row.cookie_blob)
    # xsrf_token은 Fernet 암호화 저장(inbound_sync.py:176 동일 패턴). 복호화 필수.
    xsrf = decrypt_secret(row.xsrf_token) if row.xsrf_token else ""
    return CoupangWingRgSettlementClient(cookie_header=cookie, xsrf_token=xsrf)


def _mark_red(db: Session, account_key: str) -> None:
    row = db.query(CoupangWingCookie).filter_by(account_key=account_key).first()
    if row:
        row.status = "red"
        db.commit()


def _mark_last_success(db: Session, account_key: str) -> None:
    row = db.query(CoupangWingCookie).filter_by(account_key=account_key).first()
    if row:
        row.last_success_at = kst_now()
        db.commit()


# ═══════════════════════════════════════════
# 메인 동기화 (라우터/스케줄러가 호출)
# ═══════════════════════════════════════════

def sync_rg_settlement(
    db: Session,
    account_key: str,
    *,
    start_date: str | None = None,
    end_date: str | None = None,
) -> dict:
    """RG 정산 수수료 수집·파싱·upsert. D-10: 매출인식일 기준.

    start_date/end_date: KST "YYYY-MM-DD". None=최근 90일.
    반환: {"synced": N, "account_key": ..., "period": ..., "status": "ok"|"auth_error"|"read_error"}
    fail-soft: WingAuthError→status=red+반환, WingReadError→반환(이력 유지).
    """
    today = kst_now().date()
    if end_date is None:
        end_date = today.strftime("%Y-%m-%d")
    if start_date is None:
        start_date = (today - timedelta(days=90)).strftime("%Y-%m-%d")

    log.info("RG 정산 sync 시작: %s %s~%s", account_key, start_date, end_date)

    try:
        client = _load_client(db, account_key)
    except WingAuthError as e:
        log.warning("RG 정산 쿠키 로드 실패: %s — %s", account_key, e)
        return {"synced": 0, "account_key": account_key, "period": f"{start_date}~{end_date}", "status": "auth_error", "error": str(e)}

    try:
        raw = client.get_settlement_status(start_date=start_date, end_date=end_date)
    except WingAuthError as e:
        log.warning("RG 정산 API 인증 실패: %s — %s", account_key, e)
        _mark_red(db, account_key)
        return {"synced": 0, "account_key": account_key, "period": f"{start_date}~{end_date}", "status": "auth_error", "error": str(e)}
    except WingReadError as e:
        log.error("RG 정산 API 읽기 실패: %s — %s", account_key, e)
        return {"synced": 0, "account_key": account_key, "period": f"{start_date}~{end_date}", "status": "read_error", "error": str(e)}

    try:
        rows = _parse_status_response(raw, account_key)
    except WingReadError as e:
        log.error("RG 정산 파싱 실패: %s — %s", account_key, e)
        return {"synced": 0, "account_key": account_key, "period": f"{start_date}~{end_date}", "status": "parse_error", "error": str(e)}

    # upsert: (account_key, from, to, fee_type, vendor_item_id="") 기준.
    # ★vendor_item_id="" 가드(S6): 같은 (account,from,to,fee_type)에 옵션 row(vendor_item_id=옵션ID)가
    #   공존하므로, 계정 row 쿼리에 ""를 명시하지 않으면 옵션 row를 first()로 잡아 덮어쓸 위험.
    synced = 0
    for row in rows:
        existing = db.query(CoupangRgSettlementFee).filter_by(
            account_key=row.account_key,
            recognition_date_from=row.recognition_date_from,
            recognition_date_to=row.recognition_date_to,
            fee_type=row.fee_type,
            vendor_item_id="",
        ).first()
        if existing is None:
            db.add(row)
        else:
            existing.amount = row.amount
            existing.raw_type = row.raw_type
            existing.synced_at = kst_now()
        synced += 1

    db.commit()
    _mark_last_success(db, account_key)
    log.info("RG 정산 sync 완료: %s — %d건", account_key, synced)
    return {"synced": synced, "account_key": account_key, "period": f"{start_date}~{end_date}", "status": "ok"}


def sync_all_rg_settlements(db: Session, *, start_date: str | None = None, end_date: str | None = None) -> list[dict]:
    """모든 RG 계정 정산 수수료 sync. 스케줄러에서 호출."""
    return [sync_rg_settlement(db, ak, start_date=start_date, end_date=end_date) for ak in RG_ACCOUNTS]


# ═══════════════════════════════════════════════════════════════════════
# S6 — 옵션 단위 수집 (종류별 엑셀 파서, D-2/D-8)
# ═══════════════════════════════════════════════════════════════════════
# 종류별 엑셀(WAREHOUSING_SHIPPING 등)은 시트=비용종류별. 각 시트에 주문/SKU 단위 상세(옵션ID 포함).
# ★S6 회계규칙(§8-1, 원칙22 실증): 옵션 cost = **할인적용가(A-B), VAT前** (발생비용 A=gross 아님 —
#   status/api 컴포넌트와 불일치 회피: 입출고 100,650(A) ≠ 68,625(A-B)=요약합계=status/api VAT前).
# 검산: Σ상세(A-B) == 요약합계(VAT前). +요약세액 == 요약최종 == status/api 저장값(VAT후).
# D-13 방어: 미지 시트명·헤더 부재·컬럼 부재 → 해당 시트 fail-soft skip(warn), 중단 안 함.

# 시트명(엑셀 탭) → fee_type. fee_type 체계는 status/api(_FEE_FIELD_MAP)와 정합.
# WAREHOUSING_SHIPPING 엑셀 = 입출고비+배송비 2시트. 나머지 종류는 후속(같은 패턴, 매핑만 확장).
_SHEET_FEE_TYPE_MAP: dict[str, str] = {
    "입출고비": "warehousing",
    "배송비": "delivery",
    "보관비": "storage",
    "판매수수료": "sale_fee",
    "반품 회수비": "return_shipping",
    "반품회수비": "return_shipping",
    "반품 재입고비": "return_shipping",
    "반품재입고비": "return_shipping",
    "반출비": "return_handling",
}

# 상세 헤더(2층) 컬럼 라벨 (헤더명 기반 동적 매핑 — 시트별 컬럼 위치 다름 대응)
_COL_OPTION_ID = "옵션ID"
_COL_RECOGNITION = "매출인식일"
_COL_PERIOD_END = "정산주기(종료일)"
_COL_DISCOUNTED = "할인적용가(A-B)"  # ★옵션 cost (VAT前). 엑셀 ASCII 하이픈 '-'.


def _norm_option_id(value) -> str:
    """옵션ID 정규화: 95521944483.0 → '95521944483'. 빈값/'-'/None → ''(=옵션 아님, skip 대상)."""
    if value is None:
        return ""
    s = str(value).strip()
    if s in ("", "-"):
        return ""
    if s.endswith(".0"):  # float로 읽힌 경우
        s = s[:-2]
    return s


def _parse_excel_date(value) -> date | None:
    """엑셀 날짜 셀(문자열 'YYYY-MM-DD' 또는 datetime/date) → date. 실패=None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        log.warning("RG 엑셀 날짜 파싱 실패: %s", value)
        return None


def _find_header_row(ws) -> int | None:
    """'옵션ID'를 값으로 가진 행 = 상세 헤더행. 상단 20행 내 탐색. 없으면 None(D-13)."""
    for r in range(1, min(ws.max_row, 20) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() == _COL_OPTION_ID:
                return r
    return None


def _build_col_map(ws, header_row: int) -> dict[str, int]:
    """헤더행 + 바로 아래 서브헤더행을 합쳐 {label: col} 매핑(서브 우선).

    2층 헤더 대응: row7 메인(옵션ID·매출인식일·'…입출고비(VAT별도)' 병합) +
    row8 서브('발생비용(A)'·'할인가(B)'·'할인적용가(A-B)'). 병합셀 우측은 None이라 서브가 채움.
    같은 label 중복 시 먼저 본 컬럼 유지.
    """
    col_map: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        main = ws.cell(row=header_row, column=c).value
        sub = ws.cell(row=header_row + 1, column=c).value
        for label in (sub, main):  # 서브헤더 우선
            if label is not None and str(label).strip():
                key = str(label).strip()
                col_map.setdefault(key, c)
    return col_map


def _parse_summary(ws) -> dict | None:
    """요약 행(정산주기 종료일·합계·세액·최종비용) 파싱. 헤더 위치 동적 탐색.

    상단 8행 내에서 '정산주기(종료일)' + '…합계' + '세액' + '최종비용' 헤더를 모두 가진 행을
    요약 헤더로 보고, 그 다음 행을 데이터로 읽는다. (상세 헤더 row7은 '세액'/'최종비용' 없어 매칭 안 됨.)
    """
    for r in range(1, min(ws.max_row, 8) + 1):
        labels: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip():
                labels.setdefault(str(v).strip(), c)
        end_col = labels.get(_COL_PERIOD_END)
        sum_col = next((c for k, c in labels.items() if k.endswith("합계")), None)
        tax_col = labels.get("세액")
        final_col = labels.get("최종비용")
        if end_col and sum_col and tax_col and final_col:
            dr = r + 1
            return {
                "period_end": _parse_excel_date(ws.cell(row=dr, column=end_col).value),
                "sum_discounted": _dec(ws.cell(row=dr, column=sum_col).value),
                "tax": _dec(ws.cell(row=dr, column=tax_col).value),
                "final": _dec(ws.cell(row=dr, column=final_col).value),
            }
    return None


def parse_settlement_xlsx(content: bytes) -> dict:
    """종류별 엑셀 → 옵션 단위 집계 + 검산. (D-8: 파싱은 Harness 책임, 별도 SA 없음.)

    반환:
      {"sheets": [
          {"fee_type", "sheet_name",
           "options": {(vendor_item_id, period_end): Σ(A-B) VAT前},
           "summary": {period_end, sum_discounted, tax, final} | None,
           "sum_detail": Σ상세(A-B),
           "reconcile": {"sum_detail","sum_summary","diff","match"} | None},
        ...],
       "sheets_skipped": [시트명...]}
    옵션 cost = 할인적용가(A-B) VAT前(§8-1). 집계 grain=(옵션ID, 정산주기 종료일).
    D-13: 미지 시트명·헤더/컬럼 부재 → 해당 시트 skip(중단 안 함). 파일 자체 손상만 WingReadError.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        raise WingReadError(f"RG 정산 엑셀 열기 실패: {e}") from e

    sheets_out: list[dict] = []
    skipped: list[str] = []
    try:
        for sn in wb.sheetnames:
            fee_type = _SHEET_FEE_TYPE_MAP.get(sn.strip())
            if fee_type is None:
                log.warning("RG 엑셀 미지 시트명, 건너뜀: %s", sn)
                skipped.append(sn)
                continue
            ws = wb[sn]
            hr = _find_header_row(ws)
            if hr is None:
                log.warning("RG 엑셀 '%s' 상세 헤더(옵션ID) 못 찾음, 건너뜀", sn)
                skipped.append(sn)
                continue
            col_map = _build_col_map(ws, hr)
            oid_col = col_map.get(_COL_OPTION_ID)
            cost_col = col_map.get(_COL_DISCOUNTED)
            pend_col = col_map.get(_COL_PERIOD_END)
            if not oid_col or not cost_col:
                log.warning("RG 엑셀 '%s' 옵션ID/할인적용가 컬럼 없음, 건너뜀", sn)
                skipped.append(sn)
                continue

            options: dict[tuple[str, date | None], Decimal] = {}
            sum_detail = Decimal(0)
            # 헤더+1부터 순회. 서브헤더 행은 옵션ID가 비어 자동 skip(서브헤더 유무 무관).
            for r in range(hr + 1, ws.max_row + 1):
                oid = _norm_option_id(ws.cell(row=r, column=oid_col).value)
                if not oid:
                    continue
                cost = _dec(ws.cell(row=r, column=cost_col).value)
                pend = _parse_excel_date(ws.cell(row=r, column=pend_col).value) if pend_col else None
                key = (oid, pend)
                options[key] = options.get(key, Decimal(0)) + cost
                sum_detail += cost

            summary = _parse_summary(ws)
            reconcile = None
            if summary is not None:
                diff = sum_detail - summary["sum_discounted"]
                match = abs(diff) < Decimal("0.5")
                reconcile = {
                    "sum_detail": sum_detail,
                    "sum_summary": summary["sum_discounted"],
                    "diff": diff,
                    "match": match,
                }
                if not match:
                    log.warning(
                        "RG 엑셀 '%s' 검산 불일치(원칙22): Σ상세 %s != 요약합계 %s (diff %s)",
                        sn, sum_detail, summary["sum_discounted"], diff,
                    )
            sheets_out.append({
                "fee_type": fee_type,
                "sheet_name": sn,
                "options": options,
                "summary": summary,
                "sum_detail": sum_detail,
                "reconcile": reconcile,
            })
    finally:
        wb.close()  # read_only 모드는 명시적 close 권장

    return {"sheets": sheets_out, "sheets_skipped": skipped}


def _expected_period_start(period_end: date) -> date:
    """정산주기 종료일 → 시작일 (달력 규칙, 순수 함수).

    규칙: **월요일~일요일 주별 + 달력 월 경계에서 분할**.
      start = max(그 주의 월요일, 그 달 1일)
            = max(period_end - period_end.weekday()일, period_end.replace(day=1))
    월 경계에 걸친 주는 두 조각으로 나뉜다(예: 06-29~06-30 | 07-01~07-05).

    규칙 출처: 2026-07-17 prod 계정 row 18개 전수 검증 18/18 일치(라이브 실측).
    레퍼런스 17에는 "주별"만 있고 월 경계 분할은 미기재 — 라이브가 권위(원칙22).
    구 폴백(period_end-6일 고정)은 월 경계 주간에서 반드시 틀려 6월 RG 비용을 +55%
    과대계상한 사고의 원인이었다(2026-07-17). 이 함수는 추측이 아니라 검증된 규칙이다.
    """
    monday = period_end - timedelta(days=period_end.weekday())
    month_first = period_end.replace(day=1)
    return max(monday, month_first)


def _resolve_period_start(db: Session, account_key: str, period_end: date) -> date:
    """옵션 row의 recognition_date_from 결정.

    1순위(불변): status/api 계정 row(vendor_item_id="", 같은 account_key·recognition_date_to==
      period_end)에서 from을 차용 → status/api와 정산주기(grain) 정확 일치 보장(검산 가능).
    2순위: 계정 row 없으면 검증된 달력 규칙(_expected_period_start) 사용 + 경고.
      (구 -6d 추측 폴백은 월 경계에서 틀려 삭제. 달력 규칙은 18/18 라이브 검증됨.)
    """
    acct = db.query(CoupangRgSettlementFee).filter_by(
        account_key=account_key, recognition_date_to=period_end, vendor_item_id="",
    ).first()
    if acct is not None:
        return acct.recognition_date_from
    fallback = _expected_period_start(period_end)
    log.warning(
        "RG 옵션 적재: status/api 계정 row 없음(account=%s, period_end=%s) → "
        "달력 규칙 폴백(월~일+월경계분할)=%s (추측 아님, 18/18 검증)",
        account_key, period_end, fallback,
    )
    return fallback


def ingest_settlement_xlsx(db: Session, account_key: str, content: bytes) -> dict:
    """종류별 엑셀 bytes → 파싱·검산·옵션 단위 upsert(S6).

    upsert grain=(account_key, from, to, fee_type, vendor_item_id=옵션ID). VAT前(A-B) 저장.
    반환: {"account_key", "upserted", "sheets": [{fee_type, sheet_name, options(개수),
           reconcile, vs_status_api}], "sheets_skipped", "status": "ok"|"empty"}.
    검산2(vs_status_api): 요약최종(VAT후) == status/api 계정 row amount(같은 fee_type·기간).
    """
    parsed = parse_settlement_xlsx(content)
    sheets = parsed["sheets"]
    if not sheets:
        return {
            "account_key": account_key, "upserted": 0, "sheets": [],
            "sheets_skipped": parsed["sheets_skipped"], "status": "empty",
        }

    # ── 1. fee_type 단위 병합 (codex 2R-P1) + 종료일 fallback (codex 3R-P1) ──
    # _SHEET_FEE_TYPE_MAP은 여러 시트명을 같은 fee_type으로 매핑(예 '반품 회수비'+'반품 재입고비'
    # → return_shipping). 시트별 즉시 delete는 뒤 시트가 앞 시트 삽입분을 지워 마지막 시트만 남는다.
    # → fee_type 단위로 옵션 cost를 합산한 뒤 delete-once + insert.
    # ★상세 종료일 파싱 실패 시 요약 period_end로 fallback(codex 3R-P1): summary period를 무조건
    #   삭제 대상에 넣으면 상세 종료일 불명 시 delete만 하고 insert는 skip돼 데이터 손실이 난다.
    #   → 상세 row의 종료일이 없으면 요약 종료일로 채워 insert까지 되게 하고, 요약 period는
    #   "상세 0건 정상 시트"일 때만 삭제 대상에 추가한다.
    merged: dict[str, dict[tuple[str, date | None], Decimal]] = {}
    merged_periods: dict[str, set[date]] = {}   # fee_type -> snapshot 삭제 대상 period_end 집합
    for sheet in sheets:
        ft = sheet["fee_type"]
        summ = sheet["summary"]
        fallback_pe = summ["period_end"] if summ is not None else None
        bucket = merged.setdefault(ft, {})
        pset = merged_periods.setdefault(ft, set())
        has_options = False
        for (oid, pe), cost in sheet["options"].items():
            has_options = True
            eff_pe = pe if pe is not None else fallback_pe  # 상세 종료일 없으면 요약으로 fallback
            bucket[(oid, eff_pe)] = bucket.get((oid, eff_pe), Decimal(0)) + cost
            if eff_pe is not None:
                pset.add(eff_pe)  # insert될 period만 삭제 대상(삭제만 하고 못 넣는 손실 방지)
        # 상세 0건 정상 시트(요약만 존재)도 snapshot replace 대상 — 옵션 전삭제로 stale 제거.
        if not has_options and fallback_pe is not None:
            pset.add(fallback_pe)

    # ── 2. fee_type별 snapshot replace (delete-once, period_end 단위) ─────
    for ft, pset in merged_periods.items():
        for pe in pset:
            db.query(CoupangRgSettlementFee).filter(
                CoupangRgSettlementFee.account_key == account_key,
                CoupangRgSettlementFee.fee_type == ft,
                CoupangRgSettlementFee.recognition_date_to == pe,
                CoupangRgSettlementFee.vendor_item_id != "",
            ).delete(synchronize_session=False)

    # ── 3. 병합된 옵션 적재 (delete 후라 전부 신규 insert) ─────────────────
    upserted_total = 0
    for ft, options in merged.items():
        for (oid, period_end), cost in options.items():
            if period_end is None:
                log.warning("RG 옵션 적재 skip(정산주기 종료일 불명·요약 fallback도 없음): %s/%s", ft, oid)
                continue
            period_start = _resolve_period_start(db, account_key, period_end)
            db.add(CoupangRgSettlementFee(
                account_key=account_key,
                recognition_date_from=period_start,
                recognition_date_to=period_end,
                fee_type=ft,
                vendor_item_id=oid,
                raw_type=f"xlsx:{ft}",
                amount=cost,
            ))
            upserted_total += 1

    # ── 4. 검산 리포트 ───────────────────────────────────────────────────
    # vs_status_api(codex 3R-P2): 같은 fee_type 여러 시트는 적재가 합산되므로 검산도 fee_type+
    #   period_end 단위 요약최종 합계로 status/api 계정 row와 비교(시트별 비교 시 false mismatch).
    final_by_ftpe: dict[tuple[str, date], Decimal] = {}
    for sheet in sheets:
        summary = sheet["summary"]
        if summary is not None and summary["period_end"] is not None:
            k = (sheet["fee_type"], summary["period_end"])
            final_by_ftpe[k] = final_by_ftpe.get(k, Decimal(0)) + summary["final"]
    vs_by_ftpe: dict[tuple[str, date], dict] = {}
    for (ft, pe), final_sum in final_by_ftpe.items():
        acct = db.query(CoupangRgSettlementFee).filter_by(
            account_key=account_key, recognition_date_to=pe, fee_type=ft, vendor_item_id="",
        ).first()
        if acct is not None:
            api_amt = Decimal(str(acct.amount))
            diff = final_sum - api_amt
            vs_by_ftpe[(ft, pe)] = {
                "summary_final": final_sum,
                "status_api_amount": api_amt,
                "diff": diff,
                "match": abs(diff) < Decimal("0.5"),
            }
    # 시트별 리포트(엑셀 내부 검산 reconcile은 시트 단위, vs_status_api는 fee_type+period 합계 공유)
    sheets_report: list[dict] = []
    for sheet in sheets:
        summary = sheet["summary"]
        pe = summary["period_end"] if summary is not None else None
        sheets_report.append({
            "fee_type": sheet["fee_type"],
            "sheet_name": sheet["sheet_name"],
            "options": len(sheet["options"]),
            "reconcile": sheet["reconcile"],
            "vs_status_api": vs_by_ftpe.get((sheet["fee_type"], pe)),
        })

    db.commit()
    log.info("RG 옵션 단위 적재: %s — %d건(%d시트, %d fee_type)",
             account_key, upserted_total, len(sheets), len(merged))
    return {
        "account_key": account_key,
        "upserted": upserted_total,
        "sheets": sheets_report,
        "sheets_skipped": parsed["sheets_skipped"],
        "status": "ok",
    }


# ═══════════════════════════════════════════════════════════════════════
# S6-auto: 엑셀 자동 다운로드 + 적재
# ═══════════════════════════════════════════════════════════════════════

_POLL_INTERVAL_SEC = 10   # 폴링 간격
_POLL_TIMEOUT_SEC = 300   # 최대 대기 5분


def auto_download_and_ingest(
    db: Session,
    account_key: str,
    vendor_id: str,
    *,
    poll_timeout: int = _POLL_TIMEOUT_SEC,
) -> dict:
    """S6-auto: DB의 정산 기간별 × 종류별 엑셀 자동 다운로드 → 옵션 단위 적재.

    1. DB에서 (account_key, vendor_item_id="") 계정 row의 distinct 정산 기간 조회
    2. 각 기간 × CONFIRMED_SELLER_REPORT_TYPES 조합으로 request-download/api 호출
    3. download-list/api 폴링 → COMPLETED 확인
    4. download/api/v2 → S3 URL → 엑셀 다운로드 → ingest_settlement_xlsx
    반환: {account_key, requested, completed, ingested, errors, status}
    """
    # P2-a(codex): auth 실패는 per-account error로 반환(전체 abort 방지).
    try:
        client = _load_client(db, account_key)
    except WingAuthError as e:
        return {"account_key": account_key, "requested": 0, "completed": 0,
                "ingested": 0, "errors": [str(e)], "status": "auth_error"}

    # ── 1. DB에서 정산 기간 조회 ─────────────────────────────────────────
    periods_q = (
        db.query(
            CoupangRgSettlementFee.recognition_date_from,
            CoupangRgSettlementFee.recognition_date_to,
        )
        .filter(
            CoupangRgSettlementFee.account_key == account_key,
            CoupangRgSettlementFee.vendor_item_id == "",
        )
        .distinct()
        .all()
    )
    if not periods_q:
        log.info("S6-auto: 정산 기간 없음(status/api 먼저 실행 필요) account_key=%s", account_key)
        return {"account_key": account_key, "requested": 0, "completed": 0,
                "ingested": 0, "errors": [], "status": "no_periods"}

    # ── 2. 엑셀 생성 요청 ────────────────────────────────────────────────
    # P1(codex): 각 요청에 고유한 request_time_ms 사용(같은 값 → 서버가 duplicate 처리).
    # P2-b(codex): duplicate 시 기존 요청이 24h 이전에 만들어졌을 수 있으므로
    #              poll_from_ms를 24h 전으로 넓혀 search window 안에 포함시킴.
    base_ms = int(time.time() * 1000)
    poll_from_ms = base_ms - 24 * 3600 * 1000  # 24시간 전부터 검색

    requested: list[dict] = []
    errors: list[str] = []

    req_idx = 0
    for period_from, period_to in periods_q:
        group_key = f"{vendor_id}-{period_from}-{period_to}"
        for report_type in CONFIRMED_SELLER_REPORT_TYPES:
            # 요청마다 1ms씩 증가 → 모두 고유한 requestTime
            req_time_ms = base_ms + req_idx
            req_idx += 1
            try:
                resp = client.request_download(
                    seller_report_type=report_type,
                    settlement_group_keys=[group_key],
                    request_time_ms=req_time_ms,
                )
                request_id = resp.get("requestId", "")
                if resp.get("duplicateRequest"):
                    log.info("S6-auto: 중복 요청(기존 항목 폴링) %s / %s", report_type, group_key)
                requested.append({
                    "request_id": request_id,
                    "report_type": report_type,
                    "period_from": period_from,
                    "period_to": period_to,
                    "group_key": group_key,
                })
                log.info("S6-auto: 요청 완료 %s / %s → %s", report_type, group_key, request_id)
            except WingAuthError as e:
                # P2-c(codex): 세션 만료 → cookie red 표시(sync_rg_settlement 패턴 동일).
                _mark_red(db, account_key)
                msg = f"세션 만료(cookie red 처리) {report_type}/{group_key}: {e}"
                log.warning("S6-auto: %s", msg)
                errors.append(msg)
                break  # 쿠키 만료 시 남은 요청도 실패 → 루프 탈출
            except WingReadError as e:
                msg = f"request_download 실패 {report_type}/{group_key}: {e}"
                log.warning("S6-auto: %s", msg)
                errors.append(msg)

    if not requested:
        return {"account_key": account_key, "requested": 0, "completed": 0,
                "ingested": 0, "errors": errors, "status": "all_failed"}

    # ── 3. 폴링 → COMPLETED 확인 ─────────────────────────────────────────
    pending_ids = {r["request_id"] for r in requested if r["request_id"]}
    completed_items: dict[str, dict] = {}  # request_id → download-list 항목
    deadline = time.time() + poll_timeout
    # poll_to: 가장 마지막 req_time + timeout + 여유(1min)
    poll_to_ms = base_ms + req_idx + poll_timeout * 1000 + 60_000

    while time.time() < deadline and len(completed_items) < len(pending_ids):
        try:
            items = client.get_download_list(
                request_time_from_ms=poll_from_ms,
                request_time_to_ms=poll_to_ms,
            )
        except WingAuthError as e:
            _mark_red(db, account_key)
            errors.append(f"폴링 중 세션 만료(cookie red): {e}")
            break
        except WingReadError as e:
            log.warning("S6-auto: download-list 폴링 실패: %s", e)
            break

        for item in items:
            rid = item.get("requestId", "")
            if rid in pending_ids and item.get("downloadStatus") == "COMPLETED":
                completed_items[rid] = item

        remaining = pending_ids - set(completed_items)
        if remaining:
            log.info("S6-auto: 폴링 대기 중 %d/%d ...", len(completed_items), len(pending_ids))
            time.sleep(_POLL_INTERVAL_SEC)

    timed_out = [r["request_id"] for r in requested if r["request_id"] not in completed_items]
    if timed_out:
        for rid in timed_out:
            errors.append(f"download timeout requestId={rid}")
        log.warning("S6-auto: %d개 timeout", len(timed_out))

    # ── 4. 다운로드 + 적재 ───────────────────────────────────────────────
    ingested = 0
    for req in requested:
        rid = req["request_id"]
        item = completed_items.get(rid)
        if not item:
            continue
        try:
            url = client.get_download_url(request_time_ms=item["requestTime"])
            xlsx_bytes = client.download_excel_bytes(url)
            ingest_result = ingest_settlement_xlsx(db, account_key, xlsx_bytes)
            # P2-d(codex): 0행 적재(모든 시트 skip/empty) → error로 기록.
            upserted = ingest_result.get("upserted", 0)
            if upserted == 0:
                msg = f"0행 적재(시트 파싱 실패 또는 empty) {req['report_type']}/{req['group_key']}: {ingest_result.get('sheets_skipped')}"
                log.warning("S6-auto: %s", msg)
                errors.append(msg)
            else:
                ingested += 1
                log.info("S6-auto: 적재 완료 %s / %s (%d행)", req["report_type"], req["group_key"], upserted)
        except WingAuthError as e:
            _mark_red(db, account_key)
            errors.append(f"다운로드 중 세션 만료(cookie red): {e}")
        except (WingReadError, Exception) as e:
            msg = f"다운로드/적재 실패 {req['report_type']}/{req['group_key']}: {e}"
            log.warning("S6-auto: %s", msg)
            errors.append(msg)

    status = "ok" if not errors else ("partial" if ingested > 0 else "failed")
    log.info("S6-auto 완료: %s — 요청 %d / 완료 %d / 적재 %d / 오류 %d",
             account_key, len(requested), len(completed_items), ingested, len(errors))
    return {
        "account_key": account_key,
        "requested": len(requested),
        "completed": len(completed_items),
        "ingested": ingested,
        "errors": errors,
        "status": status,
    }


def auto_download_all(db: Session, vendor_id_map: dict[str, str]) -> list[dict]:
    """모든 RG 계정 S6-auto 실행. vendor_id_map: {account_key: vendor_id}."""
    return [
        auto_download_and_ingest(db, ak, vendor_id_map[ak])
        for ak in RG_ACCOUNTS
        if ak in vendor_id_map
    ]


# ════════════════════════════════════════════════════════════════════
# S4-P2: RG 정산 자동 다운로드 갱신 트리거·heartbeat (트랙 D-8)
# ════════════════════════════════════════════════════════════════════
# Wing 페처(별도 데몬 com.ohisell.wing, D-5)가 RG 정산 엑셀을 브라우저측 다운로드 → prod push하는
# 트리거. vendor_summary_sync 패턴을 그대로 미러링(상태행=CoupangWingCookie KV)하되 account_key를
# 분리해 판매분석(COUPANG_WING_VS)·광고와 섞지 않는다(D-5 세션/상태 분리 원칙).
#   - request_refresh: UI 버튼/스케줄 → 요청 플래그 set
#   - refresh_status:  UI 폴링·페처 데몬 공용(요청 여부 + 마지막 push 시각)
#   - claim_refresh:   페처가 요청 소비(원자적 조건부 UPDATE)
#   - mark_heartbeat:  업로드(push) 성공 시 last_success_at 갱신(라우터 경계에서 호출 — 머니코드 불변)
_RG_STATE_ACCOUNT = "COUPANG_WING_RG"
# RG 정산은 주 단위 → 30일 무push까지 stale 아님(주간 캐던스 + Mac 야간 off 오탐 방지).
_RG_STALE_HOURS = 30 * 24


def _rg_state_row(db: Session) -> CoupangWingCookie | None:
    return (
        db.query(CoupangWingCookie)
        .filter(CoupangWingCookie.account_key == _RG_STATE_ACCOUNT)
        .first()
    )


def _rg_ensure_state_row(db: Session) -> CoupangWingCookie:
    row = _rg_state_row(db)
    if row is None:
        row = CoupangWingCookie(account_key=_RG_STATE_ACCOUNT)
        db.add(row)
    return row


def rg_mark_heartbeat(db: Session) -> None:
    """RG 엑셀 push(업로드) 성공 시각 갱신(staleness·스케줄 중복방지 기준). 라우터가 ingest 성공 후 호출."""
    row = _rg_ensure_state_row(db)
    row.status = "green"
    row.last_error = None
    row.last_error_at = None  # 성공 = 실패 흔적 클리어(안 지우면 오래된 실패가 화면에 남는다)
    row.last_success_at = kst_now()
    db.commit()


def rg_mark_fetch_error(db: Session, error: str) -> None:
    """Wing 페처 RG run 실패 보고 → last_error/last_error_at 기록(UI가 실패를 감지하는 유일 경로).

    ★존재 이유(PR #30이 광고비에서 먼저 고친 것과 같은 구멍): 페처가 갱신 요청을 claim한
    뒤(=플래그 이미 clear) 브라우저 에러로 죽으면 prod에 아무 흔적도 안 남았다. 성공에만
    시각(last_success_at)이 있고 실패엔 짝이 없어서 UI는 "실패"와 "아직 진행 중"을 구분할
    수단이 없었다 — 215초를 헛기다린 뒤 "Mac 응답 없음"이라는 뭉뚱그린 문구만 냈다.

    ★status는 일부러 건드리지 않는다(PR #30 codex 1R[P2]): status=red는 Layout 배너에서
    곧바로 "쿠키 만료(재설정 필요)" + 쿠키 재설정 CTA로 렌더된다(Layout.tsx:201/206).
    브라우저 크래시는 쿠키 문제가 아니라 재설정해도 헛수고다. 지속 실패는 워치독이
    last_success_at 경과로 잡는다(status 미의존).
    """
    row = _rg_ensure_state_row(db)
    row.last_error = error[:300]  # 컬럼 한계 — 긴 스택트레이스로 보고 자체가 날아가면 안 된다
    row.last_error_at = kst_now()
    db.commit()


def rg_request_refresh(db: Session) -> dict:
    """UI 'RG 정산 갱신' 버튼/스케줄 → 갱신 요청 플래그 set. Wing 페처 데몬이 다음 폴링에서 소비."""
    row = _rg_ensure_state_row(db)
    row.refresh_requested_at = kst_now()
    db.commit()
    return {"requested": True, "requested_at": row.refresh_requested_at.isoformat()}


def rg_refresh_status(db: Session) -> dict:
    """RG 갱신 요청/완료 상태. UI(버튼 후 폴링)·Wing 페처(요청 확인) 공용. 민감값 없음.

    last_error_at=마지막 실패 시각(버튼 후 이 값이 올라가면 갱신 실패) — UI가 성공/실패 둘 중
    무엇이 왔는지 이 두 시각의 변화로 가른다. 없으면 실패를 못 보고 폴링 창을 헛기다린다.
    """
    row = _rg_state_row(db)
    if row is None:
        return {"requested": False, "requested_at": None, "last_success_at": None,
                "status": "none", "last_error": None, "last_error_at": None,
                "age_hours": None, "stale": False}
    age_hours = None
    stale = False
    if row.last_success_at:
        age_hours = max(0.0, round((kst_now() - row.last_success_at).total_seconds() / 3600, 1))
        stale = age_hours > _RG_STALE_HOURS
    return {
        "requested": row.refresh_requested_at is not None,
        "requested_at": row.refresh_requested_at.isoformat() if row.refresh_requested_at else None,
        "last_success_at": row.last_success_at.isoformat() if row.last_success_at else None,
        "status": row.status,
        "last_error": row.last_error,
        "last_error_at": row.last_error_at.isoformat() if row.last_error_at else None,
        "age_hours": age_hours,
        "stale": stale,
    }


def rg_claim_refresh(db: Session) -> dict:
    """Wing 페처가 RG 갱신 요청을 '소비'(플래그 clear). 원자적 조건부 UPDATE(광고/vendor-summary 패턴)."""
    from sqlalchemy import update

    res = db.execute(
        update(CoupangWingCookie)
        .where(CoupangWingCookie.account_key == _RG_STATE_ACCOUNT)
        .where(CoupangWingCookie.refresh_requested_at.isnot(None))
        .values(refresh_requested_at=None)
    )
    db.commit()
    return {"claimed": (res.rowcount or 0) > 0}
