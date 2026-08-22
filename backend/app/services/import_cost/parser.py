"""수입 통관 서류 3종(CI·PL·통관경비서) 파서 — 순수 SA (D-CPP-48).

이 모듈은 **DB도 IO도 FastAPI도 모른다.** 입력은 bytes/str, 출력은 값 객체다.
정본은 사람이 확인한 폼 데이터고 파싱은 채워주기 편의일 뿐이다.

배선: `POST /api/import-cost/parse`(`routers/import_cost.py`)가 이 모듈을 부른다.
파싱 결과는 **저장되지 않는다** — 화면의 폼을 채워줄 뿐이고, 사람이 확인해 저장한다.

## .xls와 PDF는 «선택 의존성»으로 읽는다
- `.xlsx` → `openpyxl` (이미 requirements에 있다)
- `.xls`(BIFF8) → **`xlrd`**  · `.pdf` → **`pypdf`** — 둘 다 `requirements.txt`에 넣었지만
  **prod에 아직 설치돼 있지 않을 수 있다**(`scripts/safe_deploy.sh`에 pip 단계가 없다).
  그래서 함수 안에서 임포트하고, 없으면 «무엇을 하면 되는지»를 말하는 예외를 던진다.
  **앱이 부팅에 실패하거나 다른 기능이 죽지 않는다** — 이 모듈이 선택 경로인 이유다.

★BIFF8을 직접 구현하지 않은 이유: OLE2(CFB) 컨테이너 층 + BIFF 레코드(SST·RK·MULRK·NUMBER·
FORMULA) 복원 층을 최소로 짜도 250~350줄이다. 검증된 오픈소스(`xlrd` 2.0.1, BSD, .xls 전용)가
있으면 직접 구현보다 먼저 쓴다는 것이 우리 금지선이다.

★**OCR은 하지 않는다.** 스캔 PDF는 「글자가 없다」고 말하고 수기 경로로 보낸다 —
읽을 수 없는 것에서 숫자를 지어내는 것이 이 도메인에서 가장 나쁜 실패다.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field, replace
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from .allocator import CostLine, InvoiceLine

_ZERO = Decimal("0")
_EXCEL_EPOCH = datetime(1899, 12, 30)  # 엑셀 1900 날짜계 윤년버그를 흡수하는 표준 기준일


class CustomsDocParseError(ValueError):
    """서류 파싱 실패. 메시지에 **어느 시트·행에서 무엇을 못 읽었는지**를 담는다.

    이 예외를 잡는 쪽(라우터)은 폼을 빈 값으로라도 열어야 한다 — 파서 실패가
    시스템 실패가 되면 안 된다는 게 계약의 금지선이다.
    """


# ──────────────────────────────────────────────
# 값 객체
# ──────────────────────────────────────────────
@dataclass(frozen=True)
class InvoiceParseResult:
    lines: list[InvoiceLine]
    order_nos: list[str | None]  # 라인별 forward-fill 결과. len == len(lines)
    invoice_no: str | None
    invoice_date: date | None
    declared_total: Decimal | None  # G30 — 검산용, 값을 고치지 않는다
    declared_qty: Decimal | None  # E30
    line_total_mismatches: list[dict] = field(default_factory=list)


@dataclass(frozen=True)
class PackingLine:
    """Packing List 한 줄. allocator/reconciler엔 이 값 객체가 없어 여기서 정의한다.

    ★박스를 공유하는 라인은 qty_per_carton·carton_count·gross_weight_kg·measure·cbm이
    그 박스의 첫 행에만 원본값을 갖고 나머지는 `None`이다(병합셀을 그대로 보존한 것 —
    forward-fill하지 않는다). 배분값은 `distribute_box_metrics`가 **별도로** 만든다.
    """

    seq: int
    carton_range: str | None
    item_name: str
    quantity: Decimal
    qty_per_carton: Decimal | None
    carton_count: Decimal | None
    gross_weight_kg: Decimal | None
    measure: str | None
    cbm: Decimal | None
    remark: str | None


@dataclass(frozen=True)
class PackingParseResult:
    lines: list[PackingLine]
    total_qty: Decimal | None
    total_gross_weight_kg: Decimal | None
    total_cbm: Decimal | None
    total_cartons: Decimal | None


@dataclass(frozen=True)
class ExpenseParseResult:
    cost_lines: list[CostLine]
    hbl_no: str | None
    declaration_no: str | None
    declaration_date: date | None
    eta: date | None
    shipper_name: str | None
    vessel: str | None
    fx_rate: Decimal | None
    declared_inv_value: Decimal | None
    currency: str | None
    customs_value_krw: Decimal | None
    carton_count: int | None
    gross_weight_kg: Decimal | None
    cbm: Decimal | None


# ──────────────────────────────────────────────
# 공용 변환 헬퍼 — 못 읽으면 None이지 추정값을 만들지 않는다
# ──────────────────────────────────────────────
def _to_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        s = value.strip().replace(",", "")
        if not s:
            return None
        try:
            return Decimal(s)
        except InvalidOperation:
            return None
    return None


def _require_decimal(value: Any, ctx: str) -> Decimal:
    d = _to_decimal(value)
    if d is None:
        raise CustomsDocParseError(f"{ctx}: 숫자를 못 읽음(원본값={value!r})")
    return d


def _to_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _excel_date(value: Any) -> date | None:
    """엑셀 날짜 셀 → date. datetime(엑셀이 서식을 인식해 이미 변환한 경우)과
    순수 시리얼 숫자(변환기를 안 거친 xlsx) 둘 다 받는다 — 추정하지 않고 둘 다 실측 형태다."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        serial = float(value)
    except (TypeError, ValueError):
        return None
    return (_EXCEL_EPOCH + timedelta(days=serial)).date()


class _Cell:
    """openpyxl 셀의 `.value` 하나만 흉내낸다."""

    __slots__ = ("value",)

    def __init__(self, value: Any) -> None:
        self.value = value


class _XlsSheet:
    """xlrd 시트를 openpyxl의 `ws.cell(row=, column=).value`처럼 보이게 하는 어댑터.

    파서 본문이 쓰는 워크시트 API가 **그 하나뿐**이라(실측: `grep "ws\\."` 전건) 어댑터가
    이 정도면 충분하다. 파싱 로직을 포맷별로 두 벌 만들지 않는 것이 목적이다 —
    사본 두 벌은 한쪽만 고쳐지는 형태다.

    ★행·열은 openpyxl과 같은 **1-기반**이고 xlrd는 0-기반이라 여기서 흡수한다.
    ★범위를 벗어난 셀은 예외가 아니라 `None`이다 — openpyxl이 그렇게 동작하고,
      파서가 「빈 행이면 계속」으로 총계 행을 찾아가므로 예외를 던지면 그 탐색이 깨진다.
    """

    def __init__(self, sheet: Any) -> None:
        self._s = sheet

    def cell(self, row: int, column: int) -> _Cell:
        if row < 1 or column < 1 or row > self._s.nrows or column > self._s.ncols:
            return _Cell(None)
        v = self._s.cell_value(row - 1, column - 1)
        # xlrd는 빈 셀을 ''로 준다. openpyxl은 None이라 그쪽에 맞춘다 —
        # 파서가 `is None`으로 빈 행을 판정하기 때문이다.
        return _Cell(None if v == "" else v)


def _resolve_item_col(ws, header_row: int, guess: int, qty_col: int, probe: int = 12) -> int:
    """품목명 열을 **데이터로** 확정한다. 헤더 라벨이 한 칸 어긋나는 양식이 실재한다.

    규칙: 헤더 행 아래 `probe`행을 훑어, `order`~`qty` 사이 열 중 «숫자가 아닌 글자»가
    가장 많이 들어 있는 열을 고른다. 동점이면 헤더가 가리킨 열을 유지한다(추측을 최소화).
    후보가 하나도 없으면 헤더 값을 그대로 쓴다 — 여기서 실패로 만들지 않고,
    뒤의 「품명 자리에 숫자」 가드와 자기검산이 잡게 둔다.
    """
    lo, hi = min(guess, qty_col - 1), qty_col - 1
    if hi < 1:
        return guess
    scores: dict[int, int] = {}
    for c in range(max(lo, 1), hi + 1):
        n = 0
        for r in range(header_row + 1, header_row + 1 + probe):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = " ".join(str(v).split())
            if not s or _NUMERIC_ONLY_RE.match(s.replace(",", "")):
                continue
            if "total" in s.casefold():
                break
            n += 1
        scores[c] = n
    best = max(scores.values(), default=0)
    if best == 0:
        return guess
    if scores.get(guess, 0) == best:
        return guess
    return max((c for c, n in scores.items() if n == best))


def _pick_sheet(names: list[str], want: str) -> str:
    """시트명을 «정확히»가 아니라 «느슨하게» 고른다.

    실측(2026-08-22, 13개월 전수): 같은 공급사 파일에 `CI` 말고 **`CI (2)`**가 있었다
    (엑셀에서 시트를 복제하면 붙는 접미사다). 정확 일치만 보면 그 선적건 하나가 통째로 안 읽힌다.
    ★그래도 «아무거나»는 아니다 — 정확 일치 → 접두 일치 순으로 보고, 후보가 여럿이면
      가장 짧은 이름(원본일 가능성이 높다)을 고른다.
    """
    if want in names:
        return want
    key = want.casefold()
    cands = [n for n in names if _compact(n).casefold().startswith(key)]
    if not cands:
        raise CustomsDocParseError(f"시트 '{want}'이 없다 — 워크북 시트: {names}")
    return min(cands, key=len)


def _load_sheet(data: bytes, sheet_name: str):
    """.xlsx면 openpyxl, .xls(BIFF8)면 xlrd로 연다.

    ★`xlrd`는 **선택 의존성**이다(`requirements.txt`에 있지만 prod에 아직 설치 안 됐을 수 있다 —
    `safe_deploy.sh`에 pip 단계가 없다). 없으면 «조용히 실패»하지 않고 **무엇을 하면 되는지**를
    말한다: 이 앱의 정본은 사람이 확인한 폼 데이터라, 파서가 못 읽어도 수기 입력 경로가 살아 있다.
    """
    head = data[:8]
    is_xls = head.startswith(b"\xd0\xcf\x11\xe0")  # OLE2 복합문서 매직

    if is_xls:
        try:
            import xlrd  # noqa: PLC0415 — 선택 의존성이라 함수 안에서 임포트한다
        except ImportError as exc:
            raise CustomsDocParseError(
                ".xls(구형 엑셀) 파일인데 서버에 `xlrd`가 설치돼 있지 않다. "
                "엑셀에서 '다른 이름으로 저장 → .xlsx'로 변환해 올리거나, 수기로 입력하면 된다."
            ) from exc
        try:
            wb = xlrd.open_workbook(file_contents=data)
        except Exception as exc:
            raise CustomsDocParseError(f".xls를 열지 못했다({exc}).") from exc
        picked = _pick_sheet(wb.sheet_names(), sheet_name)
        return _XlsSheet(wb.sheet_by_name(picked))

    try:
        wb = load_workbook(BytesIO(data), data_only=True)
    except Exception as exc:  # openpyxl이 던지는 예외 종류가 다양해 포괄한다
        raise CustomsDocParseError(
            f"유효한 엑셀 파일이 아니다({exc}). .xlsx 또는 .xls를 올려야 한다."
        ) from exc
    return wb[_pick_sheet(wb.sheetnames, sheet_name)]


def _header_text(ws, row: int, col: int) -> str:
    v = ws.cell(row=row, column=col).value
    return " ".join(str(v).split()).casefold() if v is not None else ""


def _find_header_row(ws, keyword: str, max_row: int = 30, max_col: int = 20) -> tuple[int, int]:
    """헤더 셀을 **행·열 둘 다** 찾는다. 열을 고정하지 않는다.

    ★초판은 헤더를 «2열 고정»으로 찾고 데이터 열도 D·E·F로 하드코딩했다. 그런데 같은 공급사가
    보내는 서류인데도 양식마다 **열이 한 칸씩 밀린다**(실측 2026-08-22: 8/18 건은 D=품명·E=수량,
    `SO-WSOH-114` 건은 C=품명·D=수량). 그 결과 파서가 **예외 없이** 수량을 품명으로, 단가를 수량으로
    읽었다 — 값이 «있는데 틀린» 최악의 실패다. 그래서 위치가 아니라 **글자**로 찾는다.
    """
    key = keyword.casefold()
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if key in _header_text(ws, r, c):
                return r, c
    raise CustomsDocParseError(
        f"헤더 '{keyword}'를 {max_row}행 × {max_col}열 안에서 못 찾음 — 양식이 예상과 다르다."
    )


def _map_columns(ws, header_row: int, spec: dict[str, list[str]], max_col: int = 24) -> dict[str, int]:
    """헤더 행을 훑어 «논리 이름 → 열 번호» 지도를 만든다.

    `spec`의 각 항목은 후보 키워드 목록이고 **먼저 오는 것이 우선**이다(구체적인 것부터).
    한 열이 두 논리 이름에 걸리지 않게 이미 배정된 열은 건너뛴다.
    못 찾은 이름은 지도에 안 담긴다 — 호출부가 필수/선택을 판단한다.
    """
    texts = {c: _header_text(ws, header_row, c) for c in range(1, max_col + 1)}
    out: dict[str, int] = {}
    taken: set[int] = set()
    for name, keywords in spec.items():
        for kw in keywords:
            hit = next(
                (c for c, t in sorted(texts.items()) if c not in taken and kw.casefold() in t),
                None,
            )
            if hit is not None:
                out[name] = hit
                taken.add(hit)
                break
    return out


def extract_pdf_text(data: bytes) -> str:
    """통관경비서 PDF에서 텍스트를 뽑는다. 실패하면 «수기 경로»를 안내한다.

    ★`pypdf`도 선택 의존성이다(위 `xlrd`와 같은 이유).
    ★**스캔 PDF는 못 읽는다** — 텍스트 레이어가 없으면 빈 문자열이 나오고, 그때는
      「글자가 없다」고 말한다. OCR을 붙이지 않는다(계약 범위 밖이고, 틀린 숫자를 지어낼 위험이 크다).
    """
    try:
        import pypdf  # noqa: PLC0415 — 선택 의존성
    except ImportError as exc:
        raise CustomsDocParseError(
            "서버에 `pypdf`가 설치돼 있지 않아 PDF를 읽을 수 없다. "
            "경비서 내용을 수기로 입력하거나 텍스트를 붙여넣으면 된다."
        ) from exc
    try:
        reader = pypdf.PdfReader(BytesIO(data))
        text = "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception as exc:
        raise CustomsDocParseError(f"PDF를 읽지 못했다({exc}).") from exc
    if not text.strip():
        raise CustomsDocParseError(
            "PDF에서 글자를 찾지 못했다 — 스캔 이미지 PDF로 보인다. "
            "이 경우 수기 입력이나 텍스트 붙여넣기를 써야 한다(OCR은 지원하지 않는다)."
        )
    return text


# ──────────────────────────────────────────────
# §1 Commercial Invoice
# ──────────────────────────────────────────────
def parse_commercial_invoice(data: bytes) -> InvoiceParseResult:
    """CI 시트를 파싱한다. 헤더 행('Order No.')을 찾은 뒤 'Total'이 나올 때까지 읽는다.

    B열(Order No.)은 병합셀이라 데이터 라인마다 forward-fill한다. G열(Total)은
    quantity×unit_price 검산에만 쓰고, 어긋나도 값을 고치지 않고 리포트에 담는다.
    """
    ws = _load_sheet(data, "CI")

    header_row, order_col = _find_header_row(ws, "order no")
    cols = _map_columns(
        ws,
        header_row,
        {
            # 구체적인 것부터 — 'total'이 'total(cny)'와 'total qty'에 둘 다 걸리기 때문이다.
            "order": ["order no"],
            "item": ["item", "description", "goods"],
            "qty": ["quantity", "qty", "pcs"],
            "price": ["unit price", "price"],
            "total": ["total"],
        },
    )
    cols.setdefault("order", order_col)
    missing = [k for k in ("item", "qty", "price") if k not in cols]
    if missing:
        raise CustomsDocParseError(
            f"CI 헤더에서 필수 열을 못 찾음: {missing} "
            f"(찾은 것: { {k: v for k, v in cols.items()} }). 양식이 예상과 다르다."
        )
    # ★헤더 라벨은 «힌트»이고 판정은 **데이터**가 한다.
    #   실측(2026-08-22, 13개월 전수): 어떤 파일은 `Item` 라벨이 C열인데 실제 품목명은 D열이고,
    #   C열엔 분류어(`screen protector`)가 첫 행에만 있었다. 같은 공급사의 다른 파일은 라벨이 D였다.
    #   ⇒ 라벨 위치를 그대로 믿으면 그 선적건이 「품명이 비었는데 수량은 있다」로 통째로 죽는다.
    cols["item"] = _resolve_item_col(ws, header_row, cols["item"], cols["qty"])

    # 인보이스 번호·날짜도 라벨로 찾는다(열 고정 금지 — 위 헤더와 같은 이유).
    invoice_no = invoice_date = None
    try:
        r_no, c_no = _find_header_row(ws, "invoice no", max_row=header_row)
        invoice_no = _to_str(ws.cell(row=r_no, column=c_no + 1).value)
    except CustomsDocParseError:
        pass
    try:
        r_dt, c_dt = _find_header_row(ws, "invoice date", max_row=header_row)
        invoice_date = _excel_date(ws.cell(row=r_dt, column=c_dt + 1).value)
    except CustomsDocParseError:
        pass

    lines: list[InvoiceLine] = []
    order_nos: list[str | None] = []
    mismatches: list[dict] = []
    declared_total: Decimal | None = None
    declared_qty: Decimal | None = None

    seq = 0
    current_order: str | None = None
    max_row = header_row + 500
    row = header_row + 1
    while row <= max_row:
        raw_order = ws.cell(row=row, column=cols["order"]).value
        if isinstance(raw_order, str) and "total" in raw_order.strip().lower():
            declared_qty = _to_decimal(ws.cell(row=row, column=cols["qty"]).value)
            if "total" in cols:
                declared_total = _to_decimal(ws.cell(row=row, column=cols["total"]).value)
            # 총계 행은 「Total QTY: | 수량 | Total(RMB): | 금액」 꼴이라 금액이 total열이 아니라
            # 그 오른쪽에 올 수 있다. 못 읽었으면 헤더 행 오른쪽까지 훑어 숫자를 찾는다.
            if declared_total is None:
                for c in range(cols["qty"] + 1, cols["qty"] + 6):
                    cand = _to_decimal(ws.cell(row=row, column=c).value)
                    if cand is not None and (declared_qty is None or cand != declared_qty):
                        declared_total = cand
                        break
            break
        item_name = _to_str(ws.cell(row=row, column=cols["item"]).value)
        qty_raw = ws.cell(row=row, column=cols["qty"]).value
        if item_name is None and qty_raw is None:
            row += 1
            continue  # 완전 공백 행 — 총계 행을 아직 못 만났으니 계속 본다
        if item_name is None:
            raise CustomsDocParseError(f"CI {row}행: 품명이 비었는데 수량은 있다.")
        # ★품명이 숫자면 열 지도가 틀린 것이다 — 그 상태로 진행하면 수량을 품명으로 싣는다.
        if _NUMERIC_ONLY_RE.match(item_name.replace(",", "")):
            raise CustomsDocParseError(
                f"CI {row}행: 품명 자리에 숫자('{item_name}')가 있다 — 열 지도가 어긋났다. "
                f"현재 지도: {cols}. 수기 입력으로 진행해야 한다."
            )
        if isinstance(raw_order, str) and raw_order.strip():
            current_order = raw_order.strip()
        seq += 1
        qty = _require_decimal(qty_raw, f"CI {row}행 수량")
        price = _require_decimal(ws.cell(row=row, column=cols["price"]).value, f"CI {row}행 단가")
        lines.append(InvoiceLine(seq=seq, item_name=item_name, quantity=qty, unit_price_foreign=price))
        order_nos.append(current_order)
        declared_line_total = (
            _to_decimal(ws.cell(row=row, column=cols["total"]).value) if "total" in cols else None
        )
        if declared_line_total is not None:
            computed = qty * price
            if abs(computed - declared_line_total) > Decimal("0.01"):
                mismatches.append(
                    {"seq": seq, "item_name": item_name, "computed": str(computed), "declared": str(declared_line_total)}
                )
        row += 1
    else:
        raise CustomsDocParseError(f"CI 시트: {header_row + 1}~{max_row}행 안에서 총계 행('Total')을 못 찾음")

    # ★자기검산 — 뽑은 라인의 합이 서류의 총계와 같아야 한다.
    #   통관경비서에 넣은 것과 같은 장치다(그쪽은 이게 실제로 잘못된 파싱을 잡았다).
    #   CI엔 없어서 열이 밀린 서류를 **예외 없이 잘못 읽었다**(2026-08-22 SO-WSOH-114 실측).
    if declared_qty is not None:
        got_qty = sum((ln.quantity for ln in lines), _ZERO)
        if got_qty != declared_qty:
            raise CustomsDocParseError(
                f"CI 자기검산 불일치 — 라인 수량 합 {got_qty} ≠ 총계 행 {declared_qty}. "
                "라인을 빠뜨렸거나 열 지도가 어긋났다. 수기로 확인해야 한다."
            )
    if declared_total is not None:
        got_total = sum((ln.quantity * ln.unit_price_foreign for ln in lines), _ZERO)
        if abs(got_total - declared_total) > Decimal("0.01"):
            raise CustomsDocParseError(
                f"CI 자기검산 불일치 — 라인 금액 합 {got_total} ≠ 총계 행 {declared_total}. "
                "단가·수량 열이 어긋났을 수 있다. 수기로 확인해야 한다."
            )

    return InvoiceParseResult(
        lines=lines,
        order_nos=order_nos,
        invoice_no=invoice_no,
        invoice_date=invoice_date,
        declared_total=declared_total,
        declared_qty=declared_qty,
        line_total_mismatches=mismatches,
    )


# ──────────────────────────────────────────────
# §2 Packing List
# ──────────────────────────────────────────────
def parse_packing_list(data: bytes) -> PackingParseResult:
    """PL 시트를 파싱한다. 헤더 행('Ctn No.')을 찾은 뒤 'TOTAL'이 나올 때까지 읽는다.

    A열(박스 번호)·E~J열(박스 단위 수치)은 박스 공유 시 첫 행에만 값이 있다 — 여기서는
    원본 그대로(forward-fill 없이) 보존하고, 배분은 `distribute_box_metrics`가 한다.
    """
    ws = _load_sheet(data, "PL")

    header_row, ctn_col = _find_header_row(ws, "ctn no")
    cols = _map_columns(
        ws,
        header_row,
        {
            # 순서가 중요하다 — 'total g.w'가 'g.w'보다, 'ctn qty'가 'quantity'보다 먼저다.
            "ctn": ["ctn no"],
            "item": ["description", "goods", "item"],
            "ship_qty": ["shipment qty", "shipment"],
            "per_carton": ["quantity", "每箱数量", "sets"],
            "cartons": ["ctn qty", "箱数"],
            "total_gw": ["total\ng.w", "total g.w", "毛重"],
            "gw": ["g.w"],
            "measure": ["measure", "箱规"],
            "cbm": ["total volume", "volume", "cbm"],
            "remark": ["remark", "备注"],
        },
    )
    cols.setdefault("ctn", ctn_col)
    missing = [k for k in ("item", "ship_qty") if k not in cols]
    if missing:
        raise CustomsDocParseError(
            f"PL 헤더에서 필수 열을 못 찾음: {missing} (찾은 것: {cols}). 양식이 예상과 다르다."
        )

    def _c(name: str, row: int):
        col = cols.get(name)
        return ws.cell(row=row, column=col).value if col else None

    lines: list[PackingLine] = []
    total_qty = total_gw = total_cbm = total_cartons = None

    seq = 0
    max_row = header_row + 500
    row = header_row + 1
    while row <= max_row:
        raw_item = _c("item", row)
        if isinstance(raw_item, str) and "total" in raw_item.strip().lower():
            total_qty = _to_decimal(_c("ship_qty", row))
            total_cartons = _to_decimal(_c("cartons", row))
            total_gw = _to_decimal(_c("total_gw", row))
            total_cbm = _to_decimal(_c("cbm", row))
            break
        item_name = _to_str(raw_item)
        qty_raw = _c("ship_qty", row)
        if item_name is None and qty_raw is None:
            row += 1
            continue
        if item_name is None:
            raise CustomsDocParseError(f"PL {row}행: 품명이 비었는데 수량은 있다.")
        if _NUMERIC_ONLY_RE.match(item_name.replace(",", "")):
            raise CustomsDocParseError(
                f"PL {row}행: 품명 자리에 숫자('{item_name}')가 있다 — 열 지도가 어긋났다. "
                f"현재 지도: {cols}."
            )
        seq += 1
        qty = _require_decimal(qty_raw, f"PL {row}행 Shipment Qty")
        lines.append(
            PackingLine(
                seq=seq,
                carton_range=_to_str(_c("ctn", row)),
                item_name=item_name,
                quantity=qty,
                qty_per_carton=_to_decimal(_c("per_carton", row)),
                carton_count=_to_decimal(_c("cartons", row)),
                gross_weight_kg=_to_decimal(_c("total_gw", row)),  # 박스 총중량
                measure=_to_str(_c("measure", row)),
                cbm=_to_decimal(_c("cbm", row)),  # 박스 총부피
                remark=_to_str(_c("remark", row)),
            )
        )
        row += 1
    else:
        raise CustomsDocParseError(f"PL 시트: {header_row + 1}~{max_row}행 안에서 총계 행('TOTAL')을 못 찾음")

    # ★자기검산 — CI와 같은 장치.
    if total_qty is not None:
        got = sum((ln.quantity for ln in lines), _ZERO)
        if got != total_qty:
            raise CustomsDocParseError(
                f"PL 자기검산 불일치 — 라인 수량 합 {got} ≠ 총계 행 {total_qty}. "
                "라인을 빠뜨렸거나 열 지도가 어긋났다."
            )

    return PackingParseResult(
        lines=lines,
        total_qty=total_qty,
        total_gross_weight_kg=total_gw,
        total_cbm=total_cbm,
        total_cartons=total_cartons,
    )


def distribute_box_metrics(lines: list[PackingLine]) -> list[PackingLine]:
    """박스 단위로만 있는 총중량·총부피를 그룹 내 라인에 **수량 비례**로 나눈다.

    그룹 시작 판정 = `gross_weight_kg`가 값을 가진 행(그 박스의 첫 행). `carton_range`로
    가르지 않는 이유는 continuation 행의 carton_range도 원본에서 비어 있어, 그걸로는
    첫 행 자체를 못 찾기 때문이다. 원본 리스트는 건드리지 않고 새 리스트를 반환한다
    (원본 보존과 파생값을 안 섞기 위해서 — 이 함수를 별도로 둔 이유).
    """
    if not lines:
        return []
    groups: list[list[int]] = []
    for i, ln in enumerate(lines):
        if ln.gross_weight_kg is not None or not groups:
            groups.append([i])
        else:
            groups[-1].append(i)

    out = list(lines)
    for idx_group in groups:
        head = lines[idx_group[0]]
        total_qty = sum((lines[i].quantity for i in idx_group), _ZERO)
        if total_qty <= _ZERO:
            continue  # 수량 0이면 비율을 못 만든다 — 원본(대개 전부 None) 그대로 둔다
        for i in idx_group:
            share = lines[i].quantity / total_qty
            new_weight = head.gross_weight_kg * share if head.gross_weight_kg is not None else None
            new_cbm = head.cbm * share if head.cbm is not None else None
            out[i] = replace(lines[i], gross_weight_kg=new_weight, cbm=new_cbm)
    return out


# ──────────────────────────────────────────────
# §4 통관경비서 (이미 추출된 텍스트)
# ──────────────────────────────────────────────
_COST_LINE_RE = re.compile(r"^(.+?)\s+([\d,]+(?:\.\d+)?)\s*/\s*([\d,]+(?:\.\d+)?)\s*$")


def _grab(text: str, pattern: str, flags: int = 0) -> str | None:
    m = re.search(pattern, text, flags)
    return m.group(1).strip() if m else None


def _grab_date(text: str, pattern: str) -> date | None:
    s = _grab(text, pattern)
    if s is None:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def _grab_decimal(text: str, pattern: str) -> Decimal | None:
    s = _grab(text, pattern)
    return _to_decimal(s) if s is not None else None


_NUM_RE = re.compile(r"^-?[\d,]+(?:\.\d+)?$")
# 품명 자리에 이게 걸리면 열 지도가 어긋난 것이다(글자가 하나도 없는 순수 숫자).
_NUMERIC_ONLY_RE = re.compile(r"^-?\d+(?:\.\d+)?$")
_SUBTOTAL_RE = re.compile(r"subtotal|소계|합계", re.IGNORECASE)


def _compact(s: str) -> str:
    return re.sub(r"\s+", "", s)


def _near(tokens: list[str], label: str, offset: int, join_back: int = 1) -> str | None:
    """라벨 줄을 찾아 그 **앞/뒤 offset**번째 토큰을 준다.

    PDF 표는 추출기가 셀 단위로 뱉어서 「값 → 라벨」 순서가 되는 경우가 흔하다
    (실측: pypdf가 이 서류를 `209.88 / 환율` 순으로 낸다). 정규식 한 벌로는 두 순서를 다 못 잡아
    라벨 기준 상대 위치로 집는다.
    `join_back>1`이면 여러 줄로 쪼개진 값(예: 회사명 2줄)을 공백으로 이어 붙인다.
    """
    key = _compact(label)
    for i, t in enumerate(tokens):
        if _compact(t).startswith(key):
            parts = []
            for k in range(join_back):
                j = i + offset - (k if offset < 0 else -k)
                if 0 <= j < len(tokens):
                    parts.append(tokens[j])
            parts.reverse() if offset < 0 else None
            val = " ".join(p for p in parts if p).strip()
            return val or None
    return None


def _as_date(s: str | None) -> date | None:
    if not s:
        return None
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    return date(int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else None


def _extract_cost_rows(
    tokens: list[str], text: str
) -> tuple[list[CostLine], tuple[Decimal, Decimal] | None]:
    """비용 표를 뽑는다. **행마다 «이름 1 + 숫자 2»**라는 성질만 쓴다.

    왜 순서를 안 믿는가: 같은 PDF 안에서도 추출 순서가 뒤섞인다 — 실측(2026-08-22, pypdf 6.16.1)에서
    앞 7행은 `금액 → 이름 → 세액`이고 관세·부가세·통관수수료 3행은 `이름 → 금액 → 세액`이었다.
    그래서 「1 TEXT + 2 NUM이 모이면 한 행」으로만 묶고, 숫자는 **나온 순서대로** (공급가액, 세액)로 읽는다.

    소계·합계 행은 데이터가 아니라 **검산 기준**으로 따로 뽑아 돌려준다.
    """
    # 시작 지점: 신고번호 값 뒤부터가 비용 표다. 앞쪽엔 환율·과세금액 같은 헤더 숫자가 섞여 있어
    # 그대로 묶으면 가짜 행이 생긴다. 앵커를 못 찾으면 전체를 훑되 헤더 라벨은 이름 후보에서 뺀다.
    start = 0
    for i, t in enumerate(tokens):
        if _compact(t).startswith("신고번호"):
            start = i + 1
            break

    _HEADER_LABELS = (
        "환율", "신고금액", "과세금액", "면허일자", "신고일자", "신고번호", "비고", "예상세액",
        "예상공급가액", "내역", "송금계좌번호", "예금주", "처리일자", "가입금액", "잔액",
    )

    rows: list[CostLine] = []
    subtotal: tuple[Decimal, Decimal] | None = None
    name: str | None = None
    nums: list[Decimal] = []

    def _flush() -> None:
        nonlocal name, nums, subtotal
        if name is None or len(nums) < 2:
            name, nums = None, []
            return
        supply, tax = nums[0], nums[1]
        if _SUBTOTAL_RE.search(_compact(name)):
            # 「소 계」가 전체 소계다. 「B/L Sub Total」은 부분 소계라 검산 기준이 아니다.
            if "소계" in _compact(name):
                subtotal = (supply, tax)
        else:
            rows.append(
                CostLine(
                    item_name=name,
                    supply_amount=supply,
                    tax_amount=tax,
                    # ★이름 매칭 하나에 의존한다(docstring 참조). 「부가세」로 시작하면 배부 제외.
                    is_costing=not _compact(name).startswith("부가세"),
                    # 「관세」로 시작하면 귀속 대상(D-CPP-50). 「통관수수료」는 관세가 아니다 —
                    # startswith라 걸리지 않는다(‘통관…’으로 시작하므로).
                    is_duty=_compact(name).startswith("관세"),
                )
            )
        name, nums = None, []

    for t in tokens[start:]:
        c = _compact(t)
        if _NUM_RE.match(c):
            nums.append(_to_decimal(t) or _ZERO)
        else:
            if any(c.startswith(h) for h in _HEADER_LABELS):
                continue  # 헤더 라벨은 비용 항목이 아니다
            if name is not None:
                _flush()  # 이름이 연달아 나오면 앞 행은 숫자가 모자란 것 — 버린다
            name = t
        if name is not None and len(nums) >= 2:
            _flush()
            if subtotal is not None:
                # ★「소 계」를 만나면 표는 끝이다. 그 뒤는 송금계좌·잔액 구역인데, 거기에도
                #   「이름 + 숫자 2」 꼴이 있어(실측: `기업은행 : 476-… / -1,184,350 / 1,184,350`)
                #   계속 읽으면 가짜 비용 행이 생긴다.
                break
    _flush()

    return rows, subtotal


def parse_customs_expense(text: str) -> ExpenseParseResult:
    """통관경비서에서 **이미 추출된 텍스트**를 파싱한다. PDF 추출 자체는 라우터가 한다.

    ★`is_costing` 판정은 「품명이 '부가세'로 시작하는가」 **하나뿐**이다. 이름 매칭 하나에
    의존한다는 걸 여기 명시한다 — 서류 양식이 바뀌면(예: '부가세액'으로 표기) 조용히 깨진다.
    그래서 정본은 이 파서가 아니라 사람이 확인한 폼 데이터다(계약 전문).
    ★소계/합계(`Sub Total`·`소계`·`합계`, 공백이 끼워진 'B/L Sub Total' 류 포함)는 반드시
    제외한다 — 넣으면 통관비가 이중 계상된다.
    """
    tokens = [t.strip() for t in text.splitlines() if t.strip()]

    # ★후보를 여러 경로로 만들고 **모양이 맞는 것**을 고른다.
    #   정규식 하나만 믿으면 안 되는 이유(실측): 이 PDF는 라벨과 값이 줄바꿈으로 갈리고 순서도
    #   뒤섞여서, `환율\s+([\d.]+)`가 「환율\n3,434」에 걸려 **3**을 잡았다. 값이 «있는데 틀린 것»이
    #   «없는 것»보다 나쁘다 — 그래서 형태 검증을 통과한 후보만 채택한다.
    def _pick(validator, *candidates):
        for c in candidates:
            if c is None:
                continue
            v = validator(c)
            if v is not None:
                return v
        return None

    def _v_code(s: str) -> str | None:
        s = s.strip()
        return s if re.fullmatch(r"[A-Za-z][A-Za-z0-9]{5,}", s) else None

    def _v_decl_no(s: str) -> str | None:
        s = s.strip()
        return s if re.fullmatch(r"\d{3,}-\d{2}-\w+", s) else None

    def _v_pos_dec(s: str):
        d = _to_decimal(s)
        return d if d is not None and d > _ZERO else None

    def _v_text(s: str) -> str | None:
        s = " ".join(s.split())
        if not s or _NUM_RE.match(_compact(s)) or re.search(r"\d{4}-\d{2}-\d{2}", s):
            return None
        if len(_compact(s)) < 4 or not re.search(r"[A-Za-z가-힣]", s):
            return None
        # 라벨이 값 자리로 새어 들어온 경우를 거른다
        if any(_compact(s).startswith(h) for h in ("비고", "예상", "내역", "담당자", "수신")):
            return None
        return s

    hbl_no = _pick(_v_code, _near(tokens, "HBL NO", -1), _grab(text, r"HBL\s*NO\s*:?\s*(\S+)"))
    declaration_no = _pick(
        _v_decl_no, _near(tokens, "신고번호", -1), _grab(text, r"신고번호\s+(\S+)")
    )
    declaration_date = _as_date(
        _near(tokens, "신고일자", -1) or _grab(text, r"신고일자\s+(\d{4}-\d{2}-\d{2})")
    )
    eta = _as_date(_near(tokens, "ETA", -1) or _grab(text, r"ETA\s+(\d{4}-\d{2}-\d{2})"))
    shipper_name = _pick(
        _v_text,
        _near(tokens, "Shipper NM", -1, join_back=2),
        _grab(text, r"Shipper\s*NM\s+(.+?)\s*(?:·|\n|$)"),
    )
    vessel = _pick(
        _v_text,
        _near(tokens, "선명/항차", -1, join_back=2),
        _grab(text, r"선명\s*/\s*항차\s+(.+?)\s*(?:·|\n|$)"),
    )
    fx_rate = _pick(_v_pos_dec, _near(tokens, "환율", -1), _grab(text, r"환율\s+([\d,.]+)"))
    customs_value_krw = _pick(
        _v_pos_dec,
        _near(tokens, "과세금액", -1),
        _grab(text, r"과세금액\s*\(\s*₩\s*\)\s*([\d,]+)"),
    )
    gross_weight_kg = _pick(
        _v_pos_dec, _near(tokens, "Gross W/T", -1), _grab(text, r"Gross\s+W\s*/\s*T\s+([\d.]+)")
    )

    carton_m = re.search(r"(\d+)\s*CARTONS\s*/\s*([\d.]+)", text, re.IGNORECASE)
    carton_count = int(carton_m.group(1)) if carton_m else None
    cbm = _to_decimal(carton_m.group(2)) if carton_m else None

    inv_m = re.search(r"INV\s*Value\s+([A-Z]{3})\s+([\d,.]+)", text)
    if inv_m:
        currency = inv_m.group(1)
        declared_inv_value = _to_decimal(inv_m.group(2))
    else:
        # pypdf가 표를 셀 단위로 뱉으면 순서가 「23,100.00 / CNY / INV Value」로 뒤집힌다.
        currency = _near(tokens, "INV Value", -1)
        declared_inv_value = _to_decimal(_near(tokens, "INV Value", -2))
        if currency and not re.fullmatch(r"[A-Z]{3}", currency):
            currency = None

    cost_lines, subtotal = _extract_cost_rows(tokens, text)

    if not cost_lines:
        raise CustomsDocParseError(
            "통관경비서 텍스트에서 비용 라인을 하나도 못 찾음 — "
            "'항목 금액 / 세액' 꼴의 줄이 있는지, 소계·합계만 있던 건 아닌지 확인해야 한다."
        )

    # ★자기검산 — 뽑은 라인의 합이 서류의 «소계»와 같아야 한다.
    #   이게 없으면 표 순서가 조금만 달라져도 **틀린 금액이 조용히** 폼에 채워진다.
    #   맞으면 「행을 빠뜨리지도 중복하지도 않았다」가 증명되고, 틀리면 사람에게 넘긴다.
    if subtotal is not None:
        got_supply = sum((c.supply_amount for c in cost_lines), _ZERO)
        got_tax = sum((c.tax_amount for c in cost_lines), _ZERO)
        want_supply, want_tax = subtotal
        if got_supply != want_supply or got_tax != want_tax:
            raise CustomsDocParseError(
                "통관경비서 자기검산 불일치 — 뽑은 비용 라인의 합이 서류의 소계와 다르다. "
                f"공급가액 {got_supply} vs 소계 {want_supply} · "
                f"세액 {got_tax} vs 소계 {want_tax}. "
                "표 구조가 예상과 달라 라인을 빠뜨렸거나 중복했을 수 있다 — 수기로 확인해야 한다."
            )

    return ExpenseParseResult(
        cost_lines=cost_lines,
        hbl_no=hbl_no,
        declaration_no=declaration_no,
        declaration_date=declaration_date,
        eta=eta,
        shipper_name=shipper_name,
        vessel=vessel,
        fx_rate=fx_rate,
        declared_inv_value=declared_inv_value,
        currency=currency,
        customs_value_krw=customs_value_krw,
        carton_count=carton_count,
        gross_weight_kg=gross_weight_kg,
        cbm=cbm,
    )
