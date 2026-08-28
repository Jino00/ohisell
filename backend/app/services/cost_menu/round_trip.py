"""왕복 표 → 엑셀 파일 (계약 D-CPP-62 S3 「DB-생성 다운로드」).

목표 원문(계약 §1): *"엑셀에서 고치든 화면에서 고치든 같은 단가 원장에 「Jino 확인 = 승인」으로
쌓이고, **화면의 표와 다운로드 파일이 같은 모양**이라 「어느 게 최신인가」라는 질문 자체가
사라진다."*

이 모듈이 지키는 것은 그 문장의 **「같은 모양」** 쪽이다. 세 가지를 구조로 못 박는다:

① **열은 한 곳에서 온다** — `round_trip_columns.json`을 **런타임에 읽는다.** 픽스처가 아니라
   생성기의 입력이므로, 낡으면 파일이 낡게 나오고 프론트 파리티 테스트가 그 자리에서 빨개진다.
   (프론트 `costHome.ts:ROUND_TRIP_COLUMNS`는 화면을 그리고, vitest가 이 JSON을 `fs`로 직접
   읽어 key·label·editable·**순서**를 전건 대조한다.)

② **값은 화면과 같은 페이로드에서 온다** — `materials.list_materials(db)` 산출을 그대로 쓴다.
   ORM을 다시 조회하면 `price_rule.choose_price`·`resolve_inc_vat`의 **두 번째 사본**이 생기고,
   그게 이 저장소가 반복해 밟은 병이다(D-CPP-60 · 직렬화기 두 벌 `standard_cost.py`↔`recipes.py`).

③ ★**언제나 전건이다 — 화면 필터를 절대 받지 않는다.** 필터된 파일을 재업로드하면 빠진 종이
   전부 S4 「사라짐」 묶음에 서고, 확인 클릭 **한 번이 백여 종을 비활성화**한다. 이 함수들은
   필터 인자를 아예 갖지 않는다 — 안 받으면 실수로 넘길 수도 없다.
"""

from __future__ import annotations

import hashlib
import json
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from sqlalchemy.orm import Session

from app.models import CostRoundTripSnapshot
from app.utils.kst import kst_now

from . import materials as M

#: 열 스펙 정본. **이 파일을 읽는 것이 곧 스펙을 쓰는 것**이다(모듈 상수로 굳히지 않는다 —
#: import 시점에 굳으면 파일을 고쳐도 프로세스가 옛 스펙을 계속 쓴다).
_SPEC_PATH = Path(__file__).with_name("round_trip_columns.json")

#: 데이터 시트 이름. S4 파서가 이 이름으로 찾는다.
SHEET_DATA = "원가 왕복"
#: 스냅샷 메타 시트 (설계 Q3 확정: *"데이터 시트를 오염시키지 않도록 별도 시트 `_meta`를 예약"*).
SHEET_META = "_meta"


class RoundTripError(Exception):
    """스펙 자체가 깨졌을 때. **조용히 넘어가지 않는다** — 열이 빠진 파일이 나가면
    그 파일의 재업로드가 S4에서 통째로 오독된다."""


# ──────────────────────────────────────────────
# 열 스펙
# ──────────────────────────────────────────────
def load_columns() -> list[dict]:
    """열 스펙 12개를 **읽어서** 낸다.

    ★반환값에 `_why*` 주석 키는 싣지 않는다 — 스냅샷 `column_spec`에 그대로 굳는 값이라
    산문이 섞이면 스냅샷끼리의 비교가 산문 수정에도 흔들린다.
    """
    raw = json.loads(_SPEC_PATH.read_text(encoding="utf-8"))
    cols = raw.get("columns")
    if not isinstance(cols, list) or not cols:
        raise RoundTripError(f"열 스펙이 비었다: {_SPEC_PATH}")

    out: list[dict] = []
    seen: set[str] = set()
    for c in cols:
        key = c.get("key")
        label = c.get("label")
        if not key or not label:
            raise RoundTripError(f"열 스펙에 key/label이 없다: {c!r}")
        if key in seen:
            raise RoundTripError(f"열 key가 겹친다: {key}")
        seen.add(key)
        col = {"key": key, "label": label, "editable": bool(c.get("editable"))}
        if c.get("file_label"):
            col["file_label"] = c["file_label"]
        if c.get("value_labels"):
            col["value_labels"] = dict(c["value_labels"])
        out.append(col)
    return out


def header_text(col: dict) -> str:
    """파일 헤더 한 칸.

    ★읽기 전용 열은 `«읽기전용»` 접미를 단다(설계 Q3). **시트 잠금은 쓰지 않는다** —
    엑셀 잠금은 사용자가 풀 수 있어 방어가 아니고, 진짜 방어는 S4 diff가 읽기 전용 열의
    변경을 「반영 불가」 묶음으로 **세워서 보여 주는 것**이다(조용한 드롭은 자동 병합의 사촌).
    """
    base = col.get("file_label") or col["label"]
    return base if col["editable"] else f"{base} «읽기전용»"


# ──────────────────────────────────────────────
# 행 값 — 정준화
# ──────────────────────────────────────────────
def _label_of(col: dict, raw) -> str | None:
    """enum 값 → 화면과 **같은 낱말**. 매핑에 없는 값은 지어내지 않고 원값 그대로 낸다.

    ★불리언은 JSON 표기(`true`/`false`)로 찾는다. 파이썬 `str(True)`는 `"True"`라서
    JSON 스펙의 `"true"` 키와 안 맞는다 — 그러면 파일에 `×1.1` 대신 `True`가 찍히고,
    사람은 그게 무슨 뜻인지 모른다(테스트가 이걸 잡았다).
    """
    if raw is None:
        return None
    labels = col.get("value_labels") or {}
    if raw is True or raw is False:
        return labels.get("true" if raw else "false", str(raw))
    return labels.get(str(raw), str(raw))


def canonical_row(columns: list[dict], m: dict) -> dict:
    """부자재 payload 1건 → 열 key별 **정준화값**.

    ★정준화의 목적은 S4의 3-방향 대조에서 `1600` vs `1600.00`이 「변경 1건」으로 서지 않게
    하는 것이다. 그래서 돈은 **문자열로 고정 2자리**, 날짜는 ISO, **없음은 `None`**이다.
    ★`None`은 0이 아니다(계약 §2-7 · 금지선 「없음과 0 동일 표시 금지」). 파일에서는 빈 칸으로
    내려가고, 사람이 거기 `0`을 적어 올리면 그건 **「0원으로 바꾸겠다」는 주장**이라 S4가
    변경 1건으로 세운다 — 빈 칸과 0은 다른 값이다.
    """
    inc_derived = bool(m.get("latest_price_inc_derived"))
    values = {
        "id": m.get("id"),
        "name": m.get("name"),
        "form_factor": m.get("form_factor"),
        "part": m.get("part"),
        "unit": m.get("unit"),
        "price_ex": _money(m.get("latest_price_ex_vat")),
        "price_inc": _money(m.get("latest_price_inc_vat")),
        "vat_derived": True if inc_derived else None,
        "price_source": m.get("latest_price_source"),
        "effective_date": m.get("latest_price_effective_date"),
        "excel_ref": _money(m.get("excel_ref_price")),
        # ★**비고만** 싣는다 — 화면의 이 칸은 「상태 라벨 + 라이브 배지 + 비고」의 합성이라
        #   왕복이 원리적으로 불가능하다. 배지까지 쓰면 S4가 파생값 변화를 «사람의 수정»으로
        #   오독하고, 「승인 · 비고」로 합치면 그 문자열을 다시 파싱해야 한다.
        #   잃어버린 「상태」는 파일 헤더가 자백한다(`file_label`).
        "status_note": m.get("note"),
    }

    row: dict = {}
    for col in columns:
        key = col["key"]
        if key not in values:
            raise RoundTripError(f"열 `{key}`에 대응하는 값 규칙이 없다 — 스펙만 늘었다")
        raw = values[key]
        row[key] = _label_of(col, raw) if col.get("value_labels") else raw
    return row


def _money(v) -> str | None:
    """돈 문자열을 **소수 2자리로 고정**. `None`은 `None`으로 남긴다(0으로 접지 않는다)."""
    if v is None:
        return None
    try:
        return f"{Decimal(str(v)):.2f}"
    except (InvalidOperation, ValueError):
        # 숫자가 아닌 값이 오면 지어내지 않고 원문을 남긴다 — 조용히 0으로 만들지 않는다.
        return str(v)


def canonical_rows(columns: list[dict], mats: list[dict]) -> list[dict]:
    return [canonical_row(columns, m) for m in mats]


# ──────────────────────────────────────────────
# 스냅샷
# ──────────────────────────────────────────────
def snapshot_code(snap: CostRoundTripSnapshot) -> str:
    """파일에 찍히는 스냅샷 ID. S4가 이 값으로 되찾는다."""
    return f"CRT-{snap.id}"


def _content_hash(spec_json: str, rows_json: str) -> str:
    return hashlib.sha256(f"{spec_json}\n{rows_json}".encode("utf-8")).hexdigest()


def build_snapshot(db: Session) -> CostRoundTripSnapshot:
    """지금 이 순간의 표를 스냅샷으로 굳힌다. **전건이다 — 필터 인자가 없다.**

    ★직전 스냅샷과 내용이 같으면 **새 행을 만들지 않고 그 스냅샷을 그대로 재발급**한다.
    다운로드가 사실상 멱등이 되고, 스냅샷 증가가 「상태가 실제로 바뀐 횟수」에 묶이며,
    같은 상태를 두 번 받아 둘 다 올려도 S4에서 가짜 충돌이 안 난다.
    """
    columns = load_columns()
    mats = M.list_materials(db)  # ★화면과 **같은** 페이로드. 재조회하지 않는다.
    rows = canonical_rows(columns, mats)

    spec_json = json.dumps(columns, ensure_ascii=False)
    rows_json = json.dumps(rows, ensure_ascii=False)
    digest = _content_hash(spec_json, rows_json)

    latest = (
        db.query(CostRoundTripSnapshot)
        .order_by(CostRoundTripSnapshot.id.desc())
        .first()
    )
    if latest is not None and latest.content_hash == digest:
        return latest

    snap = CostRoundTripSnapshot(
        # ★KST를 앱에서 명시 세팅한다 — `server_default=func.now()`는 이 저장소에서 UTC고,
        #   파일에 찍히는 시각이 9시간 어긋나면 사람이 「어느 게 최신인가」를 바로 그 자리에서
        #   다시 묻게 된다. 그 질문을 없애는 것이 이 계약이다.
        created_at=kst_now(),
        column_spec=spec_json,
        rows=rows_json,
        row_count=len(rows),
        content_hash=digest,
    )
    db.add(snap)
    db.flush()
    return snap


# ──────────────────────────────────────────────
# 파일
# ──────────────────────────────────────────────
#: 숫자 셀로 쓰는 열 — 문자열로 쓰면 사람이 엑셀에서 계산·정렬을 못 한다.
_NUMERIC_KEYS = {"price_ex", "price_inc", "excel_ref"}


def build_workbook(snap: CostRoundTripSnapshot) -> BytesIO:
    """스냅샷 → `.xlsx` 바이트.

    ★파일은 **스냅샷에서만** 만든다 — DB를 다시 읽지 않는다. 그래야 「받은 파일」과
    「S4가 대조할 스냅샷」이 원리적으로 같은 것이 된다(둘이 각자 조회하면 그 사이의 변경이
    조용히 끼어들어, 무수정 재업로드가 「변경 N건」으로 서는 유령 diff가 난다).
    """
    columns = json.loads(snap.column_spec)
    rows = json.loads(snap.rows)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_DATA

    ws.append([header_text(c) for c in columns])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=False)
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([_cell_value(c, row.get(c["key"])) for c in columns])

    meta = wb.create_sheet(SHEET_META)
    meta.append(["항목", "값"])
    meta.append(["스냅샷 ID", snapshot_code(snap)])
    meta.append(["생성 시각 (KST)", snap.created_at.strftime("%Y-%m-%d %H:%M:%S")])
    meta.append(["행 수", snap.row_count])
    meta.append(
        [
            "안내",
            "이 시트를 지우거나 스냅샷 ID를 고치면 업로드가 「스냅샷 불명」이 되어 "
            "무엇이 바뀌었는지 대조하지 못한다. 값은 「원가 왕복」 시트에서만 고친다.",
        ]
    )
    for cell in meta[1]:
        cell.font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _cell_value(col: dict, v):
    """정준화값 → 엑셀 셀 값. `None`은 **빈 칸**이지 `0`도 `"—"`도 아니다."""
    if v is None:
        return None
    if col["key"] in _NUMERIC_KEYS:
        try:
            return float(v)
        except (TypeError, ValueError):
            return v
    return v


def filename(snap: CostRoundTripSnapshot) -> str:
    """`원가_왕복_CRT-12_20260828_1405.xlsx` — 파일명은 **편의 사본**이다.

    정본은 `_meta` 시트다(계약 §4 「스냅샷 ID가 **파일 안에** 보인다」). 파일명은 사람이
    다른 이름으로 저장하는 순간 사라지므로 그것만 믿는 경로를 만들지 않는다.
    """
    return f"원가_왕복_{snapshot_code(snap)}_{snap.created_at.strftime('%Y%m%d_%H%M')}.xlsx"
