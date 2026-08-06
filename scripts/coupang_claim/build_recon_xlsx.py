"""발주 → 발송 → 계산서 대조표 (xlsx).

한 줄 = (발주번호, SKU). 다섯 축을 한 줄에 놓아 어디서 끊겼는지 눈으로 보이게 한다:
  ①발주  ②우리 납품가능 회신  ③실제 발송  ④계산서에 적힌 수량  ⑤미정산(=③−④)
시트1 발주별 요약(124행) / 시트2 SKU 라인 상세 / 시트3 읽는 법
"""
import csv, io, json, collections, pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SC = pathlib.Path(__file__).parent
rd = lambda f, **kw: list(csv.DictReader(io.open(SC / f, encoding=kw.pop("enc", "utf-8"))))

items = rd("recon_po_items.csv")
shipped = {(r["po"], r["sku"]): r for r in rd("recon_shipped.csv")}

# 계산서 라인 → (PO, SKU) 집계
inv_qty, inv_no, inv_pay = collections.Counter(), collections.defaultdict(set), collections.defaultdict(set)
for l in rd("settle_lines.csv"):
    if l["구분"] != "발주":
        continue
    k = (l["발주번호"], l["SKU번호"])
    inv_qty[k] += int(l["수량"].replace(",", ""))
    inv_no[k].add(l["계산서번호"]); inv_pay[k].add(l["지급일"])

# 청구 대상 라인(증거등급 보유)
claim = {(r["발주번호"], r["상품번호"]): r
         for r in rd("claim_v3.csv", enc="utf-8-sig")}

rows = []
for it in items:
    k = (it["po"], it["sku"])
    ordq = int(it["ord"])
    conf = int(it["conf"]) if it["conf"] else None
    sh = shipped.get(k)
    shq = int(sh["shipped"]) if sh else 0
    invq = inv_qty.get(k, 0)
    price = int(it["price"])
    unsettled = shq - invq
    c = claim.get(k)
    if unsettled > 0:
        verdict = "① 미정산 — 발송기록 있음(확인 요청 대상)"
    elif unsettled < 0:
        verdict = "④ 발송기록보다 많이 정산됨"
    elif conf is not None and ordq > conf:
        verdict = "② 미납품 — 당사가 납품불가 회신"
    else:
        verdict = "③ 정상 — 발송 전량 정산"
    rows.append({
        "발주일": it["po_date"], "발주번호": it["po"], "물류센터": it["center"],
        "상품번호": it["sku"], "상품명": it["sku_name"],
        "①발주": ordq, "②납품가능회신": conf if conf is not None else "",
        "③발송": shq, "④계산서수량": invq, "⑤미정산(③−④)": unsettled,
        "매입단가": price, "미정산금액": unsettled * price if unsettled > 0 else 0,
        "계산서번호": ";".join(sorted(inv_no.get(k, []))),
        "지급일": ";".join(sorted(inv_pay.get(k, []))),
        "쉽먼트번호": sh["ships"] if sh else "", "송장번호": sh["waybills"] if sh else "",
        "추적(집하)": c["집하"] if c else "", "추적(하차)": c["하차"] if c else "",
        "증거등급": c["증거등급"] if c else "",
        "판정": verdict,
    })

# ── 발주별 요약 ──
by_po = collections.OrderedDict()
for r in rows:
    p = r["발주번호"]
    s = by_po.setdefault(p, {"발주일": r["발주일"], "발주번호": p, "물류센터": r["물류센터"],
                             "SKU수": 0, "①발주": 0, "②납품가능회신": 0, "③발송": 0,
                             "④계산서수량": 0, "⑤미정산": 0, "미정산금액": 0,
                             "계산서번호": set(), "미정산SKU수": 0})
    s["SKU수"] += 1
    s["①발주"] += r["①발주"]
    s["②납품가능회신"] += r["②납품가능회신"] or 0
    s["③발송"] += r["③발송"]; s["④계산서수량"] += r["④계산서수량"]
    if r["⑤미정산(③−④)"] > 0:
        s["⑤미정산"] += r["⑤미정산(③−④)"]; s["미정산금액"] += r["미정산금액"]; s["미정산SKU수"] += 1
    s["계산서번호"] |= set(filter(None, r["계산서번호"].split(";")))
for s in by_po.values():
    s["계산서번호"] = ";".join(sorted(s["계산서번호"])) or "(발행 없음)"

# ── 엑셀 ──
HDR = PatternFill("solid", fgColor="2F3B52")
FILL = {"①": PatternFill("solid", fgColor="FFF3CD"), "②": PatternFill("solid", fgColor="EEEEEE"),
        "③": PatternFill("solid", fgColor="FFFFFF"), "④": PatternFill("solid", fgColor="FFD9D9")}
THIN = Border(*[Side(style="thin", color="CCCCCC")] * 4)
wb = Workbook(); wb.remove(wb.active)


def sheet(name, data, widths, num_cols, freeze="A2"):
    ws = wb.create_sheet(name)
    cols = list(data[0].keys())
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(1, c)
        cell.fill = HDR; cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in data:
        ws.append([r[c] for c in cols])
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        v = data[i - 2].get("판정", "")
        f = FILL.get(v[:1]) if v else None
        for cell in row:
            cell.border = THIN; cell.font = Font(size=9.5)
            if f:
                cell.fill = f
            if cell.column_letter in num_cols:
                cell.number_format = "#,##0"; cell.alignment = Alignment(horizontal="right")
    for col, wdt in widths.items():
        ws.column_dimensions[col].width = wdt
    ws.freeze_panes = freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(data)+1}"
    ws.row_dimensions[1].height = 30
    return ws


sheet("발주별 요약", list(by_po.values()),
      {"A": 11, "B": 12, "C": 10, "D": 7, "E": 8, "F": 13, "G": 8, "H": 12, "I": 9,
       "J": 12, "K": 22, "L": 11},
      set("EFGHIJ"))
sheet("SKU 라인 상세", rows,
      {"A": 11, "B": 12, "C": 10, "D": 11, "E": 46, "F": 8, "G": 13, "H": 8, "I": 12,
       "J": 13, "K": 10, "L": 12, "M": 20, "N": 20, "O": 14, "P": 26, "Q": 17, "R": 17,
       "S": 14, "T": 30},
      set("FGHIJKL"))

ws = wb.create_sheet("읽는 법")
guide = [
    ("이 표가 답하는 질문", ""),
    ("", "발주 하나하나에 대해 — 우리가 무엇을 보냈고, 그중 무엇이 계산서에 실렸고, 무엇이 안 실렸는가."),
    ("", ""),
    ("다섯 개의 수량 축", ""),
    ("① 발주", "쿠팡이 발주한 수량 (supplyhub 발주상세)"),
    ("② 납품가능회신", "그중 당사가 '보낼 수 있다'고 회신한 수량. ①−②는 우리가 못 보낸다고 한 것이라 청구 대상이 아님"),
    ("③ 발송", "당사가 실제로 발송한 수량 (쉽먼트 기록, 송장번호 있음)"),
    ("④ 계산서수량", "쿠팡 계산서의 「입고상세내역」에 적힌 수량. = 쿠팡이 입고 처리하고 대금을 지급한 수량"),
    ("⑤ 미정산", "③ − ④. 보냈는데 계산서에 실리지 않은 수량"),
    ("", ""),
    ("판정 색", ""),
    ("① 노랑", "미정산 — 발송기록 있음. 쿠팡에 확인 요청할 대상"),
    ("② 회색", "미납품 — 당사가 납품불가로 회신. 정상, 청구 대상 아님"),
    ("③ 흰색", "정상 — 발송한 만큼 전액 계산서에 실림"),
    ("④ 분홍", "발송기록보다 많이 정산됨 — 확인 필요(반대 방향 오차)"),
    ("", ""),
    ("주의", ""),
    ("", "「발주리스트」의 계산서 「상태=정산완료」는 발행된 그 계산서의 지급이 끝났다는 뜻이며,"),
    ("", "발주 전량이 정산되었다는 뜻이 아닙니다. 미정산 수량은 애초에 계산서에 실리지 않으므로"),
    ("", "계산서 상태로는 확인되지 않습니다. ④ 열이 그 계산서에 실제로 적힌 수량입니다."),
    ("", ""),
    ("대상", f"입고가 완결된 발주(상태 「거래명세서확인」 계열) 중 결손이 있는 {len(by_po)}건, SKU 라인 {len(rows)}건"),
    ("기준일", "2026-08-06 (계산서 입고상세내역 전수 조회)"),
]
for k, v in guide:
    ws.append([k, v])
for row in ws.iter_rows():
    row[0].font = Font(bold=True, size=10)
    row[1].font = Font(size=10)
    row[1].alignment = Alignment(wrap_text=True, vertical="top")
ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 100

out = SC / "04_발주_발송_계산서_대조표_20260806.xlsx"
wb.save(out)

v = collections.Counter(r["판정"] for r in rows)
print(f"저장: {out.name}")
print(f"  발주 {len(by_po)}건 · SKU 라인 {len(rows)}건")
for k in sorted(v):
    amt = sum(r["미정산금액"] for r in rows if r["판정"] == k)
    print(f"  {k}: {v[k]}라인" + (f" · {amt:,}원" if amt else ""))
